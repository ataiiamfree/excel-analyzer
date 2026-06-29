#!/usr/bin/env python3
"""Batch evaluation runner for the ChatExcel agent.

The runner feeds existing Excel test files and questions into the real
orchestrator. It intentionally bypasses Chainlit, but it does not bypass the
agent pipeline: ingest, preprocess, profile, plan, code generation, sandbox
execution, checking, and report generation all stay in the loop.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import io
import json
import logging
import math
import os
import re
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


LOGGER = logging.getLogger("eval")

from scripts.answer_compare import compare_answers, extract_answer_text
from scripts.workbook_compare import compare_workbooks


@dataclass(frozen=True)
class EvalCase:
    id: str
    file_path: Path
    question: str
    source: str
    notes: list[str] = field(default_factory=list)
    assertions: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "file": str(self.file_path),
            "question": self.question,
            "source": self.source,
            "notes": self.notes,
            "assertions": self.assertions,
        }


@dataclass
class ExecutionSnapshot:
    state: dict[str, Any]
    report: str
    output_files: list[Path] = field(default_factory=list)
    step_outputs: list[str] = field(default_factory=list)
    scripts: dict[str, str] = field(default_factory=dict)
    exception: str | None = None


@dataclass
class AssertionOutcome:
    name: str
    passed: bool
    detail: str = ""
    required: bool = True

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "passed": self.passed,
            "detail": self.detail,
            "required": self.required,
        }


@dataclass(frozen=True)
class TableOutput:
    path: Path
    sheet_name: str
    rows: list[list[Any]]


def _load_dotenv() -> None:
    try:
        from dotenv import load_dotenv
    except Exception:
        return
    load_dotenv(PROJECT_ROOT / ".env")


def _safe_id(value: str) -> str:
    text = value.strip().lower()
    text = re.sub(r"\s+", "-", text)
    text = re.sub(r"[^a-z0-9_.-]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "case"


def _source_prefix(manifest_path: Path) -> str:
    if manifest_path.name == "manifest.json":
        return "complex"
    if manifest_path.suffix.lower() == ".md":
        return "simple"
    return _safe_id(manifest_path.stem)


def load_cases(manifest_path: str | Path) -> list[EvalCase]:
    """Load cases from a JSON manifest or the simple Markdown checklist."""
    path = Path(manifest_path)
    if not path.is_absolute():
        cwd_path = (Path.cwd() / path).resolve()
        path = cwd_path if cwd_path.exists() else (PROJECT_ROOT / path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"Manifest not found: {path}")

    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json_cases(path)
    if suffix in {".md", ".markdown"}:
        return _load_markdown_cases(path)
    raise ValueError(f"Unsupported manifest type: {path}")


def _load_json_cases(path: Path) -> list[EvalCase]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    items = raw.get("cases", raw) if isinstance(raw, dict) else raw
    if not isinstance(items, list):
        raise ValueError(f"JSON manifest must be a list or contain cases: {path}")

    prefix = _safe_id(raw.get("id", "")) if isinstance(raw, dict) else ""
    prefix = prefix or _source_prefix(path)
    cases: list[EvalCase] = []
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"Case #{index} is not an object in {path}")

        raw_file = item.get("file") or item.get("file_path")
        if not raw_file:
            raise ValueError(f"Case #{index} missing file in {path}")
        file_path = _resolve_case_file(path.parent, str(raw_file))

        base_id = _safe_id(str(item.get("id") or f"{prefix}-{index:02d}"))
        notes = [str(value) for value in item.get("tests", item.get("notes", []))]
        assertions = dict(item.get("assertions", {}))

        if "questions" in item:
            questions = item.get("questions") or []
            for q_index, question_item in enumerate(questions, start=1):
                question, q_assertions, q_notes = _question_parts(question_item)
                merged_assertions = {**assertions, **q_assertions}
                cases.append(EvalCase(
                    id=f"{base_id}-q{q_index:02d}",
                    file_path=file_path,
                    question=question,
                    source=str(path),
                    notes=notes + q_notes,
                    assertions=merged_assertions,
                ))
        else:
            question = str(item.get("question") or "").strip()
            if not question:
                raise ValueError(f"Case #{index} missing question in {path}")
            cases.append(EvalCase(
                id=base_id,
                file_path=file_path,
                question=question,
                source=str(path),
                notes=notes,
                assertions=assertions,
            ))
    return cases


def _question_parts(question_item: Any) -> tuple[str, dict[str, Any], list[str]]:
    if isinstance(question_item, str):
        return question_item.strip(), {}, []
    if isinstance(question_item, dict):
        question = str(question_item.get("question") or "").strip()
        assertions = dict(question_item.get("assertions", {}))
        notes = [str(value) for value in question_item.get("tests", question_item.get("notes", []))]
        if not question:
            raise ValueError("Question object is missing question")
        return question, assertions, notes
    raise ValueError(f"Unsupported question item: {question_item!r}")


def _load_markdown_cases(path: Path) -> list[EvalCase]:
    """Parse docs/test_datasets/simple checklist tables into cases."""
    cases: list[EvalCase] = []
    current_file: str | None = None
    current_capability = ""

    for line in path.read_text(encoding="utf-8").splitlines():
        heading = re.match(r"^##\s+(.+?\.xlsx)", line.strip())
        if heading:
            current_file = heading.group(1).strip()
            current_capability = ""
            continue

        if line.startswith("**") and "考察" in line:
            current_capability = line.strip("*").strip()
            continue

        if not current_file or not line.strip().startswith("|"):
            continue
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if len(cells) < 3 or not cells[0].isdigit():
            continue

        dataset_no_match = re.match(r"(\d+)_", current_file)
        dataset_no = dataset_no_match.group(1) if dataset_no_match else f"{len(cases) + 1:02d}"
        question_no = int(cells[0])
        question = cells[1]
        note = cells[2]
        file_path = path.parent / "simple" / current_file
        if not file_path.exists():
            file_path = path.parent / current_file

        notes = [value for value in (current_capability, note) if value]
        cases.append(EvalCase(
            id=f"simple-{dataset_no}-q{question_no:02d}",
            file_path=file_path.resolve(),
            question=question,
            source=str(path),
            notes=notes,
        ))
    return cases


def _resolve_case_file(base_dir: Path, raw_file: str) -> Path:
    path = Path(raw_file)
    if path.is_absolute():
        return path
    return (base_dir / path).resolve()


def select_cases(
    cases: list[EvalCase],
    case_ids: list[str] | None = None,
    limit: int | None = None,
) -> list[EvalCase]:
    selected = cases
    if case_ids:
        wanted = set(case_ids)
        selected = [case for case in selected if case.id in wanted]
        missing = sorted(wanted - {case.id for case in selected})
        if missing:
            raise ValueError(f"Unknown case id(s): {', '.join(missing)}")
    if limit is not None:
        selected = selected[:limit]
    return selected


def run_assertions(case: EvalCase, snapshot: ExecutionSnapshot) -> list[AssertionOutcome]:
    assertions = case.assertions or {}
    outcomes = [
        AssertionOutcome(
            name="no_unhandled_exception",
            passed=snapshot.exception is None,
            detail=(snapshot.exception or "")[:500],
        ),
        AssertionOutcome(
            name="workspace_completed",
            passed=snapshot.state.get("status") == "completed",
            detail=f"status={snapshot.state.get('status')!r}",
        ),
        AssertionOutcome(
            name="report_non_empty",
            passed=len((snapshot.report or "").strip()) >= 20,
            detail=f"chars={len(snapshot.report or '')}",
        ),
    ]

    required_exts = [str(ext).lower() for ext in assertions.get("required_output_exts", [])]
    for ext in required_exts:
        ext = ext if ext.startswith(".") else f".{ext}"
        matched = [path for path in snapshot.output_files if path.suffix.lower() == ext]
        outcomes.append(AssertionOutcome(
            name=f"required_output_ext:{ext}",
            passed=bool(matched),
            detail=", ".join(path.name for path in matched) or "not found",
        ))

    report_contains = [str(value) for value in assertions.get("report_contains", [])]
    for text in report_contains:
        outcomes.append(AssertionOutcome(
            name=f"report_contains:{text}",
            passed=text in snapshot.report,
            detail="found" if text in snapshot.report else "missing",
        ))

    outcomes.extend(_report_must_mention_assertions(assertions.get("report_must_mention"), snapshot.report))
    answer_text = "\n\n".join([snapshot.report, *snapshot.step_outputs])
    outcomes.extend(_expected_answer_assertions(
        assertions.get("expected_answer", assertions.get("expected_answers")),
        answer_text,
    ))
    outcomes.extend(_expected_value_assertions(assertions.get("expected_values"), snapshot.output_files))
    outcomes.extend(_expected_row_count_assertions(assertions.get("expected_row_count"), snapshot.output_files))
    outcomes.extend(_expected_row_assertions(
        assertions.get("expected_rows", assertions.get("expected_records")),
        snapshot.output_files,
    ))

    required_sheet_names = [str(value) for value in assertions.get("required_sheet_names", [])]
    required_columns = assertions.get("required_columns", {})
    if required_sheet_names or required_columns:
        outcomes.extend(_workbook_assertions(snapshot.output_files, required_sheet_names, required_columns))

    answer_workbook = assertions.get("answer_workbook", assertions.get("golden_workbook"))
    if answer_workbook:
        outcomes.extend(_answer_workbook_assertions(answer_workbook, snapshot.output_files, Path(case.source).parent))

    forbidden_patterns = [str(value) for value in assertions.get("forbidden_script_patterns", [])]
    for pattern in forbidden_patterns:
        regex = re.compile(pattern)
        matches = [
            name
            for name, text in snapshot.scripts.items()
            if regex.search(text)
        ]
        outcomes.append(AssertionOutcome(
            name=f"forbidden_script_pattern:{pattern}",
            passed=not matches,
            detail=", ".join(matches) if matches else "not found",
        ))

    non_report_outputs = [
        path for path in snapshot.output_files
        if path.name != "report.md" and path.suffix.lower() not in {".md", ".txt"}
    ]
    if not non_report_outputs:
        outcomes.append(AssertionOutcome(
            name="non_report_output_present",
            passed=False,
            detail="No downloadable data/chart artifact was produced.",
            required=False,
        ))
    return outcomes


def _report_must_mention_assertions(raw: Any, report: str) -> list[AssertionOutcome]:
    if not raw:
        return []

    specs = raw if isinstance(raw, list) else [raw]
    outcomes: list[AssertionOutcome] = []
    for index, spec in enumerate(specs, start=1):
        name = f"report_must_mention:{index}"
        required = True
        if isinstance(spec, str):
            passed = spec in report
            detail = "found" if passed else f"missing={spec!r}"
            outcomes.append(AssertionOutcome(name=name, passed=passed, detail=detail))
            continue

        if not isinstance(spec, dict):
            outcomes.append(AssertionOutcome(
                name=name,
                passed=False,
                detail=f"Unsupported report_must_mention spec: {spec!r}",
            ))
            continue

        required = bool(spec.get("required", True))
        terms = _string_list(spec.get("terms", spec.get("texts", spec.get("text"))))
        label = str(spec.get("name") or "terms")
        if not terms:
            outcomes.append(AssertionOutcome(
                name=f"report_must_mention:{label}",
                passed=False,
                detail="No text/terms configured.",
                required=required,
            ))
            continue

        within_chars = spec.get("within_chars")
        if within_chars is None or len(terms) <= 1:
            missing = [term for term in terms if term not in report]
            outcomes.append(AssertionOutcome(
                name=f"report_must_mention:{label}",
                passed=not missing,
                detail="found" if not missing else f"missing={missing}",
                required=required,
            ))
            continue

        try:
            window = max(0, int(within_chars))
        except (TypeError, ValueError):
            window = 80
        passed = _terms_appear_near(report, terms, window)
        outcomes.append(AssertionOutcome(
            name=f"report_must_mention:{label}",
            passed=passed,
            detail=f"terms={terms}; within_chars={window}",
            required=required,
        ))
    return outcomes


def _terms_appear_near(text: str, terms: list[str], window: int) -> bool:
    positions = [(term, match.start()) for term in terms for match in re.finditer(re.escape(term), text)]
    if not positions:
        return False
    for _, start in positions:
        left = max(0, start - window)
        right = min(len(text), start + window)
        snippet = text[left:right]
        if all(term in snippet for term in terms):
            return True
    return False


def _expected_answer_assertions(raw: Any, report: str) -> list[AssertionOutcome]:
    specs, errors = _normalize_expected_answer_specs(raw)
    if not specs and not errors:
        return []

    outcomes = [
        AssertionOutcome(name="expected_answer_config", passed=False, detail=error)
        for error in errors
    ]
    for index, spec in enumerate(specs, start=1):
        outcomes.append(_check_expected_answer(spec, report, index))
    return outcomes


def _normalize_expected_answer_specs(raw: Any) -> tuple[list[dict[str, Any]], list[str]]:
    if raw is None:
        return [], []
    if isinstance(raw, dict):
        if any(key in raw for key in ("value", "expected", "answer", "expected_any", "options", "answers")):
            return [dict(raw)], []
        specs = []
        for name, value in raw.items():
            specs.append({"name": str(name), "value": value})
        return specs, []
    if isinstance(raw, list):
        specs: list[dict[str, Any]] = []
        errors: list[str] = []
        for index, item in enumerate(raw, start=1):
            if isinstance(item, dict):
                specs.append(dict(item))
            elif isinstance(item, (str, int, float)):
                specs.append({"name": f"answer_{index}", "value": item})
            else:
                errors.append(f"expected_answer[{index}] must be a scalar or object, got {type(item).__name__}")
        return specs, errors
    if isinstance(raw, (str, int, float)):
        return [{"value": raw}], []
    return [], [f"expected_answer must be a scalar, list, or object, got {type(raw).__name__}"]


def _check_expected_answer(spec: dict[str, Any], report: str, index: int) -> AssertionOutcome:
    name = str(spec.get("name") or spec.get("label") or f"answer_{index}")
    required = bool(spec.get("required", True))
    expected_options = _expected_answer_options(spec)
    if not expected_options:
        return AssertionOutcome(
            name=f"expected_answer:{name}",
            passed=False,
            detail="No expected answer configured.",
            required=required,
        )

    observed, extraction_method = extract_answer_text(report, answer_regex=spec.get("answer_regex"))
    require_marked = bool(spec.get("require_marked_answer", spec.get("require_final_answer", False)))
    if require_marked and extraction_method == "full_report":
        return AssertionOutcome(
            name=f"expected_answer:{name}",
            passed=False,
            detail="No marked final answer found; expected a line like 'Final Answer: ...'.",
            required=required,
        )

    mode = str(spec.get("mode") or "auto")
    abs_tol = _float_option(spec, ("abs_tol", "tolerance", "tol"), default=1e-6)
    rel_tol = _float_option(spec, ("rel_tol",), default=0.0)
    min_score = _float_option(spec, ("min_score", "threshold"), default=0.75)
    min_token_recall = _float_option(spec, ("min_token_recall",), default=0.5)

    comparisons = [
        compare_answers(
            expected,
            observed,
            mode=mode,
            abs_tol=abs_tol,
            rel_tol=rel_tol,
            min_score=min_score,
            min_token_recall=min_token_recall,
        )
        for expected in expected_options
    ]
    best = max(comparisons, key=lambda item: (item.passed, item.score))
    detail = (
        f"method={extraction_method}; mode={best.mode}; score={best.score:.4f}; "
        f"expected={best.expected_text!r}; observed={best.observed_text[:300]!r}; {best.detail}"
    )
    return AssertionOutcome(
        name=f"expected_answer:{name}",
        passed=best.passed,
        detail=detail[:2000],
        required=required,
    )


def _expected_answer_options(spec: dict[str, Any]) -> list[Any]:
    for key in ("expected_any", "options", "answers"):
        if key not in spec:
            continue
        raw = spec.get(key)
        if isinstance(raw, (list, tuple, set)):
            return [value for value in raw if value is not None]
        if raw is not None:
            return [raw]
    for key in ("value", "expected", "answer"):
        if key in spec and spec.get(key) is not None:
            return [spec.get(key)]
    return []


def _expected_value_assertions(raw: Any, output_files: list[Path]) -> list[AssertionOutcome]:
    specs, errors = _normalize_expected_value_specs(raw)
    if not specs and not errors:
        return []

    tables, load_errors = _load_tabular_outputs(output_files)
    outcomes = [
        AssertionOutcome(
            name="tabular_output_load",
            passed=False,
            detail=error,
        )
        for error in load_errors
    ]
    for error in errors:
        outcomes.append(AssertionOutcome(
            name="expected_value_config",
            passed=False,
            detail=error,
        ))

    for index, spec in enumerate(specs, start=1):
        outcomes.append(_check_expected_value(spec, tables, index))
    return outcomes


def _normalize_expected_value_specs(raw: Any) -> tuple[list[dict[str, Any]], list[str]]:
    if not raw:
        return [], []
    if isinstance(raw, list):
        specs: list[dict[str, Any]] = []
        errors: list[str] = []
        for index, item in enumerate(raw, start=1):
            if isinstance(item, dict):
                specs.append(dict(item))
            else:
                errors.append(f"expected_values[{index}] must be an object, got {type(item).__name__}")
        return specs, errors
    if isinstance(raw, dict) and ("expected" in raw or "value" in raw):
        return [dict(raw)], []
    if isinstance(raw, dict):
        shared = {
            "abs_tol": raw.get("_abs_tol", raw.get("_tolerance", raw.get("_tol"))),
            "rel_tol": raw.get("_rel_tol"),
        }
        specs = []
        for name, expected in raw.items():
            if str(name).startswith("_"):
                continue
            spec = {"name": str(name), "expected": expected}
            spec.update({key: value for key, value in shared.items() if value is not None})
            specs.append(spec)
        return specs, []
    return [], [f"expected_values must be a list or object, got {type(raw).__name__}"]


def _check_expected_value(spec: dict[str, Any], tables: list[TableOutput], index: int) -> AssertionOutcome:
    name = str(spec.get("name") or spec.get("label") or f"value_{index}")
    required = bool(spec.get("required", True))
    expected_numbers = _expected_number_options(spec)
    if not expected_numbers:
        return AssertionOutcome(
            name=f"expected_value:{name}",
            passed=False,
            detail="Expected value is not numeric.",
            required=required,
        )

    matching_tables = _filter_tables(tables, spec)
    if not matching_tables:
        return AssertionOutcome(
            name=f"expected_value:{name}",
            passed=False,
            detail="No CSV/XLSX output table matched the assertion filters.",
            required=required,
        )

    abs_tol = _float_option(spec, ("abs_tol", "tolerance", "tol"), default=1e-9)
    rel_tol = _float_option(spec, ("rel_tol",), default=0.0)
    row_contains = spec.get("row_contains")
    column = spec.get("column")
    checked_rows = 0
    candidate_values: list[float] = []
    row_miss_tables: list[str] = []
    column_miss_tables: list[str] = []

    for table in matching_tables:
        column_index: int | None = None
        header_index: int | None = None
        if column:
            header_index, column_index = _find_column(table.rows, column, spec.get("header_row"))
            if column_index is None:
                column_miss_tables.append(_table_label(table))
                continue

        row_matched_in_table = False
        start_index = (header_index + 1) if header_index is not None and column_index is not None else 0
        for row_index, row in enumerate(table.rows[start_index:], start=start_index):
            if not _is_non_empty_row(row):
                continue
            if not _row_matches(row, row_contains):
                continue
            row_matched_in_table = True
            checked_rows += 1

            cells = [row[column_index]] if column_index is not None and column_index < len(row) else row
            numbers = [number for cell in cells for number in _numbers_from_value(cell)]
            candidate_values.extend(numbers[:10])
            for number in numbers:
                matched_expected = next(
                    (
                        expected_number
                        for expected_number in expected_numbers
                        if _numbers_close(number, expected_number, abs_tol=abs_tol, rel_tol=rel_tol)
                    ),
                    None,
                )
                if matched_expected is not None:
                    return AssertionOutcome(
                        name=f"expected_value:{name}",
                        passed=True,
                        detail=(
                            f"matched {number!r} in {_table_label(table)} row {row_index + 1}; "
                            f"expected={matched_expected!r}, abs_tol={abs_tol}, rel_tol={rel_tol}"
                        ),
                        required=required,
                    )
        if row_contains and not row_matched_in_table:
            row_miss_tables.append(_table_label(table))

    detail_parts = [
        f"expected={expected_numbers!r}",
        f"abs_tol={abs_tol}",
        f"rel_tol={rel_tol}",
        f"checked_rows={checked_rows}",
    ]
    if column_miss_tables:
        detail_parts.append(f"column_not_found={column_miss_tables[:5]}")
    if row_miss_tables:
        detail_parts.append(f"row_not_found={row_miss_tables[:5]}")
    if candidate_values:
        detail_parts.append(f"candidate_values={candidate_values[:20]}")
    return AssertionOutcome(
        name=f"expected_value:{name}",
        passed=False,
        detail="; ".join(detail_parts),
        required=required,
    )


def _expected_number_options(spec: dict[str, Any]) -> list[float]:
    if "expected_any" in spec:
        raw_values = spec.get("expected_any")
    elif "expected_options" in spec:
        raw_values = spec.get("expected_options")
    else:
        raw_values = [spec.get("expected", spec.get("value"))]

    if not isinstance(raw_values, (list, tuple, set)):
        raw_values = [raw_values]

    numbers: list[float] = []
    for raw_value in raw_values:
        number = _single_number(raw_value)
        if number is not None:
            numbers.append(number)
    return numbers


def _expected_row_count_assertions(raw: Any, output_files: list[Path]) -> list[AssertionOutcome]:
    specs, errors = _normalize_row_count_specs(raw)
    if not specs and not errors:
        return []

    tables, load_errors = _load_tabular_outputs(output_files)
    outcomes = [
        AssertionOutcome(
            name="tabular_output_load",
            passed=False,
            detail=error,
        )
        for error in load_errors
    ]
    for error in errors:
        outcomes.append(AssertionOutcome(name="expected_row_count_config", passed=False, detail=error))

    for index, spec in enumerate(specs, start=1):
        outcomes.append(_check_expected_row_count(spec, tables, index))
    return outcomes


def _normalize_row_count_specs(raw: Any) -> tuple[list[dict[str, Any]], list[str]]:
    if raw is None:
        return [], []
    if isinstance(raw, int):
        return [{"name": "any_table", "count": raw}], []
    if isinstance(raw, dict):
        return [dict(raw)], []
    if isinstance(raw, list):
        specs: list[dict[str, Any]] = []
        errors: list[str] = []
        for index, item in enumerate(raw, start=1):
            if isinstance(item, int):
                specs.append({"name": f"row_count_{index}", "count": item})
            elif isinstance(item, dict):
                specs.append(dict(item))
            else:
                errors.append(f"expected_row_count[{index}] must be an integer or object, got {type(item).__name__}")
        return specs, errors
    return [], [f"expected_row_count must be an integer, list, or object, got {type(raw).__name__}"]


def _check_expected_row_count(spec: dict[str, Any], tables: list[TableOutput], index: int) -> AssertionOutcome:
    name = str(spec.get("name") or f"row_count_{index}")
    required = bool(spec.get("required", True))
    expected = spec.get("count", spec.get("expected"))
    if not isinstance(expected, int):
        return AssertionOutcome(
            name=f"expected_row_count:{name}",
            passed=False,
            detail=f"Expected row count must be an integer: {expected!r}",
            required=required,
        )

    matching_tables = _filter_tables(tables, spec)
    if not matching_tables:
        return AssertionOutcome(
            name=f"expected_row_count:{name}",
            passed=False,
            detail="No CSV/XLSX output table matched the assertion filters.",
            required=required,
        )

    counts = []
    for table in matching_tables:
        count = _data_row_count(table.rows, spec)
        counts.append((_table_label(table), count))
        if count == expected:
            return AssertionOutcome(
                name=f"expected_row_count:{name}",
                passed=True,
                detail=f"matched {_table_label(table)} with {count} rows",
                required=required,
            )
    return AssertionOutcome(
        name=f"expected_row_count:{name}",
        passed=False,
        detail=f"expected={expected}; observed={counts}",
        required=required,
    )


def _expected_row_assertions(raw: Any, output_files: list[Path]) -> list[AssertionOutcome]:
    specs, errors = _normalize_expected_row_specs(raw)
    if not specs and not errors:
        return []

    tables, load_errors = _load_tabular_outputs(output_files)
    outcomes = [
        AssertionOutcome(
            name="tabular_output_load",
            passed=False,
            detail=error,
        )
        for error in load_errors
    ]
    for error in errors:
        outcomes.append(AssertionOutcome(name="expected_rows_config", passed=False, detail=error))

    for index, spec in enumerate(specs, start=1):
        outcomes.append(_check_expected_row(spec, tables, index))
    return outcomes


def _normalize_expected_row_specs(raw: Any) -> tuple[list[dict[str, Any]], list[str]]:
    if not raw:
        return [], []
    if isinstance(raw, dict):
        return [dict(raw)], []
    if isinstance(raw, list):
        specs: list[dict[str, Any]] = []
        errors: list[str] = []
        for index, item in enumerate(raw, start=1):
            if isinstance(item, dict):
                specs.append(dict(item))
            else:
                errors.append(f"expected_rows[{index}] must be an object, got {type(item).__name__}")
        return specs, errors
    return [], [f"expected_rows must be a list or object, got {type(raw).__name__}"]


def _check_expected_row(spec: dict[str, Any], tables: list[TableOutput], index: int) -> AssertionOutcome:
    name = str(spec.get("name") or f"row_{index}")
    required = bool(spec.get("required", True))
    contains = spec.get("row_contains", spec.get("contains"))
    if not contains:
        return AssertionOutcome(
            name=f"expected_row:{name}",
            passed=False,
            detail="expected_rows requires row_contains/contains.",
            required=required,
        )

    matching_tables = _filter_tables(tables, spec)
    for table in matching_tables:
        for row_index, row in enumerate(table.rows, start=1):
            if _is_non_empty_row(row) and _row_matches(row, contains):
                return AssertionOutcome(
                    name=f"expected_row:{name}",
                    passed=True,
                    detail=f"found in {_table_label(table)} row {row_index}",
                    required=required,
                )
    return AssertionOutcome(
        name=f"expected_row:{name}",
        passed=False,
        detail=f"row not found; contains={_string_list(contains)}",
        required=required,
    )


def _load_tabular_outputs(output_files: list[Path]) -> tuple[list[TableOutput], list[str]]:
    tables: list[TableOutput] = []
    errors: list[str] = []
    seen: set[Path] = set()
    for raw_path in output_files:
        path = Path(raw_path)
        suffix = path.suffix.lower()
        if suffix not in {".csv", ".tsv", ".xlsx", ".xlsm"}:
            continue
        if not path.exists():
            continue
        resolved = path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)

        if suffix in {".xlsx", ".xlsm"}:
            try:
                tables.extend(_read_workbook_tables(path))
            except Exception as exc:
                errors.append(f"{path.name}: {type(exc).__name__}: {exc}")
        else:
            try:
                tables.append(TableOutput(path=path, sheet_name=path.stem, rows=_read_delimited_table(path)))
            except Exception as exc:
                errors.append(f"{path.name}: {type(exc).__name__}: {exc}")
    return tables, errors


def _read_workbook_tables(path: Path) -> list[TableOutput]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        tables = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            tables.append(TableOutput(path=path, sheet_name=sheet_name, rows=rows))
        return tables
    finally:
        workbook.close()


def _read_delimited_table(path: Path) -> list[list[str]]:
    text = path.read_text(encoding="utf-8-sig")
    if not text:
        return []
    sample = text[:4096]
    if path.suffix.lower() == ".tsv":
        dialect = csv.excel_tab
    else:
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
    return [list(row) for row in csv.reader(io.StringIO(text), dialect)]


def _filter_tables(tables: list[TableOutput], spec: dict[str, Any]) -> list[TableOutput]:
    file_exts = _string_list(spec.get("file_ext", spec.get("file_exts")))
    file_exts = [ext.lower() if ext.startswith(".") else f".{ext.lower()}" for ext in file_exts]
    file_contains = _string_list(spec.get("file_contains", spec.get("filename_contains")))
    sheet = spec.get("sheet", spec.get("sheet_name"))
    sheet_contains = spec.get("sheet_contains")

    matched = []
    for table in tables:
        if file_exts and table.path.suffix.lower() not in file_exts:
            continue
        if file_contains and not all(token in table.path.name for token in file_contains):
            continue
        if sheet is not None and str(table.sheet_name) != str(sheet):
            continue
        if sheet_contains is not None and str(sheet_contains) not in str(table.sheet_name):
            continue
        matched.append(table)
    return matched


def _find_column(rows: list[list[Any]], column: Any, header_row: Any = None) -> tuple[int | None, int | None]:
    aliases = _string_list(column)
    if not aliases:
        return None, None

    if header_row is not None:
        try:
            row_indexes = [max(0, int(header_row) - 1)]
        except (TypeError, ValueError):
            row_indexes = []
    else:
        row_indexes = list(range(min(len(rows), 20)))

    for row_index in row_indexes:
        if row_index >= len(rows):
            continue
        row = rows[row_index]
        for col_index, cell in enumerate(row):
            text = _cell_text(cell)
            if any(_label_matches(text, alias) for alias in aliases):
                return row_index, col_index
    return None, None


def _label_matches(text: str, expected: str) -> bool:
    text_norm = _normal_text(text)
    expected_norm = _normal_text(expected)
    return bool(expected_norm) and (text_norm == expected_norm or expected_norm in text_norm)


def _data_row_count(rows: list[list[Any]], spec: dict[str, Any]) -> int:
    include_header = bool(spec.get("include_header", False))
    header_row = spec.get("header_row")
    if header_row is not None:
        try:
            start = max(0, int(header_row) - (1 if include_header else 0))
        except (TypeError, ValueError):
            start = 0
    elif include_header:
        start = 0
    else:
        start = _first_non_empty_row_index(rows)
        start = start + 1 if start is not None else 0

    contains = spec.get("row_contains")
    return sum(1 for row in rows[start:] if _is_non_empty_row(row) and _row_matches(row, contains))


def _first_non_empty_row_index(rows: list[list[Any]]) -> int | None:
    for index, row in enumerate(rows):
        if _is_non_empty_row(row):
            return index
    return None


def _row_matches(row: list[Any], contains: Any) -> bool:
    terms = _string_list(contains)
    if not terms:
        return True
    row_text = _normal_text("\t".join(_cell_text(cell) for cell in row))
    return all(_normal_text(term) in row_text for term in terms)


def _is_non_empty_row(row: list[Any]) -> bool:
    return any(_cell_text(cell) for cell in row)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normal_text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value).strip().casefold())


def _string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(item) for item in value if str(item)]
    return [str(value)]


def _single_number(value: Any) -> float | None:
    numbers = _numbers_from_value(value)
    return numbers[0] if numbers else None


def _numbers_from_value(value: Any) -> list[float]:
    if value is None or isinstance(value, bool):
        return []
    if isinstance(value, (int, float)):
        number = float(value)
        return [number] if math.isfinite(number) else []

    text = str(value).strip()
    if not text:
        return []
    if re.fullmatch(r"\d{4}[-/年]\d{1,2}([-月/]\d{1,2}日?)?", text):
        return []

    normalized = (
        text.replace("，", ",")
        .replace("％", "%")
        .replace("－", "-")
        .replace("−", "-")
        .replace("￥", "")
        .replace("¥", "")
    )
    normalized = re.sub(r"(?<=\d),(?=\d{3}(\D|$))", "", normalized)

    parenthesized_negative = bool(re.fullmatch(r"\(.*\)", normalized))
    numbers: list[float] = []
    for match in re.finditer(r"[-+]?\d+(?:\.\d+)?%?", normalized):
        raw = match.group()
        is_percent = raw.endswith("%")
        raw_number = raw[:-1] if is_percent else raw
        try:
            number = float(raw_number)
        except ValueError:
            continue
        if parenthesized_negative and number > 0:
            number = -number
        if math.isfinite(number):
            numbers.append(number)
            if is_percent:
                numbers.append(number / 100)
    return numbers


def _float_option(spec: dict[str, Any], names: tuple[str, ...], default: float) -> float:
    for name in names:
        value = spec.get(name)
        if value is None:
            continue
        number = _single_number(value)
        if number is not None:
            return float(number)
    return default


def _numbers_close(actual: float, expected: float, *, abs_tol: float, rel_tol: float) -> bool:
    tolerance = max(abs_tol, rel_tol * abs(expected))
    return abs(actual - expected) <= tolerance


def _table_label(table: TableOutput) -> str:
    return f"{table.path.name}:{table.sheet_name}"


def _workbook_assertions(
    output_files: list[Path],
    required_sheet_names: list[str],
    required_columns: Any,
) -> list[AssertionOutcome]:
    workbook_files = [path for path in output_files if path.suffix.lower() in {".xlsx", ".xlsm"}]
    if not workbook_files:
        return [
            AssertionOutcome(
                name="workbook_exists_for_schema_checks",
                passed=False,
                detail="No .xlsx/.xlsm output file found.",
            )
        ]

    try:
        import openpyxl
    except Exception as exc:
        return [
            AssertionOutcome(
                name="openpyxl_available",
                passed=False,
                detail=f"{type(exc).__name__}: {exc}",
            )
        ]

    loaded: list[tuple[Path, Any]] = []
    for workbook_path in workbook_files:
        try:
            loaded.append((workbook_path, openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)))
        except Exception as exc:
            return [
                AssertionOutcome(
                    name=f"workbook_loadable:{workbook_path.name}",
                    passed=False,
                    detail=f"{type(exc).__name__}: {exc}",
                )
            ]

    outcomes: list[AssertionOutcome] = []
    all_sheet_names = {sheet for _, wb in loaded for sheet in wb.sheetnames}
    for sheet_name in required_sheet_names:
        outcomes.append(AssertionOutcome(
            name=f"required_sheet:{sheet_name}",
            passed=sheet_name in all_sheet_names,
            detail="found" if sheet_name in all_sheet_names else f"available={sorted(all_sheet_names)}",
        ))

    if isinstance(required_columns, list):
        expected_by_sheet = {"*": [str(value) for value in required_columns]}
    elif isinstance(required_columns, dict):
        expected_by_sheet = {
            str(sheet): [str(value) for value in columns]
            for sheet, columns in required_columns.items()
        }
    else:
        expected_by_sheet = {}

    for sheet_name, columns in expected_by_sheet.items():
        present_columns = _columns_for_sheet(loaded, sheet_name)
        missing = [column for column in columns if column not in present_columns]
        outcomes.append(AssertionOutcome(
            name=f"required_columns:{sheet_name}",
            passed=not missing,
            detail="found" if not missing else f"missing={missing}; available={sorted(present_columns)}",
        ))

    for _, wb in loaded:
        wb.close()
    return outcomes


def _answer_workbook_assertions(raw: Any, output_files: list[Path], base_dir: Path) -> list[AssertionOutcome]:
    specs = raw if isinstance(raw, list) else [raw]
    outcomes: list[AssertionOutcome] = []
    candidates = _candidate_answer_workbooks(output_files)

    for index, item in enumerate(specs, start=1):
        if isinstance(item, str):
            spec = {"path": item}
        elif isinstance(item, dict):
            spec = dict(item)
        else:
            outcomes.append(AssertionOutcome(
                name=f"answer_workbook:{index}",
                passed=False,
                detail=f"Unsupported answer_workbook spec: {item!r}",
            ))
            continue

        required = bool(spec.get("required", True))
        expected_path = _resolve_expected_workbook_path(base_dir, spec.get("path", spec.get("file")))
        if expected_path is None or not expected_path.exists():
            outcomes.append(AssertionOutcome(
                name=f"answer_workbook:{index}",
                passed=False,
                detail=f"Golden workbook not found: {expected_path}",
                required=required,
            ))
            continue
        if not candidates:
            outcomes.append(AssertionOutcome(
                name=f"answer_workbook:{expected_path.name}",
                passed=False,
                detail="No generated .xlsx/.xlsm workbook found outside normalized/raw inputs.",
                required=required,
            ))
            continue

        abs_tol = _float_option(spec, ("abs_tol", "tolerance", "tol"), default=1e-6)
        rel_tol = _float_option(spec, ("rel_tol",), default=0.0)
        min_match_ratio = _float_option(spec, ("min_match_ratio", "min_ratio"), default=1.0)
        max_mismatches = _int_option(spec, ("max_mismatches",), default=0)

        comparisons = []
        errors = []
        for candidate in candidates:
            try:
                comparisons.append(compare_workbooks(
                    candidate,
                    expected_path,
                    ranges=spec.get("ranges", spec.get("range", spec.get("answer_position"))),
                    default_sheet=spec.get("sheet", spec.get("answer_sheet")),
                    abs_tol=abs_tol,
                    rel_tol=rel_tol,
                    data_only=bool(spec.get("data_only", True)),
                ))
            except Exception as exc:
                errors.append(f"{candidate.name}: {type(exc).__name__}: {exc}")

        if not comparisons:
            outcomes.append(AssertionOutcome(
                name=f"answer_workbook:{expected_path.name}",
                passed=False,
                detail="; ".join(errors)[:1000] or "No comparable workbook candidates.",
                required=required,
            ))
            continue

        best = max(comparisons, key=lambda result: (result.match_ratio, -result.mismatched_cells))
        passed = best.passed(min_match_ratio=min_match_ratio, max_mismatches=max_mismatches)
        detail = best.summary()
        if errors:
            detail = f"{detail}; candidate_errors={errors[:3]}"
        outcomes.append(AssertionOutcome(
            name=f"answer_workbook:{expected_path.name}",
            passed=passed,
            detail=detail[:2000],
            required=required,
        ))
    return outcomes


def _candidate_answer_workbooks(output_files: list[Path]) -> list[Path]:
    candidates: list[Path] = []
    seen: set[Path] = set()
    input_suffixes = ("_input.xlsx", "_init.xlsx", "initial.xlsx")
    for raw_path in output_files:
        path = Path(raw_path)
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        if not path.exists():
            continue
        parts = {part.lower() for part in path.parts}
        if "normalized" in parts or "raw" in parts:
            continue
        if path.name.lower().endswith(input_suffixes):
            continue
        resolved = path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        candidates.append(path)
    return candidates


def _resolve_expected_workbook_path(base_dir: Path, raw_path: Any) -> Path | None:
    if raw_path is None:
        return None
    path = Path(str(raw_path))
    if path.is_absolute():
        return path
    return (base_dir / path).resolve()


def _int_option(spec: dict[str, Any], names: tuple[str, ...], default: int) -> int:
    for name in names:
        value = spec.get(name)
        if value is None:
            continue
        try:
            return int(value)
        except (TypeError, ValueError):
            continue
    return default


def _columns_for_sheet(loaded_workbooks: list[tuple[Path, Any]], sheet_name: str) -> set[str]:
    columns: set[str] = set()
    for _, workbook in loaded_workbooks:
        candidate_sheets = workbook.sheetnames if sheet_name == "*" else [sheet_name]
        for candidate in candidate_sheets:
            if candidate not in workbook.sheetnames:
                continue
            worksheet = workbook[candidate]
            first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            columns.update(str(value).strip() for value in first_row if value is not None)
    return columns


async def run_case(case: EvalCase, run_dir: Path, args: argparse.Namespace) -> dict[str, Any]:
    _load_dotenv()
    from app.agent.orchestrator import build_orchestrator
    from app.config import Config
    from app.session import Session

    case_dir = run_dir / "cases" / case.id
    workspace_root = case_dir / "workspace"
    case_dir.mkdir(parents=True, exist_ok=True)
    workspace_root.mkdir(parents=True, exist_ok=True)

    if not case.file_path.exists():
        raise FileNotFoundError(f"Case file not found: {case.file_path}")

    config = Config()
    config.workspace_dir = str(workspace_root)
    if args.sandbox_timeout is not None:
        config.sandbox_timeout = args.sandbox_timeout
    if args.max_repair_attempts is not None:
        config.max_repair_attempts = args.max_repair_attempts

    if not config.llm_api_key:
        raise RuntimeError("DEEPSEEK_API_KEY is empty. Set it in .env or the environment before running evals.")

    session = Session.create(str(case.file_path))
    orchestrator = build_orchestrator(config)
    started = time.perf_counter()
    before_workspaces = _workspace_ids(workspace_root)
    exception_text: str | None = None
    task_result = None
    step_outputs: list[str] = []

    async def on_step_start(step: Any, step_index: int, total_steps: int) -> None:
        LOGGER.info("[%s] start step %s/%s %s: %s", case.id, step_index, total_steps, step.id, step.description)

    async def on_step_end(step: Any, result: Any) -> None:
        stdout = getattr(result, "stdout", "") or ""
        error = getattr(result, "error", "") or ""
        if stdout or error:
            step_outputs.append(f"## {step.id}: {step.description}\n{stdout}\n{error}".strip())
        LOGGER.info(
            "[%s] end step %s: failed=%s files=%d",
            case.id,
            step.id,
            getattr(result, "failed", False),
            len(getattr(result, "files", []) or []),
        )

    try:
        task_result = await orchestrator.run(
            case.question,
            session,
            on_step_start=on_step_start,
            on_step_end=on_step_end,
        )
    except Exception:
        exception_text = traceback.format_exc()
        LOGGER.exception("[%s] failed with unhandled exception", case.id)

    duration = time.perf_counter() - started
    workspace_path = _resolve_workspace_path(workspace_root, session.tasks, before_workspaces)
    snapshot = _build_snapshot(workspace_path, task_result, exception_text, step_outputs=step_outputs)
    outcomes = run_assertions(case, snapshot)
    passed = all(outcome.passed for outcome in outcomes if outcome.required)

    result = {
        "case": case.to_dict(),
        "passed": passed,
        "duration_seconds": round(duration, 3),
        "workspace": str(workspace_path) if workspace_path else None,
        "task_ids": list(session.tasks),
        "state": snapshot.state,
        "output_files": [str(path) for path in snapshot.output_files],
        "step_outputs": [text[:4000] for text in snapshot.step_outputs],
        "assertions": [outcome.to_dict() for outcome in outcomes],
        "exception": exception_text,
    }
    result["failed_assertions"] = [
        outcome["name"]
        for outcome in result["assertions"]
        if outcome["required"] and not outcome["passed"]
    ]
    result["failure_category"] = classify_failure(result)
    _write_json(case_dir / "result.json", result)
    (case_dir / "report.md").write_text(snapshot.report or "", encoding="utf-8")
    return result


def _workspace_ids(workspace_root: Path) -> set[str]:
    if not workspace_root.exists():
        return set()
    return {path.name for path in workspace_root.iterdir() if path.is_dir()}


def _resolve_workspace_path(
    workspace_root: Path,
    task_ids: list[str],
    before_workspaces: set[str],
) -> Path | None:
    if task_ids:
        return workspace_root / task_ids[-1]
    after = [path for path in workspace_root.iterdir() if path.is_dir()] if workspace_root.exists() else []
    new_paths = [path for path in after if path.name not in before_workspaces]
    candidates = new_paths or after
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def _build_snapshot(
    workspace_path: Path | None,
    task_result: Any,
    exception_text: str | None,
    *,
    step_outputs: list[str] | None = None,
) -> ExecutionSnapshot:
    state: dict[str, Any] = {}
    report = getattr(task_result, "report", "") if task_result is not None else ""
    output_files = [Path(path) for path in (getattr(task_result, "files", []) or [])]
    scripts: dict[str, str] = {}

    if workspace_path and workspace_path.exists():
        state = _read_json(workspace_path / "state.json", default={}) or {}
        report_path = workspace_path / "output" / "report.md"
        if not report and report_path.exists():
            report = report_path.read_text(encoding="utf-8")
        # Discover output files from output/, normalized/, and workspace root
        _tabular_suffixes = {".csv", ".tsv", ".xlsx", ".xlsm"}
        search_dirs = [workspace_path / "output", workspace_path / "normalized", workspace_path]
        for search_dir in search_dirs:
            if not search_dir.exists():
                continue
            discovered = [path for path in search_dir.iterdir() if path.is_file()]
            if search_dir == workspace_path:
                # Only pick up tabular files from workspace root (skip state.json etc.)
                discovered = [p for p in discovered if p.suffix.lower() in _tabular_suffixes]
            known = {path.resolve() for path in output_files if path.exists()}
            output_files.extend(path for path in discovered if path.resolve() not in known)
        scripts_dir = workspace_path / "scripts"
        if scripts_dir.exists():
            for script_path in sorted(scripts_dir.glob("*.py")):
                scripts[script_path.name] = script_path.read_text(encoding="utf-8", errors="replace")

    return ExecutionSnapshot(
        state=state,
        report=report or "",
        output_files=output_files,
        step_outputs=step_outputs or [],
        scripts=scripts,
        exception=exception_text,
    )


def _read_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def classify_failure(result: dict[str, Any]) -> str | None:
    if result.get("passed"):
        return None

    exception = str(result.get("exception") or "")
    if exception:
        if _is_transient_error(exception):
            return "transient_error"
        return "unhandled_exception"

    assertions = result.get("assertions") or []
    failed = [
        item
        for item in assertions
        if item.get("required", True) and not item.get("passed")
    ]
    names = [str(item.get("name") or "") for item in failed]
    details = "\n".join(str(item.get("detail") or "") for item in failed)

    if any(name == "workspace_completed" for name in names):
        return "workspace_not_completed"
    if any(name == "report_non_empty" for name in names):
        return "empty_report"
    if any(name.startswith("expected_answer") for name in names):
        if "No marked final answer" in details:
            return "no_final_answer"
        if "numbers_matched=False" in details or "expected_numbers=" in details:
            return "wrong_numeric_answer"
        return "wrong_answer"
    if any(name.startswith("answer_workbook") for name in names):
        return "wrong_workbook"
    if any(name.startswith("required_output_ext") for name in names):
        return "missing_artifact"
    if any(name.endswith("_config") for name in names):
        return "eval_config_error"
    return "assertion_failed"


def write_summary(run_dir: Path, results: list[dict[str, Any]]) -> None:
    passed = [result for result in results if result.get("passed")]
    failed = [result for result in results if not result.get("passed")]
    category_counts = _count_by(results, lambda item: item.get("failure_category") or "passed")
    summary = {
        "run_dir": str(run_dir),
        "total": len(results),
        "passed": len(passed),
        "failed": len(failed),
        "failure_categories": category_counts,
        "results": results,
    }
    _write_json(run_dir / "summary.json", summary)
    _write_failures_csv(run_dir / "failures.csv", failed)
    _write_count_csv(run_dir / "by_failure_category.csv", category_counts)
    _write_count_csv(run_dir / "by_note.csv", _count_by_note(results))

    lines = [
        "# ChatExcel Eval Summary",
        "",
        f"- Total: {len(results)}",
        f"- Passed: {len(passed)}",
        f"- Failed: {len(failed)}",
        "",
        "## Failure Categories",
        "",
        "| Category | Count |",
        "|---|---:|",
    ]
    for category, count in sorted(category_counts.items(), key=lambda item: (-item[1], item[0])):
        lines.append(f"| {category} | {count} |")
    lines.extend([
        "",
        "| Case | Result | Duration | Workspace |",
        "|---|---:|---:|---|",
    ])
    for result in results:
        case = result["case"]
        status = "PASS" if result.get("passed") else "FAIL"
        if result.get("failure_category"):
            status = f"{status} ({result['failure_category']})"
        lines.append(
            f"| {case['id']} | {status} | {result.get('duration_seconds', 0)}s | {result.get('workspace') or ''} |"
        )
    (run_dir / "summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def _count_by(results: list[dict[str, Any]], key_fn: Any) -> dict[str, int]:
    counts: dict[str, int] = {}
    for result in results:
        key = str(key_fn(result) or "unknown")
        counts[key] = counts.get(key, 0) + 1
    return counts


def _count_by_note(results: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for result in results:
        for note in result.get("case", {}).get("notes") or []:
            key = str(note)
            counts[key] = counts.get(key, 0) + 1
    return counts


def _write_count_csv(path: Path, counts: dict[str, int]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["key", "count"])
        writer.writeheader()
        for key, count in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
            writer.writerow({"key": key, "count": count})


def _write_failures_csv(path: Path, failed_results: list[dict[str, Any]]) -> None:
    fieldnames = [
        "case_id",
        "failure_category",
        "failed_assertions",
        "duration_seconds",
        "source",
        "notes",
        "workspace",
    ]
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for result in failed_results:
            case = result.get("case", {})
            writer.writerow({
                "case_id": case.get("id", ""),
                "failure_category": result.get("failure_category") or "",
                "failed_assertions": ";".join(result.get("failed_assertions") or []),
                "duration_seconds": result.get("duration_seconds", ""),
                "source": case.get("source", ""),
                "notes": ";".join(str(note) for note in case.get("notes", []) or []),
                "workspace": result.get("workspace") or "",
            })


def default_manifests() -> list[Path]:
    candidates = [
        PROJECT_ROOT / "docs" / "test_datasets" / "manifest.json",
        PROJECT_ROOT / "docs" / "test_datasets" / "简单测试清单.md",
    ]
    return [path for path in candidates if path.exists()]


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run ChatExcel batch evaluations.")
    parser.add_argument(
        "--manifest",
        action="append",
        default=[],
        help="JSON manifest or Markdown checklist. May be passed multiple times.",
    )
    parser.add_argument(
        "--benchmark",
        action="append",
        choices=["sheetbench", "spreadsheetbench", "spreadsheetbench-v2", "all"],
        default=[],
        help=(
            "Materialize and run a public benchmark. May repeat. "
            "Defaults: sheetbench=qa, spreadsheetbench=verified, spreadsheetbench-v2=example."
        ),
    )
    parser.add_argument(
        "--benchmark-variant",
        help="Benchmark variant to use, e.g. qa/complex-qa/verified/example/full.",
    )
    parser.add_argument(
        "--benchmark-data-dir",
        default=str(PROJECT_ROOT / "eval_datasets"),
        help="Directory for benchmark archives, extracted files, and generated manifests.",
    )
    parser.add_argument(
        "--benchmark-force",
        action="store_true",
        help="Redownload/re-extract benchmark archives before generating manifests.",
    )
    parser.add_argument(
        "--benchmark-no-download",
        action="store_true",
        help="Use existing benchmark archives only; fail if they are missing.",
    )
    parser.add_argument("--case-id", action="append", default=[], help="Run only this case id. May repeat.")
    parser.add_argument("--limit", type=int, help="Run only the first N selected cases.")
    parser.add_argument("--dry-run", action="store_true", help="List cases without calling the agent.")
    parser.add_argument(
        "--output-dir",
        help="Directory for eval artifacts. Defaults to eval_runs/<timestamp> under project root.",
    )
    parser.add_argument("--sandbox-timeout", type=int, help="Override sandbox timeout seconds.")
    parser.add_argument("--max-repair-attempts", type=int, help="Override max repair attempts.")
    parser.add_argument("--fail-fast", action="store_true", help="Stop after the first failing case.")
    parser.add_argument("--retries", type=int, default=2, help="Max attempts per case on transient (network) errors.")
    parser.add_argument("--log-level", default=os.getenv("LOG_LEVEL", "INFO"), help="Python logging level.")
    return parser


def _configure_logging(run_dir: Path | None, level_name: str) -> None:
    level = getattr(logging, level_name.upper(), logging.INFO)
    handlers: list[logging.Handler] = [logging.StreamHandler()]
    if run_dir is not None:
        run_dir.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(run_dir / "eval.log", encoding="utf-8"))
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%H:%M:%S",
        handlers=handlers,
        force=True,
    )


_TRANSIENT_ERROR_PATTERNS = (
    "nodename nor servname provided",
    "Name or service not known",
    "Temporary failure in name resolution",
    "Connection refused",
    "Connection reset by peer",
    "Connection timed out",
    "TimeoutError",
    "HTTPSConnectionPool",
    "RemoteDisconnected",
    "status code: 500",
    "status code: 502",
    "status code: 503",
    "status code: 529",
)


def _is_transient_error(exception_text: str) -> bool:
    """Check if an exception looks like a transient network/server error."""
    return any(pattern in exception_text for pattern in _TRANSIENT_ERROR_PATTERNS)


async def async_main(args: argparse.Namespace) -> int:
    manifest_paths = [Path(value) for value in args.manifest]
    if args.benchmark:
        from scripts.benchmark_data import materialize_benchmarks

        manifest_paths.extend(materialize_benchmarks(
            args.benchmark,
            output_dir=args.benchmark_data_dir,
            variant=args.benchmark_variant,
            force=args.benchmark_force,
            download=not args.benchmark_no_download,
        ))
    if not manifest_paths:
        manifest_paths = default_manifests()
    all_cases: list[EvalCase] = []
    for manifest in manifest_paths:
        all_cases.extend(load_cases(manifest))
    cases = select_cases(all_cases, case_ids=args.case_id, limit=args.limit)

    if args.dry_run:
        _configure_logging(None, args.log_level)
        print(f"Loaded {len(all_cases)} cases; selected {len(cases)} cases.")
        for case in cases:
            exists = "OK" if case.file_path.exists() else "MISSING"
            print(f"{case.id}\t{exists}\t{case.file_path}\t{case.question}")
        return 0

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = Path(args.output_dir) if args.output_dir else PROJECT_ROOT / "eval_runs" / timestamp
    if not run_dir.is_absolute():
        run_dir = (PROJECT_ROOT / run_dir).resolve()
    _configure_logging(run_dir, args.log_level)

    LOGGER.info("Selected %d cases from %d loaded cases", len(cases), len(all_cases))
    _write_json(run_dir / "cases.json", [case.to_dict() for case in cases])

    max_retries = getattr(args, "retries", 1) or 1
    results: list[dict[str, Any]] = []
    for index, case in enumerate(cases, start=1):
        LOGGER.info("Running case %d/%d: %s", index, len(cases), case.id)
        result: dict[str, Any] | None = None
        for attempt in range(1, max_retries + 1):
            try:
                result = await run_case(case, run_dir, args)
            except Exception:
                error = traceback.format_exc()
                LOGGER.exception("Case bootstrap failed: %s", case.id)
                result = {
                    "case": case.to_dict(),
                    "passed": False,
                    "duration_seconds": 0,
                    "workspace": None,
                    "task_ids": [],
                    "state": {},
                    "output_files": [],
                    "assertions": [
                        AssertionOutcome(
                            name="case_bootstrap",
                            passed=False,
                            detail=error[:1000],
                        ).to_dict()
                    ],
                    "exception": error,
                }
            if result.get("passed") or attempt >= max_retries:
                break
            # Retry only on network / transient errors
            exc_text = result.get("exception") or ""
            if not _is_transient_error(exc_text):
                break
            LOGGER.warning("[%s] transient error on attempt %d/%d, retrying...", case.id, attempt, max_retries)
        results.append(result)
        write_summary(run_dir, results)
        if args.fail_fast and not result.get("passed"):
            break

    failed = [result for result in results if not result.get("passed")]
    LOGGER.info("Eval finished: total=%d passed=%d failed=%d", len(results), len(results) - len(failed), len(failed))
    return 1 if failed else 0


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    return asyncio.run(async_main(args))


if __name__ == "__main__":
    raise SystemExit(main())
