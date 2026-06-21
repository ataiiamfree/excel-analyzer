"""Workbook value comparison helpers for benchmark evaluation."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries


@dataclass(frozen=True)
class RangeSpec:
    sheet_name: str | None
    coordinate: str | None


@dataclass
class CellMismatch:
    sheet_name: str
    coordinate: str
    expected: Any
    actual: Any

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet_name,
            "cell": self.coordinate,
            "expected": _jsonable(self.expected),
            "actual": _jsonable(self.actual),
        }


@dataclass
class WorkbookComparison:
    candidate_path: Path
    expected_path: Path
    checked_cells: int = 0
    matched_cells: int = 0
    mismatched_cells: int = 0
    missing_sheets: list[str] = field(default_factory=list)
    range_errors: list[str] = field(default_factory=list)
    sample_mismatches: list[CellMismatch] = field(default_factory=list)

    @property
    def match_ratio(self) -> float:
        if self.checked_cells == 0:
            return 1.0 if not self.missing_sheets and not self.range_errors else 0.0
        return self.matched_cells / self.checked_cells

    def passed(self, *, min_match_ratio: float = 1.0, max_mismatches: int = 0) -> bool:
        return (
            not self.missing_sheets
            and not self.range_errors
            and self.match_ratio >= min_match_ratio
            and self.mismatched_cells <= max_mismatches
        )

    def summary(self) -> str:
        parts = [
            f"candidate={self.candidate_path.name}",
            f"match_ratio={self.match_ratio:.4f}",
            f"matched={self.matched_cells}",
            f"checked={self.checked_cells}",
            f"mismatches={self.mismatched_cells}",
        ]
        if self.missing_sheets:
            parts.append(f"missing_sheets={self.missing_sheets[:5]}")
        if self.range_errors:
            parts.append(f"range_errors={self.range_errors[:3]}")
        if self.sample_mismatches:
            parts.append(f"samples={[item.to_dict() for item in self.sample_mismatches[:3]]}")
        return "; ".join(parts)

    def to_dict(self) -> dict[str, Any]:
        return {
            "candidate": str(self.candidate_path),
            "expected": str(self.expected_path),
            "checked_cells": self.checked_cells,
            "matched_cells": self.matched_cells,
            "mismatched_cells": self.mismatched_cells,
            "match_ratio": self.match_ratio,
            "missing_sheets": self.missing_sheets,
            "range_errors": self.range_errors,
            "sample_mismatches": [item.to_dict() for item in self.sample_mismatches],
        }


def compare_workbooks(
    candidate_path: str | Path,
    expected_path: str | Path,
    *,
    ranges: Any = None,
    default_sheet: Any = None,
    abs_tol: float = 1e-6,
    rel_tol: float = 0.0,
    data_only: bool = True,
    max_samples: int = 20,
) -> WorkbookComparison:
    """Compare a produced workbook against a golden workbook.

    Empty cells where both workbooks are blank are skipped. Cells where either
    side has a value are checked, which catches missing outputs and unexpected
    extra values inside the answer range.
    """
    candidate_path = Path(candidate_path)
    expected_path = Path(expected_path)
    comparison = WorkbookComparison(candidate_path=candidate_path, expected_path=expected_path)

    candidate_wb = load_workbook(candidate_path, data_only=data_only, read_only=True)
    expected_wb = load_workbook(expected_path, data_only=data_only, read_only=True)
    try:
        specs = parse_range_specs(ranges, default_sheet=default_sheet)
        if not specs:
            specs = [RangeSpec(sheet_name=sheet_name, coordinate=None) for sheet_name in expected_wb.sheetnames]

        for spec in specs:
            sheet_name = spec.sheet_name or _default_sheet_name(default_sheet) or expected_wb.sheetnames[0]
            if sheet_name not in expected_wb.sheetnames:
                comparison.range_errors.append(f"Expected workbook has no sheet {sheet_name!r}.")
                continue
            if sheet_name not in candidate_wb.sheetnames:
                comparison.missing_sheets.append(sheet_name)
                continue

            expected_ws = expected_wb[sheet_name]
            candidate_ws = candidate_wb[sheet_name]
            try:
                bounds = _bounds_for_range(spec.coordinate, expected_ws, candidate_ws)
            except ValueError as exc:
                comparison.range_errors.append(f"{sheet_name}!{spec.coordinate}: {exc}")
                continue

            min_col, min_row, max_col, max_row = bounds
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    expected_value = expected_ws.cell(row=row, column=col).value
                    actual_value = candidate_ws.cell(row=row, column=col).value
                    if _is_blank(expected_value) and _is_blank(actual_value):
                        continue
                    comparison.checked_cells += 1
                    if _values_equal(actual_value, expected_value, abs_tol=abs_tol, rel_tol=rel_tol):
                        comparison.matched_cells += 1
                    else:
                        comparison.mismatched_cells += 1
                        if len(comparison.sample_mismatches) < max_samples:
                            comparison.sample_mismatches.append(
                                CellMismatch(
                                    sheet_name=sheet_name,
                                    coordinate=f"{get_column_letter(col)}{row}",
                                    expected=expected_value,
                                    actual=actual_value,
                                )
                            )
        return comparison
    finally:
        candidate_wb.close()
        expected_wb.close()


def parse_range_specs(ranges: Any, *, default_sheet: Any = None) -> list[RangeSpec]:
    if not ranges:
        return []
    if isinstance(ranges, RangeSpec):
        return [ranges]
    if isinstance(ranges, str):
        tokens = _split_range_tokens(ranges)
    elif isinstance(ranges, (list, tuple, set)):
        tokens = []
        for item in ranges:
            if isinstance(item, RangeSpec):
                tokens.append(item)
            else:
                tokens.extend(_split_range_tokens(str(item)))
    else:
        tokens = _split_range_tokens(str(ranges))

    specs: list[RangeSpec] = []
    for token in tokens:
        if isinstance(token, RangeSpec):
            specs.append(token)
            continue
        sheet_name, coordinate = _parse_range_token(str(token), default_sheet=default_sheet)
        specs.append(RangeSpec(sheet_name=sheet_name, coordinate=coordinate))
    return specs


def _split_range_tokens(text: str) -> list[str]:
    tokens: list[str] = []
    current: list[str] = []
    in_quote = False
    i = 0
    while i < len(text):
        char = text[i]
        if char == "'":
            if i + 1 < len(text) and text[i + 1] == "'":
                current.append("''")
                i += 2
                continue
            in_quote = not in_quote
            current.append(char)
        elif char == "," and not in_quote:
            token = "".join(current).strip()
            if token:
                tokens.append(token)
            current = []
        else:
            current.append(char)
        i += 1
    token = "".join(current).strip()
    if token:
        tokens.append(token)
    return tokens


def _parse_range_token(token: str, *, default_sheet: Any = None) -> tuple[str | None, str | None]:
    token = token.strip().strip(",")
    if not token:
        return _default_sheet_name(default_sheet), None

    if "!" in token:
        sheet_part, coordinate = token.rsplit("!", 1)
        return _clean_sheet_name(sheet_part), _clean_coordinate(coordinate)
    return _default_sheet_name(default_sheet), _clean_coordinate(token)


def _default_sheet_name(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (list, tuple, set)):
        for item in value:
            name = _default_sheet_name(item)
            if name:
                return name
        return None
    text = str(value).strip()
    if not text:
        return None
    first = _split_range_tokens(text)[0] if "," in text else text
    return _clean_sheet_name(first)


def _clean_sheet_name(value: str) -> str:
    text = value.strip()
    if text.startswith("'") and text.endswith("'") and len(text) >= 2:
        text = text[1:-1]
    return text.strip("'").replace("''", "'")


def _clean_coordinate(value: str) -> str | None:
    text = value.strip().strip("'")
    return text or None


def _bounds_for_range(coordinate: str | None, expected_ws: Any, candidate_ws: Any) -> tuple[int, int, int, int]:
    if not coordinate:
        return (
            1,
            1,
            max(expected_ws.max_column or 1, candidate_ws.max_column or 1),
            max(expected_ws.max_row or 1, candidate_ws.max_row or 1),
        )

    normalized = coordinate.replace("$", "")
    min_col, min_row, max_col, max_row = range_boundaries(normalized)
    min_col = min_col or 1
    max_col = max_col or max(expected_ws.max_column or 1, candidate_ws.max_column or 1)
    min_row = min_row or 1
    max_row = max_row or max(expected_ws.max_row or 1, candidate_ws.max_row or 1)
    if min_col > max_col or min_row > max_row:
        raise ValueError(f"Invalid range bounds: {coordinate!r}")
    return min_col, min_row, max_col, max_row


def _values_equal(actual: Any, expected: Any, *, abs_tol: float, rel_tol: float) -> bool:
    if _is_blank(actual) and _is_blank(expected):
        return True
    if _is_blank(actual) or _is_blank(expected):
        return False

    actual_number = _number(actual)
    expected_number = _number(expected)
    if actual_number is not None and expected_number is not None:
        tolerance = max(abs_tol, rel_tol * abs(expected_number))
        return abs(actual_number - expected_number) <= tolerance

    if isinstance(actual, (datetime, date, time)) or isinstance(expected, (datetime, date, time)):
        return _datetime_text(actual) == _datetime_text(expected)

    return _normal_text(actual) == _normal_text(expected)


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?%?", text.replace(",", "")):
        is_percent = text.endswith("%")
        number_text = text[:-1] if is_percent else text
        try:
            number = float(number_text.replace(",", ""))
        except ValueError:
            return None
        return number / 100 if is_percent else number
    return None


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _normal_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value).strip())


def _datetime_text(value: Any) -> str:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value).strip()


def _jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value
