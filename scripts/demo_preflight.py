"""Read-only preflight checks for the ChatExcel v0.9 Max demo.

This script never starts or stops services, edits configuration, deletes
conversations, or prints secrets. Start the release-like server first, then run:

    python scripts/demo_preflight.py
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable
from urllib.request import urlopen

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parent.parent
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))
DEFAULT_WORKBOOK = REPO / "docs" / "test_datasets" / "simple" / "01_门店月度销售.xlsx"
DEFAULT_WEB_DIST = REPO / "web" / "dist"
EXPECTED_TOTALS = {
    "杭州西湖店": 753.3,
    "北京旗舰店": 739.7,
    "上海南京路店": 664.9,
    "广州天河店": 664.8,
    "成都春熙路店": 566.7,
}


@dataclass(frozen=True)
class CheckResult:
    name: str
    ok: bool
    detail: str


JsonFetcher = Callable[[str, float], dict[str, Any]]


def fetch_json(url: str, timeout: float = 3.0) -> dict[str, Any]:
    with urlopen(url, timeout=timeout) as response:  # noqa: S310 - explicit local demo URL
        return json.load(response)


def check_service(base_url: str, fetcher: JsonFetcher = fetch_json) -> list[CheckResult]:
    base = base_url.rstrip("/")
    try:
        health = fetcher(f"{base}/api/health", 3.0)
        openapi = fetcher(f"{base}/openapi.json", 3.0)
    except Exception as exc:  # noqa: BLE001 - turn network failures into a concise preflight result
        return [CheckResult("分析服务", False, f"无法连接：{type(exc).__name__}")]

    return [
        CheckResult("健康检查", health.get("status") == "ok", f"status={health.get('status')!r}"),
        CheckResult(
            "API 版本",
            openapi.get("info", {}).get("version") == "0.9.2",
            f"version={openapi.get('info', {}).get('version')!r}",
        ),
    ]


def check_workbook(path: Path) -> list[CheckResult]:
    if not path.is_file():
        return [CheckResult("演示工作簿", False, f"文件不存在：{path}")]

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "月度销售" not in workbook.sheetnames:
            return [CheckResult("演示工作簿", False, "缺少工作表：月度销售")]
        sheet = workbook["月度销售"]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
    except Exception as exc:  # noqa: BLE001 - corrupted/encrypted workbook should be a readable failure
        return [CheckResult("演示工作簿", False, f"无法读取：{type(exc).__name__}")]

    if not rows:
        return [CheckResult("演示工作簿", False, "工作表为空")]

    headers = [str(value or "").strip() for value in rows[0]]
    required = {"门店", "月份", "销售额(万元)"}
    missing = sorted(required.difference(headers))
    if missing:
        return [CheckResult("演示工作簿", False, f"缺少列：{', '.join(missing)}")]

    store_index = headers.index("门店")
    sales_index = headers.index("销售额(万元)")
    totals: dict[str, float] = {}
    data_rows = 0
    for row in rows[1:]:
        store = str(row[store_index] or "").strip()
        sales = row[sales_index]
        if not store or sales is None:
            continue
        data_rows += 1
        totals[store] = totals.get(store, 0.0) + float(sales)

    rounded = {store: round(total, 1) for store, total in totals.items()}
    return [
        CheckResult("演示工作簿结构", data_rows == 30, f"有效数据 {data_rows} 行"),
        CheckResult("演示标准答案", rounded == EXPECTED_TOTALS, "5 家门店汇总值已核对"),
    ]


def check_web_build(path: Path) -> list[CheckResult]:
    index_ok = (path / "index.html").is_file()
    asset_dir = path / "assets"
    js_ok = asset_dir.is_dir() and any(asset_dir.glob("*.js"))
    css_ok = asset_dir.is_dir() and any(asset_dir.glob("*.css"))
    return [
        CheckResult(
            "前端生产构建",
            index_ok and js_ok and css_ok,
            "index.html、JS、CSS 均存在" if index_ok and js_ok and css_ok else f"构建不完整：{path}",
        )
    ]


def check_runtime() -> list[CheckResult]:
    from app.agent.runtime import build_pi_rpc_transport
    from app.config import Config

    config = Config()
    transport = build_pi_rpc_transport(config)
    command = transport.command
    key_configured = bool(config.llm_api_key and not config.llm_api_key.startswith("your-"))
    pi_available = bool(shutil.which(command[0]))
    context_guard = "--no-context-files" in command or "-nc" in command
    tools_guard = "--tools" in command and "bash" in command
    git_guard = bool(transport.env and transport.env.get("GIT_DIR") == os.devnull)

    return [
        CheckResult("LLM 配置", key_configured, f"模型 {config.llm_model}；密钥不显示"),
        CheckResult("Pi 命令", pi_available, "可执行文件已找到" if pi_available else f"未找到 {command[0]}"),
        CheckResult("Pi 上下文隔离", context_guard, "未加载仓库上下文文件"),
        CheckResult("Pi 工具限制", tools_guard, "仅保留分析桥接所需 bash"),
        CheckResult("Pi Git 防线", git_guard, "子进程无法发现应用仓库"),
    ]


def run_preflight(
    *,
    base_url: str,
    workbook_path: Path,
    web_dist_path: Path,
    fetcher: JsonFetcher = fetch_json,
) -> list[CheckResult]:
    return [
        *check_service(base_url, fetcher),
        *check_workbook(workbook_path),
        *check_web_build(web_dist_path),
        *check_runtime(),
    ]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="ChatExcel v0.9 Max 演示前只读预检")
    parser.add_argument("--base-url", default="http://127.0.0.1:8000")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--web-dist", type=Path, default=DEFAULT_WEB_DIST)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    results = run_preflight(
        base_url=args.base_url,
        workbook_path=args.workbook,
        web_dist_path=args.web_dist,
    )

    print("ChatExcel v0.9 Max 演示预检")
    for result in results:
        marker = "✓" if result.ok else "✗"
        print(f"  {marker} {result.name} — {result.detail}")

    failed = [result for result in results if not result.ok]
    print(f"\n结果：{len(results) - len(failed)}/{len(results)} 项通过")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
