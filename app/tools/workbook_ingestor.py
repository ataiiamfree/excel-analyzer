"""Workbook structure scanning.

The ingestor does not modify workbooks. It produces a manifest describing
sheets and rough table candidates so preprocessing has an explicit contract.
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


class WorkbookIngestor:
    # 连续空行数 ≥ 此阈值时视为表间分隔
    _GAP_ROWS = 2

    def scan(self, file_path: str | Path) -> dict[str, Any]:
        path = Path(file_path)
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
        sheets = []
        for worksheet in workbook.worksheets:
            auto_filter_ref = self._auto_filter_ref(worksheet)
            auto_filter_header = self._auto_filter_header(auto_filter_ref)
            regions = self._detect_table_regions(worksheet)
            tables = []
            for idx, bounds in enumerate(regions, start=1):
                min_row, min_col, max_row, max_col = bounds
                header_candidates = self._header_candidates(worksheet, bounds)
                notes: list[str] = []
                confidence = 0.7
                if auto_filter_header is not None and min_row <= auto_filter_header <= max_row:
                    header_candidates = [auto_filter_header]
                    notes.append(f"AutoFilter 表头强信号: {auto_filter_ref}")
                    confidence = 0.95
                tables.append(
                    {
                        "table_id": f"{worksheet.title}_t{idx}",
                        "range": (
                            f"{get_column_letter(min_col)}{min_row}:"
                            f"{get_column_letter(max_col)}{max_row}"
                        ),
                        "header_candidates": header_candidates,
                        "confidence": confidence,
                        "notes": notes,
                    }
                )
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_col": worksheet.max_column,
                    "hidden_rows": [
                        idx
                        for idx, dim in worksheet.row_dimensions.items()
                        if getattr(dim, "hidden", False)
                    ],
                    "hidden_cols": [
                        key
                        for key, dim in worksheet.column_dimensions.items()
                        if getattr(dim, "hidden", False)
                    ],
                    "auto_filter_ref": auto_filter_ref,
                    "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
                    "tables": tables,
                }
            )
        return {
            "manifest_path": "workbook_manifest.json",
            "files": [{"path": str(path), "sheets": sheets}],
        }

    def _detect_table_regions(
        self, worksheet: Any
    ) -> list[tuple[int, int, int, int]]:
        """Split a worksheet into table regions separated by blank-row gaps.

        Returns a list of (min_row, min_col, max_row, max_col) bounding boxes.
        Adjacent non-empty rows belong to the same region; a gap of ≥ _GAP_ROWS
        consecutive blank rows starts a new region.
        """
        overall = self._non_empty_bounds(worksheet)
        if overall is None:
            return []

        min_row, min_col, max_row, max_col = overall

        # Build a set of row indices that have at least one non-empty cell
        non_empty_rows: set[int] = set()
        for row in worksheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                if cell.value not in (None, ""):
                    non_empty_rows.add(cell.row)
                    break

        if not non_empty_rows:
            return []

        # Walk rows top-to-bottom, split into blocks separated by ≥ _GAP_ROWS blanks
        sorted_rows = sorted(non_empty_rows)
        regions: list[tuple[int, int, int, int]] = []
        block_start = sorted_rows[0]
        prev_row = sorted_rows[0]

        for row_idx in sorted_rows[1:]:
            gap = row_idx - prev_row - 1
            if gap >= self._GAP_ROWS:
                # Close current block — refine column bounds
                bounds = self._refine_col_bounds(worksheet, block_start, prev_row, min_col, max_col)
                regions.append(bounds)
                block_start = row_idx
            prev_row = row_idx

        # Close last block
        bounds = self._refine_col_bounds(worksheet, block_start, prev_row, min_col, max_col)
        regions.append(bounds)

        return regions

    def _refine_col_bounds(
        self,
        worksheet: Any,
        row_start: int,
        row_end: int,
        global_min_col: int,
        global_max_col: int,
    ) -> tuple[int, int, int, int]:
        """Tighten column bounds to the actual non-empty range within a row block."""
        actual_min_col = global_max_col
        actual_max_col = global_min_col
        for row in worksheet.iter_rows(
            min_row=row_start,
            max_row=row_end,
            min_col=global_min_col,
            max_col=global_max_col,
        ):
            for cell in row:
                if cell.value not in (None, ""):
                    actual_min_col = min(actual_min_col, cell.column)
                    actual_max_col = max(actual_max_col, cell.column)
        return (row_start, actual_min_col, row_end, actual_max_col)

    def _non_empty_bounds(self, worksheet: Any) -> tuple[int, int, int, int] | None:
        min_row = min_col = None
        max_row = max_col = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value in (None, ""):
                    continue
                min_row = cell.row if min_row is None else min(min_row, cell.row)
                min_col = cell.column if min_col is None else min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
        if min_row is None or min_col is None:
            return None
        return min_row, min_col, max_row, max_col

    def _auto_filter_ref(self, worksheet: Any) -> str | None:
        value = getattr(getattr(worksheet, "auto_filter", None), "ref", None)
        return str(value) if value else None

    def _auto_filter_header(self, auto_filter_ref: str | None) -> int | None:
        if not auto_filter_ref:
            return None
        try:
            _, min_row, _, _ = range_boundaries(auto_filter_ref)
        except ValueError:
            return None
        return min_row

    def _header_candidates(
        self, worksheet: Any, bounds: tuple[int, int, int, int]
    ) -> list[int]:
        """Find header row candidates — consecutive text-heavy rows at the top.

        Stops scanning after the first non-candidate row following a candidate,
        so data rows with text content don't get mistakenly included.

        Density is computed against **merge-expanded** row values: cells inside
        a merged range use the range's top-left value. Without this, parent
        group rows like `Total Fundraising` merged across 5 columns look sparse
        (only column B has the value) and get rejected, which erases the whole
        upper level of a multi-level header.
        """
        min_row, min_col, max_row, max_col = bounds
        candidates: list[int] = []
        scan_end = min(max_row, min_row + 19)
        width = max_col - min_col + 1
        merged_map = self._build_merged_value_map(
            worksheet, min_row, scan_end, min_col, max_col
        )
        # Pre-compute per-row signals so the primary scan (top-down, density-
        # gated) and the sparse backward extension can share the same view.
        row_signals: dict[int, dict[str, Any]] = {}
        for row_idx in range(min_row, scan_end + 1):
            raw_values = [
                worksheet.cell(row_idx, col_idx).value
                for col_idx in range(min_col, max_col + 1)
            ]
            effective_values = [
                merged_map.get((row_idx, col_idx), raw_values[col_idx - min_col])
                for col_idx in range(min_col, max_col + 1)
            ]
            non_empty = [v for v in effective_values if v not in (None, "")]
            # Calendar years appearing as ints (2022, 2023, …) are header
            # labels in practice, not data measurements. Treat them as text
            # equivalents when judging header-ness so leaf year rows without
            # string quoting still get picked up.
            text_like = [
                v for v in non_empty
                if isinstance(v, str) or self._is_year_like_int(v)
            ]
            data_values = [
                v for v in non_empty
                if isinstance(v, (int, float, datetime.date, datetime.datetime))
                and not self._is_year_like_int(v)
            ]
            unique_texts = {
                str(v).strip()
                for v in text_like
                if str(v).strip()
            }
            row_signals[row_idx] = {
                "blank": not non_empty,
                "non_empty_count": len(non_empty),
                "text_like_count": len(text_like),
                "data_count": len(data_values),
                "density": len(non_empty) / width if width else 0,
                "text_ratio": len(text_like) / len(non_empty) if non_empty else 0,
                "unique_texts": unique_texts,
            }

        found_first = False
        for row_idx in range(min_row, scan_end + 1):
            sig = row_signals[row_idx]
            if sig["blank"]:
                if found_first:
                    break
                continue
            if sig["data_count"] > 0:
                # First row carrying real data → end of the header search.
                break
            base_header_like = (
                sig["density"] >= 0.4
                and sig["text_ratio"] >= 0.6
            )
            # A single-value row on its own is almost always a report title
            # (e.g. "Q3 results" merged across every column). We only accept
            # unique=1 rows when they extend an already-detected header block,
            # e.g. a `£(000)` unit row sandwiched between the parent group row
            # and the leaf year row.
            is_primary = base_header_like and (
                len(sig["unique_texts"]) >= 2 or found_first
            )
            if is_primary:
                candidates.append(row_idx)
                found_first = True
            elif found_first and sig["unique_texts"]:
                # Text-only annotation row nested between primaries (e.g. row
                # of `%` markers between region row and leaf year row). Include
                # so per-cell annotations reach the leaf lineage, but do not
                # advance found_first blocks — subsequent primaries handle it.
                candidates.append(row_idx)

        # Backward extension: sparse parent-group rows (e.g. `Landings into` in
        # one cell, `Total landings` in another) are below the density
        # threshold on their own, but sandwiched consecutively above a real
        # leaf header they clearly belong to the same block. Only extend
        # through rows that are text-only, have ≥2 distinct labels, and sit
        # directly above the current earliest candidate.
        if candidates:
            first_row = candidates[0]
            for row_idx in range(first_row - 1, min_row - 1, -1):
                sig = row_signals[row_idx]
                if sig["blank"]:
                    break
                if sig["data_count"] > 0:
                    break
                if len(sig["unique_texts"]) < 2:
                    break
                if sig["text_like_count"] == 0:
                    break
                candidates.insert(0, row_idx)

        # Cap very deep report headers without discarding the leaf row. The
        # leaf carries the actual field names; clipping from the end would keep
        # distant report metadata while turning the leaf into a data record.
        return candidates[-6:] or [min_row]

    def _is_year_like_int(self, value: Any) -> bool:
        if not isinstance(value, int) or isinstance(value, bool):
            return False
        return 1900 <= value <= 2100

    def _build_merged_value_map(
        self,
        worksheet: Any,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> dict[tuple[int, int], Any]:
        """Return {(row, col): value} for cells inside merged ranges in scan area.

        Excel merged ranges store the value only at the top-left cell; treating
        every covered cell as if it held that value lets header detection see
        the true logical density of a parent-group row.
        """
        mapping: dict[tuple[int, int], Any] = {}
        for merged_range in worksheet.merged_cells.ranges:
            r_min = merged_range.min_row
            r_max = merged_range.max_row
            c_min = merged_range.min_col
            c_max = merged_range.max_col
            if r_max < min_row or r_min > max_row:
                continue
            if c_max < min_col or c_min > max_col:
                continue
            top_left_value = worksheet.cell(r_min, c_min).value
            if top_left_value in (None, ""):
                continue
            for r in range(max(r_min, min_row), min(r_max, max_row) + 1):
                for c in range(max(c_min, min_col), min(c_max, max_col) + 1):
                    mapping[(r, c)] = top_left_value
        return mapping
