"""Excel preprocessing helpers.

This module intentionally starts conservative: suspicious rows are flagged, but
business data is not excluded unless the evidence is stronger than a keyword.
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.utils.cell import range_boundaries


@dataclass
class NormalizedTable:
    table_id: str
    source_file: str
    source_sheet: str
    source_range: str
    parquet_path: str
    preview_xlsx_path: str
    columns: list[dict[str, Any]]
    row_count: int
    enum_columns: dict[str, list[str]] = field(default_factory=dict)
    oversized_cells: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    # Map final column name → header lineage from top-level group to leaf column.
    # Always populated: single-level columns have a single-element list [name].
    header_paths: dict[str, list[str]] = field(default_factory=dict)


@dataclass
class PreprocessResult:
    workbook_manifest_path: str | None
    tables: list[NormalizedTable]
    report: dict[str, Any]


class ExcelPreprocessor:
    summary_keywords = ("合计", "小计", "总计", "汇总", "合 计")
    enum_max_unique_values = 15
    enum_max_unique_ratio = 0.5
    oversized_cell_chars = 200

    def process(
        self,
        file_path: str | Path,
        manifest: dict[str, Any],
        output_dir: str | Path | None = None,
    ) -> PreprocessResult:
        file_path = Path(file_path)
        # data_only=False so we can manipulate merged cells; values read via .value
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        workbook_values = openpyxl.load_workbook(file_path, data_only=True)
        workspace_dir = (
            file_path.parent.parent
            if file_path.parent.name == "raw"
            else file_path.parent
        )
        normalized_dir = Path(output_dir) if output_dir is not None else workspace_dir / "normalized"
        normalized_dir.mkdir(parents=True, exist_ok=True)

        tables: list[NormalizedTable] = []
        for file_info in manifest.get("files", []):
            for sheet_info in file_info.get("sheets", []):
                worksheet = workbook[sheet_info["name"]]
                ws_values = workbook_values[sheet_info["name"]]
                # Step 1: 拆分合并单元格并填充
                self._unmerge_and_fill(worksheet, ws_values)
                previous_headers_by_width: dict[int, list[str]] = {}
                for table_candidate in sheet_info.get("tables", []):
                    fallback_headers = previous_headers_by_width.get(
                        self._table_candidate_width(str(table_candidate["range"]))
                    )
                    table = self._normalize_table(
                        file_path=file_path,
                        worksheet=worksheet,
                        ws_values=ws_values,
                        table_candidate=table_candidate,
                        normalized_dir=normalized_dir,
                        fallback_headers=fallback_headers,
                    )
                    tables.append(table)
                    visible_headers = [
                        str(column.get("name"))
                        for column in table.columns
                        if column.get("name") and not str(column.get("name")).startswith("_source_")
                    ]
                    previous_headers_by_width[len(visible_headers)] = visible_headers

        return PreprocessResult(
            workbook_manifest_path=self.normalize_manifest_path(manifest),
            tables=tables,
            report={
                "table_count": len(tables),
                "warnings": [w for table in tables for w in table.warnings],
            },
        )

    def _normalize_table(
        self,
        file_path: Path,
        worksheet: Any,
        ws_values: Any,
        table_candidate: dict[str, Any],
        normalized_dir: Path,
        fallback_headers: list[str] | None = None,
    ) -> NormalizedTable:
        range_ref = str(table_candidate["range"])
        bounds = range_boundaries(range_ref)
        min_col = self._range_bound(bounds[0], range_ref)
        min_row = self._range_bound(bounds[1], range_ref)
        max_col = self._range_bound(bounds[2], range_ref)
        max_row = self._range_bound(bounds[3], range_ref)
        warnings: list[str] = []
        rows = [
            [
                self._cell_value(
                    worksheet=worksheet,
                    ws_values=ws_values,
                    row=row_idx,
                    col=col_idx,
                    warnings=warnings,
                )
                for col_idx in range(min_col, max_col + 1)
            ]
            for row_idx in range(min_row, max_row + 1)
        ]
        header_candidates = table_candidate.get("header_candidates") or [min_row]
        is_continuation = self._looks_like_headerless_continuation(
            rows=rows,
            header_candidates=header_candidates,
            min_row=min_row,
            fallback_headers=fallback_headers,
        )
        raw_header_paths: list[list[str]] = []
        if is_continuation and fallback_headers:
            header_rel = 0
            header_depth = 0
            headers = list(fallback_headers)
            raw_header_paths = [[str(name)] for name in headers]
            warnings.append("检测到疑似续表块，已复用上一段同宽表头")
        else:
            header_start_abs, header_end_abs = self._detect_header_range(
                header_candidates, default=min_row
            )
            header_rel = max(1, header_end_abs - min_row + 1)
            header_depth = header_end_abs - header_start_abs + 1

            # Step 5: 多层表头合并为单层
            if header_depth > 1:
                header_start_rel = max(1, header_start_abs - min_row + 1)
                raw_header_paths = self._merge_multi_level_headers(
                    rows, header_rel, start_rel=header_start_rel
                )
            else:
                raw_header_paths = [
                    [str(value).strip()] if value not in (None, "") else []
                    for value in rows[header_rel - 1]
                ]
            headers = self._dedupe_headers(rows[header_rel - 1])

        # 检测数据区域边界（从底部向上过滤脚注行）
        data_end = self._detect_data_end(rows, header_rel)

        row_flags = self.classify_rows(rows, header_row=header_rel, data_end=data_end)
        labeled_context_rows, labeled_context_columns = self._repeating_labeled_context_rows(
            rows=rows,
            headers=headers,
            header_rel=header_rel,
            data_end=data_end,
        )
        if labeled_context_columns:
            warnings.append(
                "检测到重复键值分组行，已将 "
                f"{', '.join(labeled_context_columns)} 下传到后续数值明细行"
            )

        records: list[dict[str, Any]] = []
        current_context_group: str | None = None
        current_labeled_context: dict[str, Any] = {}
        detected_context_groups = False
        for rel_idx, values in enumerate(rows, start=1):
            if rel_idx <= header_rel:
                continue
            if rel_idx in labeled_context_rows:
                current_labeled_context = {
                    column: labeled_context_rows[rel_idx].get(column)
                    for column in labeled_context_columns
                }
                continue
            context_label = self._context_group_label(
                values,
                headers,
                following_rows=rows[rel_idx:],
            )
            if context_label:
                current_context_group = context_label
                detected_context_groups = True
                warnings.append(
                    f"row {min_row + rel_idx - 1}: 检测到父级/分组标题行 `{context_label}`，"
                    "已写入后续明细的 _context_group"
                )
                continue
            flag = row_flags.get(rel_idx, {})
            if flag.get("warning"):
                warnings.append(f"row {min_row + rel_idx - 1}: {flag['warning']}")
            if flag.get("exclude"):
                continue
            record = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
            }
            record.update(current_labeled_context)
            if detected_context_groups or current_context_group is not None:
                record["_context_group"] = current_context_group
            record["_source_file"] = file_path.name
            record["_source_sheet"] = worksheet.title
            record["_source_row"] = min_row + rel_idx - 1
            records.append(record)

        data_columns = [*headers, *labeled_context_columns]
        if detected_context_groups:
            for record in records:
                record.setdefault("_context_group", None)
            data_columns.append("_context_group")
        dataframe = pd.DataFrame.from_records(
            records,
            columns=[*data_columns, "_source_file", "_source_sheet", "_source_row"],
        )
        table_id = self._safe_table_id(str(table_candidate["table_id"]))
        data_path = normalized_dir / f"{table_id}.parquet"
        preview_path = normalized_dir / f"{table_id}_preview.xlsx"
        try:
            dataframe.to_parquet(data_path, index=False)
        except Exception as exc:  # pragma: no cover - depends on optional parquet engine
            warnings.append(f"parquet 写入失败，已降级为 xlsx: {exc}")
            data_path = normalized_dir / f"{table_id}.xlsx"
            dataframe.to_excel(data_path, index=False)
        dataframe.head(50).to_excel(preview_path, index=False)

        enum_columns = self._detect_enum_columns(dataframe, data_columns)
        oversized_cells = self._detect_oversized_cells(dataframe, headers)
        if oversized_cells:
            warnings.append(f"检测到 {len(oversized_cells)} 个超长文本单元格，profile 预览会截断")

        header_paths = self._build_header_paths(
            deduped_headers=headers,
            raw_paths=raw_header_paths,
            dataframe_columns=[str(col) for col in dataframe.columns],
        )
        column_formats = self._build_column_format_metadata(
            worksheet=worksheet,
            rows=rows,
            row_flags=row_flags,
            headers=headers,
            min_row=min_row,
            min_col=min_col,
            header_rel=header_rel,
            data_end=data_end,
        )

        columns = [
            self._column_metadata(
                dataframe,
                str(col),
                enum_columns,
                header_paths,
                column_formats,
            )
            for col in dataframe.columns
        ]
        return NormalizedTable(
            table_id=table_id,
            source_file=file_path.name,
            source_sheet=worksheet.title,
            source_range=range_ref,
            parquet_path=str(data_path),
            preview_xlsx_path=str(preview_path),
            columns=columns,
            row_count=len(dataframe),
            enum_columns=enum_columns,
            oversized_cells=oversized_cells,
            warnings=warnings,
            header_paths=header_paths,
        )

    def _repeating_labeled_context_rows(
        self,
        *,
        rows: list[list[Any]],
        headers: list[str],
        header_rel: int,
        data_end: int,
    ) -> tuple[dict[int, dict[str, Any]], list[str]]:
        """Find repeating key/value rows that introduce following numeric detail.

        Report-style workbooks often place entity metadata on a separate row,
        such as ``User Name: Alice`` followed by one or more metric rows. A row
        is treated as context only when the same label recurs in another block
        and numeric detail exists before the next labeled row. Those structural
        guards keep ordinary flat rows containing colons from being removed.
        """
        candidates: dict[int, dict[str, Any]] = {}
        for rel_idx in range(header_rel + 1, data_end + 1):
            visible = rows[rel_idx - 1][: len(headers)]
            if any(isinstance(value, (int, float)) for value in visible):
                continue
            pairs = self._extract_labeled_context_pairs(visible)
            if pairs:
                candidates[rel_idx] = pairs

        eligible: dict[int, dict[str, Any]] = {}
        candidate_rows = sorted(candidates)
        for index, rel_idx in enumerate(candidate_rows):
            next_rel = candidate_rows[index + 1] if index + 1 < len(candidate_rows) else data_end + 1
            following = rows[rel_idx : next_rel - 1]
            if any(
                any(isinstance(value, (int, float)) for value in row[: len(headers)])
                for row in following
            ):
                eligible[rel_idx] = candidates[rel_idx]

        counts: dict[str, int] = {}
        first_seen: list[str] = []
        for pairs in eligible.values():
            for column, value in pairs.items():
                if value in (None, ""):
                    continue
                counts[column] = counts.get(column, 0) + 1
                if column not in first_seen:
                    first_seen.append(column)
        repeated = [column for column in first_seen if counts.get(column, 0) >= 2]
        if not repeated:
            return {}, []

        context_rows = {
            rel_idx: {column: pairs.get(column) for column in repeated}
            for rel_idx, pairs in eligible.items()
            if any(pairs.get(column) not in (None, "") for column in repeated)
        }
        return context_rows, repeated

    def _extract_labeled_context_pairs(self, values: list[Any]) -> dict[str, Any]:
        label_runs: list[tuple[int, int, str]] = []
        index = 0
        while index < len(values):
            column = self._context_column_for_label(values[index])
            if not column:
                index += 1
                continue
            end = index
            while end + 1 < len(values) and self._context_column_for_label(values[end + 1]) == column:
                end += 1
            label_runs.append((index, end, column))
            index = end + 1

        pairs: dict[str, Any] = {}
        for run_index, (_, end, column) in enumerate(label_runs):
            next_start = label_runs[run_index + 1][0] if run_index + 1 < len(label_runs) else len(values)
            candidates = [
                value
                for value in values[end + 1 : next_start]
                if value not in (None, "") and self._context_column_for_label(value) is None
            ]
            pairs[column] = candidates[0] if candidates else None
        return pairs

    def _context_column_for_label(self, value: Any) -> str | None:
        if not isinstance(value, str):
            return None
        label = value.strip()
        if not label.endswith((":", "：")):
            return None
        label = label[:-1].strip().casefold()
        if not label or len(label) > 64:
            return None
        safe_label = self._safe_table_id(re.sub(r"\s+", "_", label)).strip("_")
        return f"_context_{safe_label}" if safe_label else None

    def _context_group_label(
        self,
        values: list[Any],
        headers: list[str],
        following_rows: list[list[Any]] | None = None,
        *,
        require_downstream_detail: bool = True,
    ) -> str | None:
        visible_values = values[: len(headers)]
        indexed_values = [
            (index, value)
            for index, value in enumerate(visible_values)
            if value not in (None, "")
            and not self._is_context_ignorable_column(headers[index] if index < len(headers) else "")
        ]
        if not indexed_values:
            return None

        first_index, first_value = indexed_values[0]
        if not isinstance(first_value, str):
            return None
        label = first_value.strip()
        if len(label) < 2 or self._contains_summary_keyword([label]):
            return None
        if any(isinstance(value, (int, float)) for _, value in indexed_values):
            return None

        normalized_values = {
            str(value).strip().casefold()
            for _, value in indexed_values
            if str(value).strip()
        }
        if (
            len(normalized_values) == 1
            and len(indexed_values) >= max(3, len(headers) // 2)
        ):
            return label

        first_header = headers[0].strip().lower() if headers else ""
        first_header_tokens = self._header_tokens(first_header)
        first_header_looks_like_id = bool(
            first_header_tokens.intersection({"serial", "no", "编号", "序号", "id"})
            or first_header == "#"
        )
        if (
            len(indexed_values) == 1
            and first_index == 0
            and len(headers) >= 3
            and first_header_looks_like_id
            and (
                not require_downstream_detail
                or self._has_downstream_detail_row(following_rows or [], headers)
            )
        ):
            return label
        return None

    def _header_tokens(self, header: str) -> set[str]:
        tokens = {
            token
            for token in re.split(r"[\s_/\-]+", str(header or "").strip().lower())
            if token
        }
        if str(header or "").strip() == "#":
            tokens.add("#")
        return tokens

    def _has_downstream_detail_row(
        self,
        following_rows: list[list[Any]],
        headers: list[str],
    ) -> bool:
        for row in following_rows:
            visible_values = row[: len(headers)]
            if not any(value not in (None, "") for value in visible_values):
                continue
            if self._context_group_label(
                visible_values,
                headers,
                following_rows=None,
                require_downstream_detail=False,
            ):
                continue
            if self._row_has_context_detail_values(visible_values, headers):
                return True
        return False

    def _row_has_context_detail_values(self, values: list[Any], headers: list[str]) -> bool:
        indexed_values = [
            value
            for index, value in enumerate(values[: len(headers)])
            if value not in (None, "")
            and not self._is_context_ignorable_column(headers[index] if index < len(headers) else "")
        ]
        if any(isinstance(value, (int, float)) for value in indexed_values):
            return True
        return len(indexed_values) >= 2

    def _is_context_ignorable_column(self, header: str) -> bool:
        normalized = str(header or "").strip().lower()
        return any(
            token in normalized
            for token in ("remark", "note", "comment", "备注", "说明")
        )

    def _column_metadata(
        self,
        dataframe: pd.DataFrame,
        column: str,
        enum_columns: dict[str, list[str]],
        header_paths: dict[str, list[str]],
        column_formats: dict[str, dict[str, Any]],
    ) -> dict[str, Any]:
        metadata: dict[str, Any] = {"name": column, "dtype": str(dataframe[column].dtype)}
        if column in enum_columns:
            metadata["enum_values"] = enum_columns[column]
        metadata["header_path"] = list(header_paths.get(column) or [column])
        metadata.update(column_formats.get(column) or {})
        return metadata

    def _build_column_format_metadata(
        self,
        *,
        worksheet: Any,
        rows: list[list[Any]],
        row_flags: dict[int, dict[str, Any]],
        headers: list[str],
        min_row: int,
        min_col: int,
        header_rel: int,
        data_end: int,
    ) -> dict[str, dict[str, Any]]:
        """Preserve Excel display scaling without changing stored values.

        A trailing comma in an Excel number format divides the displayed value
        by 1,000. Normalized tables retain the raw stored value, so downstream
        analysis needs this metadata to avoid applying the display scaling a
        second time merely because a header also says ``(000)``.
        """
        metadata: dict[str, dict[str, Any]] = {}
        for column_offset, header in enumerate(headers):
            formats: list[str] = []
            numeric_count = 0
            for rel_idx in range(header_rel + 1, data_end + 1):
                if row_flags.get(rel_idx, {}).get("exclude"):
                    continue
                row = rows[rel_idx - 1] if rel_idx - 1 < len(rows) else []
                value = row[column_offset] if column_offset < len(row) else None
                if isinstance(value, bool) or not isinstance(value, (int, float)):
                    continue
                numeric_count += 1
                cell = worksheet.cell(
                    row=min_row + rel_idx - 1,
                    column=min_col + column_offset,
                )
                number_format = str(cell.number_format or "General")
                if number_format.lower() != "general":
                    formats.append(number_format)

            if not formats:
                continue
            counts = Counter(formats)
            dominant_format, dominant_count = counts.most_common(1)[0]
            column_metadata: dict[str, Any] = {
                "excel_number_formats": [
                    number_format for number_format, _ in counts.most_common(3)
                ]
            }
            display_divisor = self._excel_display_divisor(dominant_format)
            if (
                display_divisor > 1
                and numeric_count > 0
                and dominant_count / numeric_count >= 0.8
            ):
                column_metadata["excel_display_divisor"] = display_divisor
            metadata[header] = column_metadata
        return metadata

    def _excel_display_divisor(self, number_format: str) -> int:
        """Return the display-only divisor encoded by trailing format commas."""
        section = str(number_format or "").split(";", 1)[0]
        cleaned: list[str] = []
        index = 0
        while index < len(section):
            char = section[index]
            if char == '"':
                index += 1
                while index < len(section) and section[index] != '"':
                    index += 1
            elif char == "[":
                index += 1
                while index < len(section) and section[index] != "]":
                    index += 1
            elif char in {"\\", "_", "*"}:
                index += 1
            else:
                cleaned.append(char)
            index += 1

        compact = "".join(cleaned)
        placeholders = [
            index for index, char in enumerate(compact) if char in {"0", "#", "?"}
        ]
        if not placeholders:
            return 1
        trailing = compact[placeholders[-1] + 1 :]
        comma_count = trailing.count(",")
        return 1000 ** comma_count

    def _build_header_paths(
        self,
        *,
        deduped_headers: list[str],
        raw_paths: list[list[str]],
        dataframe_columns: list[str],
    ) -> dict[str, list[str]]:
        """Assemble the per-column header lineage map.

        - Detected data headers keep their multi-level lineage.
        - Synthetic columns (`_context_group`, `_source_*`) get a single-level
          path equal to the column name so downstream code can treat every
          column uniformly.
        """
        header_paths: dict[str, list[str]] = {}
        for name, raw in zip(deduped_headers, raw_paths):
            if raw:
                header_paths[name] = list(raw)
            else:
                header_paths[name] = [name]
        for column in dataframe_columns:
            header_paths.setdefault(column, [column])
        return header_paths

    def _range_bound(self, value: Any, range_ref: str) -> int:
        if value is None:
            raise ValueError(f"Invalid table range: {range_ref}")
        return int(value)

    def _table_candidate_width(self, range_ref: str) -> int:
        min_col, _, max_col, _ = range_boundaries(range_ref)
        return int(max_col) - int(min_col) + 1

    def _detect_header_range(self, candidates: list[int], default: int) -> tuple[int, int]:
        """Find the start and end rows of a consecutive header block.

        Returns (header_start, header_end) as absolute row numbers.
        [1, 2] → (1, 2): two-level header.
        [3] → (3, 3): single header at row 3.
        [1, 2, 5] → (1, 2): first consecutive block only.
        """
        if not candidates:
            return (default, default)
        first = self._row_index(candidates[0], default)
        last = first
        for c in candidates[1:]:
            val = self._row_index(c, -1)
            if val == last + 1:
                last = val
            else:
                break
        return (first, last)

    def _row_index(self, value: Any, default: int) -> int:
        if value is None:
            return default
        return int(value)

    def _looks_like_headerless_continuation(
        self,
        *,
        rows: list[list[Any]],
        header_candidates: list[int],
        min_row: int,
        fallback_headers: list[str] | None,
    ) -> bool:
        if not rows or not fallback_headers or len(fallback_headers) != len(rows[0]):
            return False
        # If the first candidate is the region's first row, keep the normal
        # header path. Continuations usually have no header at the top, so a
        # later text-heavy data row may be mistaken for a header candidate.
        if header_candidates and int(header_candidates[0]) == min_row:
            return False
        first_row = rows[0]
        non_empty = [value for value in first_row if value not in (None, "")]
        if not non_empty:
            return False
        numeric_like = sum(isinstance(value, (int, float)) for value in non_empty)
        text_like = sum(isinstance(value, str) for value in non_empty)
        # A continuation's first row often mixes group labels with serial
        # numbers or measurements. A real header row should not contain data
        # numerics in this conservative detector.
        return numeric_like >= 1 and text_like >= 1

    def _cell_value(
        self,
        worksheet: Any,
        ws_values: Any,
        row: int,
        col: int,
        warnings: list[str],
    ) -> Any:
        raw_value = worksheet.cell(row=row, column=col).value
        computed_value = ws_values.cell(row=row, column=col).value
        if isinstance(raw_value, str) and raw_value.startswith("="):
            if computed_value is not None:
                return computed_value
            warnings.append(
                f"{worksheet.title}!{worksheet.cell(row=row, column=col).coordinate}: "
                "公式没有缓存计算值，保留公式文本"
            )
        return computed_value if computed_value is not None else raw_value

    def classify_rows(
        self,
        rows: list[list[Any]],
        header_row: int,
        data_end: int | None = None,
    ) -> dict[int, dict[str, Any]]:
        """Classify rows using 1-based row indexes.

        A summary keyword alone is not enough to exclude a row. It only becomes
        an auto-excluded summary row when it looks like an aggregate label near
        the table boundary and has numeric aggregate cells.
        """

        flags: dict[int, dict[str, Any]] = {}
        data_end = data_end or len(rows)
        for row_idx, values in enumerate(rows, start=1):
            non_empty = [value for value in values if value not in (None, "")]
            if row_idx < header_row:
                flags[row_idx] = {"kind": "title", "exclude": True, "confidence": 0.9}
                continue
            if row_idx > data_end:
                flags[row_idx] = {"kind": "footnote", "exclude": True, "confidence": 0.8}
                continue
            if not non_empty:
                flags[row_idx] = {"kind": "blank", "exclude": True, "confidence": 1.0}
                continue

            if self._looks_like_summary_row(values, row_idx=row_idx, data_end=data_end):
                flags[row_idx] = {"kind": "summary", "exclude": True, "confidence": 0.85}
                continue

            if self._contains_summary_keyword(values):
                flags[row_idx] = {
                    "kind": "possible_summary",
                    "exclude": False,
                    "confidence": 0.45,
                    "warning": "包含汇总关键词，但不满足自动排除条件",
                }
                continue

            flags[row_idx] = {"kind": "data", "exclude": False, "confidence": 0.9}
        return flags

    def _unmerge_and_fill(self, ws: Any, ws_values: Any) -> None:
        """拆分合并单元格，用左上角的值填充所有被合并的格子。

        ws: 以 data_only=False 打开的 worksheet（可操作 merged_cells）
        ws_values: 以 data_only=True 打开的 worksheet（读取公式计算后的值）
        """
        for merged_range in list(ws.merged_cells.ranges):
            min_r, min_c = merged_range.min_row, merged_range.min_col
            # 优先取计算值，公式单元格取 data_only=True 的值
            value = ws_values.cell(min_r, min_c).value
            if value is None:
                value = ws.cell(min_r, min_c).value
            ws.unmerge_cells(str(merged_range))
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    ws.cell(row, col).value = value

    def _merge_multi_level_headers(
        self, rows: list[list[Any]], header_row: int, start_rel: int = 1
    ) -> list[list[str]]:
        """多层表头合并为单层：'采购金额' + '计划' → '采购金额_计划'。

        直接修改 rows 中 header_row-1 位置的行（rows 是 0-indexed list，header_row 是 1-based）。
        start_rel: 表头起始行（1-based），默认为 1。只合并 start_rel..header_row 范围内的行。

        For sparse non-leaf header rows (e.g. `Landings into` at col B only,
        implicitly spanning B..K until the next label at col L), forward-fill
        each non-empty cell rightward until the next non-empty cell. The leaf
        (bottom) header row is left as-is so per-column labels stay specific.

        Returns per-column header lineage lists (top-level group → leaf column,
        empty levels dropped, adjacent duplicates dedupped). Callers use this to
        keep track of the original hierarchy after the flat merge destroys it.
        """
        if header_row <= 1:
            return []
        num_cols = len(rows[0]) if rows else 0
        leaf_rel = header_row - 1
        header_depth = header_row - start_rel + 1
        # The row immediately above the leaf is frequently a per-cell
        # annotation row (e.g. `%` markers on change columns, `Tonnes` on
        # tonnage columns). Forward-filling it contaminates non-annotated
        # columns with those annotations. Skip it when there are ≥3 header
        # rows; for shallower blocks the "penultimate" IS the top group row
        # and must be filled.
        skip_penultimate = header_depth >= 3
        for row_rel in range(start_rel - 1, leaf_rel):
            if skip_penultimate and row_rel == leaf_rel - 1:
                continue
            rows[row_rel] = self._forward_fill_sparse_header_row(
                rows[row_rel], num_cols
            )
        merged_headers = []
        header_paths: list[list[str]] = []
        for col_idx in range(num_cols):
            parts: list[str] = []
            for row_idx in range(start_rel - 1, header_row):  # start_rel-1..header_row-1
                val = rows[row_idx][col_idx] if col_idx < len(rows[row_idx]) else None
                if val is not None and str(val).strip():
                    parts.append(str(val).strip())
            # 去重（合并单元格填充后上下层可能相同）
            seen: list[str] = []
            for p in parts:
                if p not in seen:
                    seen.append(p)
            merged_headers.append("_".join(seen) if seen else f"列{col_idx + 1}")
            header_paths.append(seen)
        rows[header_row - 1] = merged_headers
        return header_paths

    def _forward_fill_sparse_header_row(
        self, row: list[Any], num_cols: int
    ) -> list[Any]:
        """Fill each non-empty label rightward until the next non-empty cell.

        Only applied to visibly sparse rows (< 60% density) so dense leaf rows
        or annotation rows keep their per-column specificity. Cells to the
        LEFT of the first non-empty value stay empty — some tables leave the
        row-label column blank at the top level.
        """
        effective_row = list(row)
        while len(effective_row) < num_cols:
            effective_row.append(None)
        filled_count = sum(
            1
            for value in effective_row[:num_cols]
            if value not in (None, "") and str(value).strip()
        )
        if num_cols == 0 or filled_count / num_cols >= 0.6:
            return effective_row
        result = list(effective_row)
        last_value: Any = None
        for i in range(num_cols):
            value = result[i]
            if value not in (None, "") and str(value).strip():
                last_value = value
            elif last_value is not None:
                result[i] = last_value
        return result

    def _detect_data_end(self, rows: list[list[Any]], header_row: int) -> int:
        """从底部向上扫描，过滤脚注行，返回最后一个有效数据行的 1-based 索引。"""
        footer_keywords = ("备注", "编制", "审核", "注：", "注:", "说明")
        total = len(rows)
        for row_idx in range(total, header_row, -1):
            values = rows[row_idx - 1]  # convert to 0-based
            non_empty = [v for v in values if v not in (None, "")]
            if not non_empty:
                continue
            row_text = " ".join(str(v) for v in non_empty)
            if any(kw in row_text for kw in footer_keywords):
                continue
            if len(non_empty) / max(len(values), 1) > 0.3:
                return row_idx
        return total

    def normalize_manifest_path(self, manifest: dict[str, Any]) -> str | None:
        return manifest.get("manifest_path") or manifest.get("path")

    def _detect_enum_columns(
        self,
        dataframe: pd.DataFrame,
        headers: list[str],
    ) -> dict[str, list[str]]:
        enum_columns: dict[str, list[str]] = {}
        for header in headers:
            if header not in dataframe.columns:
                continue
            series = dataframe[header]
            if (
                pd.api.types.is_numeric_dtype(series)
                or pd.api.types.is_datetime64_any_dtype(series)
            ):
                continue
            values = series.dropna().astype(str).str.strip()
            values = values[values != ""]
            if values.empty:
                continue
            unique_values = sorted(values.unique().tolist())
            unique_ratio = len(unique_values) / len(values)
            is_context_column = str(header).startswith("_context_")
            if len(unique_values) <= self.enum_max_unique_values and (
                is_context_column or unique_ratio < self.enum_max_unique_ratio
            ):
                enum_columns[header] = unique_values
        return enum_columns

    def _detect_oversized_cells(
        self,
        dataframe: pd.DataFrame,
        headers: list[str],
    ) -> list[dict[str, Any]]:
        oversized: list[dict[str, Any]] = []
        if "_source_row" not in dataframe.columns:
            return oversized
        for header in headers:
            if header not in dataframe.columns:
                continue
            for row_index, value in dataframe[header].items():
                if not isinstance(value, str) or len(value) <= self.oversized_cell_chars:
                    continue
                oversized.append(
                    {
                        "column": header,
                        "source_row": int(dataframe.at[row_index, "_source_row"]),
                        "length": len(value),
                        "preview": value[: self.oversized_cell_chars],
                    }
                )
        return oversized

    def _dedupe_headers(self, values: list[Any]) -> list[str]:
        counts: dict[str, int] = {}
        headers = []
        for index, value in enumerate(values, start=1):
            base = str(value).strip() if value not in (None, "") else f"列{index}"
            count = counts.get(base, 0)
            counts[base] = count + 1
            headers.append(base if count == 0 else f"{base}_{count + 1}")
        return headers

    def _safe_table_id(self, table_id: str) -> str:
        return "".join(char if char.isalnum() or char in ("-", "_") else "_" for char in table_id)

    def _contains_summary_keyword(self, values: list[Any]) -> bool:
        row_text = " ".join(str(value) for value in values if value not in (None, ""))
        return any(keyword in row_text for keyword in self.summary_keywords)

    def _looks_like_summary_row(
        self,
        values: list[Any],
        row_idx: int,
        data_end: int,
    ) -> bool:
        if not self._contains_summary_keyword(values):
            return False
        # Conservative default: only the final row is auto-excluded as a
        # summary. Intermediate group subtotals need stronger structural
        # detection later; until then they remain data with a warning.
        if row_idx != data_end:
            return False
        numeric_count = sum(isinstance(value, (int, float)) for value in values)
        text_values = [str(value).strip() for value in values if isinstance(value, str)]
        first_text = text_values[0] if text_values else ""
        label_like = any(keyword in first_text for keyword in self.summary_keywords)
        return label_like and numeric_count >= 1
