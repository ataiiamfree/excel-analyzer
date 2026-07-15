"""Shared validation for Excel uploads."""

from __future__ import annotations

import uuid
from pathlib import Path

import openpyxl
from fastapi import HTTPException, UploadFile


EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
UPLOAD_CHUNK_SIZE = 1024 * 1024


def save_validated_excel(
    upload: UploadFile,
    target: Path,
    *,
    max_size_mb: int,
) -> tuple[int, int, int]:
    """Save an upload and return size, sheet count, and row count.

    Validation happens before the caller creates or mutates persistent records.
    Partial and unreadable files are removed on every failure path.
    """
    suffix = Path(upload.filename or "").suffix.lower()
    if suffix not in EXCEL_EXTENSIONS:
        raise HTTPException(
            status_code=415,
            detail="目前请上传 .xlsx 或 .xlsm 文件",
        )

    target.parent.mkdir(parents=True, exist_ok=True)
    staging = target.with_name(
        f".{target.stem}.{uuid.uuid4().hex}.uploading{target.suffix}"
    )
    size_limit = max_size_mb * 1024 * 1024
    size = 0
    try:
        with staging.open("wb") as fh:
            while chunk := upload.file.read(UPLOAD_CHUNK_SIZE):
                size += len(chunk)
                if size > size_limit:
                    raise HTTPException(
                        status_code=413,
                        detail=f"文件超过 {max_size_mb}MB 上限",
                    )
                fh.write(chunk)

        workbook = openpyxl.load_workbook(
            staging,
            read_only=True,
            data_only=True,
            keep_vba=suffix == ".xlsm",
        )
        try:
            sheet_count = len(workbook.sheetnames)
            row_count = sum(workbook[sheet].max_row or 0 for sheet in workbook.sheetnames)
        finally:
            workbook.close()
        staging.replace(target)
    except HTTPException:
        staging.unlink(missing_ok=True)
        raise
    except Exception as exc:
        staging.unlink(missing_ok=True)
        raise HTTPException(
            status_code=422,
            detail="Excel 文件无法读取，请确认文件未损坏且未加密",
        ) from exc

    return size, sheet_count, row_count
