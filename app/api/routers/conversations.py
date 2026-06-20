"""Conversation REST endpoints."""

from __future__ import annotations

import shutil
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Annotated

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile

from app.api.deps import SessionRegistry, get_config, get_session_registry, get_store
from app.api.persistence.store import Store, utc_now_iso
from app.api.schemas import (
    ArtifactOut,
    ConversationGroup,
    ConversationListOut,
    ConversationOut,
    ConversationUpdate,
    MessageOut,
    UserRunRequest,
)
from app.api.ws.runner import artifact_out
from app.api.ws.runner import run_conversation_query
from app.config import Config
from app.workspace import Workspace

router = APIRouter(prefix="/api/conversations", tags=["conversations"])

EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}


def title_for_query(query: str | None, file_name: str | None) -> str:
    raw = (query or "").strip() or (file_name or "新的 Excel 分析")
    return raw[:24] + ("..." if len(raw) > 24 else "")


def group_label(created_at: str) -> str:
    dt = datetime.fromisoformat(created_at)
    today = datetime.now(timezone.utc).date()
    if dt.date() == today:
        return "今天"
    if (today - dt.date()).days == 1:
        return "昨天"
    if (today - dt.date()).days <= 7:
        return "本周"
    return "更早"


def grouped(conversations: list[dict]) -> ConversationListOut:
    order = ["今天", "昨天", "本周", "更早"]
    buckets: dict[str, list[ConversationOut]] = {label: [] for label in order}
    for item in conversations:
        buckets[group_label(item["created_at"])].append(ConversationOut(**item))
    return ConversationListOut(
        groups=[
            ConversationGroup(label=label, conversations=items)
            for label, items in buckets.items()
            if items
        ]
    )


def profile_workbook(path: Path) -> tuple[int | None, int | None]:
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_count = len(workbook.sheetnames)
        row_count = sum(workbook[sheet].max_row or 0 for sheet in workbook.sheetnames)
        workbook.close()
        return sheet_count, row_count
    except Exception:
        return None, None


def save_upload_to_workspace(upload: UploadFile, conversation_id: str, config: Config) -> tuple[str, int]:
    suffix = Path(upload.filename or "").suffix.lower()
    if suffix not in EXCEL_EXTENSIONS:
        raise HTTPException(status_code=400, detail="目前请上传 .xlsx 或 .xlsm 文件")
    workspace = Workspace(root=config.workspace_dir, task_id=conversation_id)
    target = Path(workspace.path) / "raw" / Path(upload.filename or "upload.xlsx").name
    size = 0
    with target.open("wb") as fh:
        while chunk := upload.file.read(1024 * 1024):
            size += len(chunk)
            fh.write(chunk)
    return str(target.resolve()), size


@router.get("", response_model=ConversationListOut)
async def list_conversations(store: Store = Depends(get_store)) -> ConversationListOut:
    return grouped(store.list_conversations())


@router.post("", response_model=ConversationOut)
async def create_conversation(
    file: Annotated[UploadFile, File()],
    query: Annotated[str | None, Form()] = None,
    store: Store = Depends(get_store),
    config: Config = Depends(get_config),
    sessions: SessionRegistry = Depends(get_session_registry),
) -> ConversationOut:
    conversation_id = uuid.uuid4().hex
    file_path, file_size = save_upload_to_workspace(file, conversation_id, config)
    sheet_count, row_count = profile_workbook(Path(file_path))
    row = store.create_conversation(
        conversation_id=conversation_id,
        title=title_for_query(query, file.filename),
        file_name=file.filename,
        file_size=file_size,
        local_file_path=file_path,
        sheet_count=sheet_count,
        row_count=row_count,
    )
    sessions.replace_file(conversation_id, file_path)
    return ConversationOut(**row)


@router.get("/{conversation_id}", response_model=ConversationOut)
async def get_conversation(conversation_id: str, store: Store = Depends(get_store)) -> ConversationOut:
    try:
        return ConversationOut(**store.get_conversation(conversation_id))
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="会话不存在") from exc


@router.patch("/{conversation_id}", response_model=ConversationOut)
async def update_conversation(
    conversation_id: str,
    update: ConversationUpdate,
    store: Store = Depends(get_store),
) -> ConversationOut:
    values = {}
    if update.title is not None:
        values["title"] = update.title
    if update.starred is not None:
        values["starred"] = update.starred
    if update.archived is not None:
        values["archived_at"] = utc_now_iso() if update.archived else None
    try:
        return ConversationOut(**store.update_conversation(conversation_id, **values))
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="会话不存在") from exc


@router.delete("/{conversation_id}", status_code=204)
async def delete_conversation(
    conversation_id: str,
    store: Store = Depends(get_store),
    config: Config = Depends(get_config),
) -> None:
    try:
        store.get_conversation(conversation_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="会话不存在") from exc
    store.delete_conversation(conversation_id)
    shutil.rmtree(Path(config.workspace_dir) / conversation_id, ignore_errors=True)


@router.get("/{conversation_id}/messages", response_model=list[MessageOut])
async def list_messages(conversation_id: str, store: Store = Depends(get_store)) -> list[MessageOut]:
    try:
        store.get_conversation(conversation_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="会话不存在") from exc
    return [MessageOut(**item) for item in store.list_messages(conversation_id)]


@router.get("/{conversation_id}/artifacts", response_model=list[ArtifactOut])
async def list_artifacts(conversation_id: str, store: Store = Depends(get_store)) -> list[ArtifactOut]:
    try:
        store.get_conversation(conversation_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="会话不存在") from exc
    return [ArtifactOut(**artifact_out(item)) for item in store.list_artifacts(conversation_id)]


@router.post("/{conversation_id}/files", response_model=ConversationOut)
async def replace_file(
    conversation_id: str,
    file: Annotated[UploadFile, File()],
    store: Store = Depends(get_store),
    config: Config = Depends(get_config),
    sessions: SessionRegistry = Depends(get_session_registry),
) -> ConversationOut:
    try:
        store.get_conversation(conversation_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="会话不存在") from exc
    file_path, file_size = save_upload_to_workspace(file, conversation_id, config)
    sheet_count, row_count = profile_workbook(Path(file_path))
    row = store.update_conversation(
        conversation_id,
        title=file.filename or "新的 Excel 分析",
        sheet_count=sheet_count,
        row_count=row_count,
    )
    with store._lock, store._conn:
        store._conn.execute(
            """
            UPDATE conversations
            SET file_name = ?, file_size = ?, local_file_path = ?, updated_at = ?
            WHERE id = ?
            """,
            (file.filename, file_size, file_path, utc_now_iso(), conversation_id),
        )
    sessions.replace_file(conversation_id, file_path)
    return ConversationOut(**store.get_conversation(conversation_id))


@router.post("/{conversation_id}/runs")
async def run_conversation(
    conversation_id: str,
    request: UserRunRequest,
    store: Store = Depends(get_store),
    config: Config = Depends(get_config),
    sessions: SessionRegistry = Depends(get_session_registry),
):
    try:
        store.get_conversation(conversation_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="会话不存在") from exc
    return await run_conversation_query(
        store=store,
        config=config,
        sessions=sessions,
        conversation_id=conversation_id,
        query=request.query,
    )
