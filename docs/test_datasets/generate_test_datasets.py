from __future__ import annotations

import json
import random
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


RNG = random.Random(20260526)
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "generated"


UNITS = ["罗湖局", "福田局", "南山局", "宝安局", "龙岗局", "龙华局", "坪山局", "光明局", "盐田局"]
TEAMS = ["一班", "二班", "三班", "重点客户组", "政企组"]
SUPPLIERS = ["华南电气", "深能物资", "鹏城智造", "前海设备", "粤海线缆", "星河自动化", "南粤科技", "港湾仪表"]
CATEGORIES = ["配电设备", "线缆", "仪表", "自动化", "备品备件", "施工服务"]
WAREHOUSES = ["宝安中心仓", "龙岗区域仓", "南山前置仓", "坪山备件仓"]
PRODUCT_CATS = ["智能终端", "配件", "线缆耗材", "计量设备", "低压电器"]
REGIONS = ["华南", "华东", "华北", "西南", "海外"]


def rand_date(start: date, end: date) -> date:
    return start + timedelta(days=RNG.randint(0, (end - start).days))


def maybe_date(value: date | None, string_rate: float = 0.18) -> date | str | None:
    if value is None:
        return None
    if RNG.random() < string_rate:
        fmt = RNG.choice(["%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"])
        return value.strftime(fmt)
    return value


def maybe_capacity(value: int | float) -> int | float | str | None:
    roll = RNG.random()
    if roll < 0.04:
        return None
    if roll < 0.15:
        return f"{value:,.0f}"
    if roll < 0.22:
        return f"{value:.0f}kVA"
    return value


def write_table(
    ws,
    title: str,
    headers: list[str],
    rows: list[list[object]],
    *,
    start_row: int = 3,
    note: str = "",
    freeze: bool = True,
) -> None:
    max_col = len(headers)
    if start_row > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(1, 1, title)
        ws.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
        ws.cell(1, 1).fill = PatternFill("solid", fgColor="1F4E78")
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
    if start_row > 2 and note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        ws.cell(2, 1, note)
        ws.cell(2, 1).font = Font(italic=True, color="666666")
    header_row = start_row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="A6A6A6"))
    for row_idx, row in enumerate(rows, start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)
    last_row = header_row + len(rows)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{last_row}"
    if freeze:
        ws.freeze_panes = f"A{header_row + 1}"
    for col in range(1, max_col + 1):
        values = [str(ws.cell(r, col).value or "") for r in range(header_row, min(last_row, header_row + 60) + 1)]
        width = min(max(max(len(v) for v in values) + 2, 10), 26)
        ws.column_dimensions[get_column_letter(col)].width = width


def save_workbook(wb: Workbook, filename: str) -> str:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / filename
    wb.save(path)
    return str(path)


def create_utility_workbook() -> dict:
    wb = Workbook()
    wb.remove(wb.active)

    in_progress_rows = []
    archive_rows = []
    duplicate_ids: list[str] = []
    for i in range(1, 261):
        order_id = f"YG2026{10000 + i}"
        if i <= 18:
            duplicate_ids.append(order_id)
        accepted = rand_date(date(2025, 1, 1), date(2026, 4, 28))
        power_date = None
        if RNG.random() < 0.45:
            power_date = accepted + timedelta(days=RNG.randint(20, 180))
            if power_date > date(2026, 4, 30):
                power_date = None
        status = RNG.choices(["运行", "归档", "中止", "作废"], [0.62, 0.25, 0.09, 0.04])[0]
        unit = RNG.choice(UNITS)
        cap = RNG.choice([315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500])
        if RNG.random() < 0.07:
            cap = -RNG.choice([100, 200, 315, 500])
        in_progress_rows.append([
            i, order_id, f"U{RNG.randint(100000, 999999)}", RNG.choice(["高压新装", "高压增容", "高压减容", "高压临时用电"]),
            status, unit, maybe_date(accepted), maybe_date(power_date), maybe_capacity(cap),
            RNG.choice(["是", "否"]), RNG.choice(["刘工", "王工", "李工", "陈工"]), RNG.choice(["", "客户延期", "资料补充", "'=VLOOKUP(B2,辅表!A:B,2,FALSE)"])
        ])

    for i in range(1, 361):
        if i <= len(duplicate_ids):
            order_id = duplicate_ids[i - 1]
        else:
            order_id = f"GD2026{20000 + i}"
        accepted = rand_date(date(2024, 12, 1), date(2026, 4, 26))
        power_date = accepted + timedelta(days=RNG.randint(15, 220))
        if power_date > date(2026, 4, 30):
            power_date = rand_date(date(2026, 1, 1), date(2026, 4, 30))
        status = RNG.choices(["归档", "中止"], [0.95, 0.05])[0]
        cap1 = RNG.choice([0, 315, 400, 630, 800, 1000, 1250, 1600, 2000])
        cap2 = RNG.choice([0, -100, -200, 100, 315])
        archive_rows.append([
            i, order_id, f"U{RNG.randint(100000, 999999)}", RNG.choice(["高压新装", "高压增容", "高压减容", "临时新装"]),
            status, RNG.choice(UNITS), maybe_date(accepted), maybe_date(power_date),
            maybe_capacity(cap1), maybe_capacity(cap2), RNG.choice(["张三", "李四", "赵六"]), RNG.choice(["", "跨年送电", "历史归档"])
        ])

    ws1 = wb.create_sheet("在途工单")
    write_table(ws1, "高压业扩在途工单台账", [
        "序号", "工单编号", "用户编号", "业务子类", "状态", "区局", "正式受理日期", "接火送电",
        "增减容量(kVA)", "是否临时", "客户经理", "备注"
    ], in_progress_rows, start_row=4, note="注意：第 1、2 个 sheet 才是本题数据源；容量可能是数字、文本或带 kVA 后缀。")

    ws2 = wb.create_sheet("已送电归档")
    write_table(ws2, "高压业扩已送电归档", [
        "序号", "工单号", "用户编号", "业务类别", "工单状态", "供电单位", "受理时间", "送电日期",
        "新增容量", "退补容量", "联系人", "备注"
    ], archive_rows, start_row=3, note="工单号与在途表可能重复，重复时应按工单号去重。")

    helper = wb.create_sheet("辅助-口径")
    write_table(helper, "口径说明", ["字段", "说明"], [
        ["报装时间", "在途工单.正式受理日期 / 已送电归档.受理时间"],
        ["送电时间", "在途工单.接火送电 / 已送电归档.送电日期"],
        ["容量", "在途工单.增减容量(kVA)；归档表为 新增容量 + 退补容量"],
        ["排除", "状态为 中止/作废 的工单不计入"],
    ], start_row=2)
    helper.sheet_state = "hidden"
    path = save_workbook(wb, "测试01_业扩报装送电统计_多Sheet口径.xlsx")
    return {
        "file": path,
        "question": "针对上传文件做业扩报装/送电统计。仅分析第一个、第二个 sheet，两个 sheet 分开统计后再合并；报装时间分别对应“正式受理日期/受理时间”，送电时间分别对应“接火送电/送电日期”；容量字段在第一个 sheet 为“增减容量(kVA)”，第二个 sheet 为“新增容量+退补容量”。要求：1）排除状态为中止/作废的数据；2）按工单编号/工单号去重，重复时只保留送电日期或归档状态更完整的一条；3）“本月”按数据中最新报装或送电日期所在月份，不按系统当前月份；4）输出本月报装单数、本月报装容量、本月送电单数、本月送电容量、全年报装单数及同比、全年报装容量及同比、全年送电单数及同比、全年送电容量及同比；5）输出今年和去年同期用于核对的明细记录与分表汇总，最后结果用表格输出。",
        "tests": ["前两张 sheet 才是数据源", "字段同义映射", "容量文本清洗", "去重", "按数据最新月而非系统月", "过程明细输出"],
    }


def create_overload_workbook() -> dict:
    wb = Workbook()
    wb.remove(wb.active)

    user_rows = []
    totals = {"公变": 0, "专变": 0, "低压用户": 0}
    for idx, unit in enumerate(UNITS, start=1):
        public = RNG.randint(18000, 98000)
        special = RNG.randint(1200, 8600)
        low = RNG.randint(80000, 360000)
        totals["公变"] += public
        totals["专变"] += special
        totals["低压用户"] += low
        user_rows.append([idx, unit, public, special, low, RNG.choice(["东片", "西片", "中心片"])])
    user_rows.append(["", "深圳局", totals["公变"], totals["专变"], totals["低压用户"], "汇总行"])

    meter_pool = [f"MP{100000 + i}" for i in range(500)]
    current_rows = []
    prev_rows = []
    all_rows = []
    for i in range(1, 241):
        unit = RNG.choice(UNITS)
        meter = RNG.choice(meter_pool)
        typ = RNG.choice(["过", "超"])
        repeat = RNG.choice([0, 0, 0, 1, 2])
        current_rows.append([f"C{i:04d}", f"用户{i}", f"T{RNG.randint(1000,9999)}", unit, unit + "供电所", meter, typ, repeat, maybe_date(rand_date(date(2026, 5, 1), date(2026, 5, 25)))])
    for i in range(1, 210):
        unit = RNG.choice(UNITS)
        meter = RNG.choice(meter_pool)
        typ = RNG.choice(["过", "超"])
        prev_rows.append([f"P{i:04d}", f"上月用户{i}", f"T{RNG.randint(1000,9999)}", unit, unit + "供电所", meter, typ, maybe_date(rand_date(date(2026, 4, 1), date(2026, 4, 30)))])
    period_cols = ["总重过载类型", "尖期重过载类型", "尖峰期重过载类型", "平期重过载类型", "谷期重过载类型"]
    for i in range(1, 900):
        unit = RNG.choice(UNITS)
        meter = RNG.choice(meter_pool + [None, ""])
        vals = [RNG.choices(["正常", "过载", "超载", ""], [0.55, 0.22, 0.18, 0.05])[0] for _ in period_cols]
        all_rows.append([maybe_date(rand_date(date(2026, 5, 1), date(2026, 5, 25))), f"CU{i:05d}", f"用户{i}", unit, unit + "供电所", meter, *vals, RNG.choice(["配变", "专变"])])

    ws = wb.create_sheet("用户数")
    write_table(ws, "各单位用户数", ["序号", "单位", "公变", "专变", "低压用户", "片区"], user_rows, start_row=2, note="最后一行深圳局为汇总行，分单位统计时不要重复计入。")
    ws = wb.create_sheet("过超去重")
    write_table(ws, "本月过超去重记录", ["记录ID", "用户名称", "资产编号", "供电区局", "供电单位", "计量点编号", "类型", "重复计算", "发生日期"], current_rows, start_row=1)
    ws = wb.create_sheet("上月过超去重")
    write_table(ws, "上月过超去重记录", ["记录ID", "用户名称", "资产编号", "供电区局", "供电单位", "计量点编号", "类型", "发生日期"], prev_rows, start_row=3, note="表头不在第一行，测试自动识别。")
    ws = wb.create_sheet("过超总")
    write_table(ws, "本月过超总明细", ["数据日期", "用户编号", "用户名称", "供电区局", "供电单位", "计量点编号", *period_cols, "用户类型"], all_rows, start_row=2)
    path = save_workbook(wb, "测试02_配变重过载统计_去重汇总.xlsx")
    return {
        "file": path,
        "question": "根据上传表格生成三个统计表，每个表最后一行以“深圳局”为单位汇总各区局。表一“各单位过超信息统计表”：数据源为“过超去重”和“上月过超去重”，字段为单位、上月、本月、环比、重复、新增，其中重复为本月重复计算字段大于0的数量，新增=本月-重复。表二“各单位重过载记录统计表”：数据源为“用户数”和“过超去重”，单位取用户数中所有区局但排除已有汇总行，字段为单位、公变、超载、过载、小计、超负荷占比（小计/公变，按千分比输出）。表三“各单位超负荷用户数统计表”：数据源为“过超总”，对总/尖/峰/平/谷五个重过载类型字段分别统计；每个字段先筛选本单位且类型为过载，按非空计量点编号去重计数，再筛选类型为超载按同样规则计数，二者相加。输出计算过程和可核对明细。",
        "tests": ["汇总行排除再重算", "跨 sheet 联表", "重复计算", "非空去重计数", "千分比", "表头偏移"],
    }


def create_procurement_workbook() -> dict:
    wb = Workbook()
    wb.remove(wb.active)
    orders = []
    receipts = []
    last_year = []
    for i in range(1, 521):
        po = f"PO26{100000 + i}"
        supplier = RNG.choice(SUPPLIERS)
        cat = RNG.choice(CATEGORIES)
        order_date = rand_date(date(2026, 1, 2), date(2026, 5, 24))
        promise = order_date + timedelta(days=RNG.randint(7, 60))
        qty = RNG.randint(5, 500)
        price = round(RNG.uniform(80, 8600), 2)
        status = RNG.choices(["有效", "已取消", "关闭"], [0.82, 0.08, 0.10])[0]
        currency = RNG.choices(["CNY", "USD"], [0.9, 0.1])[0]
        orders.append([po, RNG.randint(1, 3), supplier, f"S{SUPPLIERS.index(supplier)+1:03d}", cat, RNG.choice(["运维部", "市场部", "工程部", "信息中心"]), maybe_date(order_date), maybe_date(promise), qty, price, currency, status, RNG.choice(["是", "否"]), ""])
        received_total = 0
        for r in range(RNG.randint(0, 3)):
            if received_total >= qty:
                break
            got = min(qty - received_total, RNG.randint(1, max(2, qty // 2)))
            received_total += got
            receipt_date = promise + timedelta(days=RNG.randint(-10, 25))
            receipts.append([f"RC{po[-5:]}-{r+1}", po, got, maybe_date(receipt_date), RNG.choice(["合格", "合格", "待复检", "不合格"]), RNG.choice(WAREHOUSES), round(got * price, 2)])
    for i in range(1, 430):
        supplier = RNG.choice(SUPPLIERS)
        cat = RNG.choice(CATEGORIES)
        order_date = rand_date(date(2025, 1, 2), date(2025, 5, 24))
        qty = RNG.randint(5, 480)
        price = round(RNG.uniform(70, 7600), 2)
        last_year.append([f"PO25{100000+i}", supplier, f"S{SUPPLIERS.index(supplier)+1:03d}", cat, maybe_date(order_date), qty, price, RNG.choice(["有效", "关闭"])])

    ws = wb.create_sheet("采购订单")
    write_table(ws, "采购订单明细", ["PO号", "行号", "供应商名称", "供应商编码", "品类", "需求部门", "下单日期", "承诺到货日期", "采购数量", "含税单价", "币种", "订单状态", "是否紧急", "备注"], orders, start_row=4, note="订单可能分批到货，金额=数量*含税单价；取消订单排除。")
    ws = wb.create_sheet("到货验收")
    write_table(ws, "到货验收明细", ["验收单号", "PO号", "到货数量", "验收日期", "质检结果", "入库仓库", "到货金额"], receipts, start_row=2)
    ws = wb.create_sheet("供应商主数据")
    vendor_rows = [[f"S{i+1:03d}", s, RNG.choice(["A", "B", "C"]), RNG.choice(REGIONS), RNG.choice(["战略", "普通", "观察"])] for i, s in enumerate(SUPPLIERS)]
    write_table(ws, "供应商主数据", ["供应商编码", "标准供应商", "等级", "区域", "合作类型"], vendor_rows, start_row=1)
    ws = wb.create_sheet("去年同期订单")
    write_table(ws, "去年同期采购订单", ["PO号", "供应商名称", "供应商编码", "品类", "下单日期", "采购数量", "含税单价", "订单状态"], last_year, start_row=3)
    ws = wb.create_sheet("品类映射")
    write_table(ws, "品类映射", ["品类", "大类", "是否关键物资"], [[c, RNG.choice(["设备", "材料", "服务"]), RNG.choice(["是", "否"])] for c in CATEGORIES], start_row=1)
    path = save_workbook(wb, "测试03_采购履约与供应商绩效.xlsx")
    return {
        "file": path,
        "question": "请基于采购订单、到货验收、供应商主数据和去年同期订单做供应商履约分析。要求以数据最新下单日期所在月份为本月，输出按供应商和品类两个维度的统计表：本月下单金额、本月到货金额、全年采购金额、全年采购金额同比、订单数、已到货订单数、准时交付率、平均交付天数、超期未到货金额、质检不合格金额。取消订单不计入；订单金额=采购数量*含税单价，USD按汇率7.2折算人民币；同一个PO可能多次到货，要按PO汇总到货数量和最晚验收日期判断是否准时。最后输出超期未到货明细、去年同期基数明细和供应商Top/Bottom清单。",
        "tests": ["一对多到货汇总", "取消订单排除", "币种折算", "同比", "供应商主数据映射", "超期未到货"],
    }


def create_sales_workbook() -> dict:
    wb = Workbook()
    wb.remove(wb.active)
    opp_rows = []
    payments = []
    snapshot = []
    customers = []
    customer_ids = [f"CUST{i:04d}" for i in range(1, 170)]
    for cid in customer_ids:
        customers.append([cid, f"客户{cid[-4:]}", RNG.choice(REGIONS), RNG.choice(["制造", "金融", "互联网", "政企", "能源"]), RNG.choice(["KA", "A", "B", "C"]), round(RNG.uniform(5, 300), 2), round(RNG.uniform(4, 360), 2)])
    for i in range(1, 481):
        opp = f"OPP26{i:05d}"
        cid = RNG.choice(customer_ids)
        create_date = rand_date(date(2025, 12, 1), date(2026, 5, 25))
        expected = create_date + timedelta(days=RNG.randint(15, 120))
        actual = expected + timedelta(days=RNG.randint(-20, 45)) if RNG.random() < 0.55 else None
        stage = RNG.choices(["线索", "方案", "报价", "谈判", "赢单", "输单"], [0.10, 0.18, 0.22, 0.20, 0.20, 0.10])[0]
        amount = round(RNG.uniform(3, 420), 2)
        currency = RNG.choices(["CNY", "USD"], [0.88, 0.12])[0]
        renew = RNG.choice(["是", "否"])
        lost = RNG.choice(["价格", "竞品", "预算取消", "需求变化", ""]) if stage == "输单" else ""
        opp_rows.append([opp, cid, f"客户{cid[-4:]}", RNG.choice(REGIONS), RNG.choice(TEAMS), RNG.choice(["Alice", "Bob", "Cindy", "David", "Eva"]), stage, maybe_date(create_date), maybe_date(expected), maybe_date(actual), amount, currency, renew, lost, maybe_date(rand_date(date(2026, 4, 1), date(2026, 5, 25)))])
        if stage == "赢单":
            for p in range(RNG.randint(1, 3)):
                pay_date = actual + timedelta(days=RNG.randint(0, 45)) if actual else rand_date(date(2026, 1, 1), date(2026, 5, 25))
                payments.append([f"PAY{opp[-5:]}-{p+1}", opp, maybe_date(pay_date), round(amount * RNG.uniform(0.2, 0.8), 2), currency, RNG.choice(["已确认", "已确认", "待确认"])])
        if RNG.random() < 0.65:
            snapshot.append([opp, RNG.choice(["线索", "方案", "报价", "谈判"]), round(amount * RNG.uniform(0.8, 1.2), 2), maybe_date(date(2026, 4, 30))])

    ws = wb.create_sheet("商机明细")
    write_table(ws, "销售商机明细", ["商机ID", "客户ID", "客户名称", "区域", "销售团队", "销售", "阶段", "创建日期", "预计关闭日期", "实际关闭日期", "合同金额", "币种", "是否续费", "流失原因", "最后更新时间"], opp_rows, start_row=3, note="金额单位：万元；USD按汇率表折算。")
    ws = wb.create_sheet("合同回款")
    write_table(ws, "合同回款流水", ["回款单号", "商机ID", "回款日期", "回款金额", "币种", "确认状态"], payments, start_row=2)
    ws = wb.create_sheet("上月商机快照")
    write_table(ws, "上月商机快照", ["商机ID", "上月阶段", "上月预计金额", "快照日期"], snapshot, start_row=1)
    ws = wb.create_sheet("客户主数据")
    write_table(ws, "客户主数据", ["客户ID", "客户名称", "区域", "行业", "客户等级", "上年ARR", "当前ARR"], customers, start_row=2)
    ws = wb.create_sheet("汇率表")
    write_table(ws, "汇率表", ["币种", "人民币汇率"], [["CNY", 1], ["USD", 7.2]], start_row=1)
    path = save_workbook(wb, "测试04_销售续费漏斗与回款.xlsx")
    return {
        "file": path,
        "question": "请分析销售商机、回款和续费情况。要求按区域、销售团队分别输出：本月新增商机数和金额、本月赢单数和金额、本月输单数、期初pipeline金额、期末pipeline金额、pipeline净变化、阶段转化率、确认回款金额、回款达成率、续费客户数、续费金额、GRR和NRR。金额统一折算人民币，USD按汇率表。期初pipeline来自“上月商机快照”，期末pipeline来自当前商机中非赢单/非输单阶段；本月以数据最新更新时间所在月份为准；待确认回款不计入。输出输单原因Top5、回款明细和续费客户核对明细。",
        "tests": ["快照对比", "多币种", "漏斗阶段", "回款确认状态", "续费/ARR", "Top原因"],
    }


def create_inventory_workbook() -> dict:
    wb = Workbook()
    wb.remove(wb.active)
    sku_rows = []
    skus = [f"SKU{i:04d}" for i in range(1, 140)]
    for sku in skus:
        sku_rows.append([sku, f"物料{sku[-4:]}", RNG.choice(PRODUCT_CATS), RNG.choice(["自有", "代工", "进口"]), RNG.choice(["个", "箱", "米"]), RNG.randint(20, 500), RNG.randint(7, 45), RNG.choice(["是", "否"])])
    stock_rows = []
    sales_rows = []
    flow_rows = []
    last_stockout = []
    for wh in WAREHOUSES:
        for sku in RNG.sample(skus, 90):
            available = RNG.randint(0, 1200)
            transit = RNG.randint(0, 500)
            bad = RNG.randint(0, 40)
            stock_rows.append([date(2026, 5, 25), wh, sku, available, transit, bad, f"B{RNG.randint(100,999)}"])
            if available < 30 and RNG.random() < 0.7:
                last_stockout.append([date(2026, 4, 30), wh, sku, RNG.randint(1, 12)])
    for i in range(1, 1800):
        wh = RNG.choice(WAREHOUSES)
        sku = RNG.choice(skus)
        dt = rand_date(date(2026, 4, 1), date(2026, 5, 25))
        qty = RNG.randint(1, 80)
        doc_type = RNG.choices(["销售出库", "销售退货", "样品出库"], [0.82, 0.10, 0.08])[0]
        if doc_type == "销售退货":
            qty = -qty
        sales_rows.append([maybe_date(dt), f"SO{i:06d}", wh, sku, qty, doc_type, RNG.choice(["直营网店", "经销商", "项目", "售后"])])
    for i in range(1, 2200):
        wh = RNG.choice(WAREHOUSES)
        sku = RNG.choice(skus)
        dt = rand_date(date(2026, 1, 1), date(2026, 5, 25))
        typ = RNG.choice(["采购入库", "销售出库", "调拨入", "调拨出", "盘点调整"])
        sign = -1 if typ in ["销售出库", "调拨出"] else 1
        qty = sign * RNG.randint(1, 180)
        flow_rows.append([maybe_date(dt), f"INV{i:06d}", wh, sku, qty, typ, RNG.choice(["良品", "良品", "不良品"])])
    ws = wb.create_sheet("SKU主数据")
    write_table(ws, "SKU主数据", ["SKU", "品名", "类别", "品牌", "单位", "安全库存", "采购提前期(天)", "是否重点SKU"], sku_rows, start_row=1)
    ws = wb.create_sheet("库存快照")
    write_table(ws, "库存快照", ["快照日期", "仓库", "SKU", "可用库存", "在途库存", "不良品库存", "批次"], stock_rows, start_row=3)
    ws = wb.create_sheet("销售出库")
    write_table(ws, "销售出库明细", ["日期", "单据号", "仓库", "SKU", "数量", "单据类型", "渠道"], sales_rows, start_row=2)
    ws = wb.create_sheet("库存流水")
    write_table(ws, "库存流水", ["日期", "单据号", "仓库", "SKU", "变动数量", "类型", "库存状态"], flow_rows, start_row=2)
    ws = wb.create_sheet("上月缺货记录")
    write_table(ws, "上月缺货记录", ["统计日期", "仓库", "SKU", "缺货天数"], last_stockout, start_row=1)
    path = save_workbook(wb, "测试05_库存缺货与补货分析.xlsx")
    return {
        "file": path,
        "question": "请基于SKU主数据、库存快照、销售出库、库存流水和上月缺货记录生成库存健康分析。按仓库和类别输出：本月净销量、日均销量、当前可用库存、在途库存、可售天数、低于安全库存SKU数、缺货天数、新增缺货SKU数、重复缺货SKU数、安全库存缺口、建议补货量。销售退货数量为负数，应抵减销量；样品出库不计入销售；不良品库存不能计入可用库存。缺货定义为某SKU某日推算库存<=0；本月以销售出库中最新日期所在月份为准。输出缺货SKU明细、建议补货明细和仓库汇总表。",
        "tests": ["负数退货", "样品排除", "库存流水推算缺货", "上月重复缺货", "安全库存缺口", "仓库/类别多维汇总"],
    }


def create_ar_workbook() -> dict:
    wb = Workbook()
    wb.remove(wb.active)
    customers = [f"客户{i:04d}" for i in range(1, 150)]
    customer_rows = [[c, RNG.choice(REGIONS), RNG.choice(["制造", "金融", "能源", "政企", "零售"]), RNG.choice(["KA", "A", "B", "C"]), RNG.randint(30, 120)] for c in customers]
    invoices = []
    payments = []
    writeoffs = []
    ly_invoices = []
    for i in range(1, 700):
        inv = f"INV26{i:06d}"
        customer = RNG.choice(customers)
        inv_date = rand_date(date(2025, 11, 1), date(2026, 5, 24))
        term = RNG.choice([30, 45, 60, 90])
        due = inv_date + timedelta(days=term)
        amount = round(RNG.uniform(0.5, 260), 2)
        status = RNG.choices(["有效", "红冲", "作废"], [0.86, 0.08, 0.06])[0]
        currency = RNG.choices(["CNY", "USD"], [0.92, 0.08])[0]
        invoices.append([inv, customer, RNG.choice(["配网事业部", "营销事业部", "数字化事业部"]), maybe_date(inv_date), maybe_date(due), amount, round(amount * 0.06, 2), currency, status, term, RNG.choice(["张经理", "李经理", "陈经理"])])
        paid_total = 0
        for p in range(RNG.randint(0, 3)):
            pay = round(min(amount - paid_total, amount * RNG.uniform(0.2, 0.7)), 2)
            if pay <= 0:
                break
            paid_total += pay
            payments.append([f"REC{i:06d}-{p+1}", inv, customer, maybe_date(inv_date + timedelta(days=RNG.randint(10, 140))), pay, currency, RNG.choice(["已核销", "已核销", "待认领"])])
        if RNG.random() < 0.04:
            writeoffs.append([f"WO{i:05d}", inv, customer, maybe_date(rand_date(date(2026, 1, 1), date(2026, 5, 24))), round(amount * RNG.uniform(0.1, 0.5), 2), RNG.choice(["坏账审批", "争议折让"])])
    for i in range(1, 580):
        inv_date = rand_date(date(2025, 1, 1), date(2025, 5, 24))
        amount = round(RNG.uniform(0.5, 240), 2)
        ly_invoices.append([f"INV25{i:06d}", RNG.choice(customers), RNG.choice(["配网事业部", "营销事业部", "数字化事业部"]), maybe_date(inv_date), amount, RNG.choice(["有效", "红冲"])])
    ws = wb.create_sheet("发票明细")
    write_table(ws, "发票明细", ["发票号", "客户名称", "事业部", "开票日期", "到期日", "不含税金额", "税额", "币种", "发票状态", "账期天数", "销售经理"], invoices, start_row=4, note="金额单位：万元；红冲/作废按口径处理。")
    ws = wb.create_sheet("收款明细")
    write_table(ws, "收款明细", ["收款单号", "发票号", "客户名称", "收款日期", "收款金额", "币种", "核销状态"], payments, start_row=2)
    ws = wb.create_sheet("客户主数据")
    write_table(ws, "客户主数据", ["客户名称", "区域", "行业", "客户等级", "信用账期"], customer_rows, start_row=1)
    ws = wb.create_sheet("坏账核销")
    write_table(ws, "坏账核销", ["核销单号", "发票号", "客户名称", "核销日期", "核销金额", "原因"], writeoffs, start_row=2)
    ws = wb.create_sheet("去年同期发票")
    write_table(ws, "去年同期发票", ["发票号", "客户名称", "事业部", "开票日期", "不含税金额", "发票状态"], ly_invoices, start_row=3)
    ws = wb.create_sheet("汇率")
    write_table(ws, "汇率", ["币种", "人民币汇率"], [["CNY", 1], ["USD", 7.2]], start_row=1)
    path = save_workbook(wb, "测试06_应收账款账龄与回款核销.xlsx")
    return {
        "file": path,
        "question": "请做应收账款账龄和回款分析。以发票明细、收款明细、客户主数据、坏账核销、去年同期发票为数据源，金额统一折算人民币，USD按汇率7.2。统计口径：作废发票排除，红冲发票金额按负数处理；只计已核销收款，待认领不计入已收；净应收=发票含税金额-已核销收款-坏账核销。以数据中最新收款/开票日期作为分析日，输出按事业部、区域、客户等级的本月开票金额、本月回款金额、回款率、期末净应收、逾期金额、账龄0-30/31-60/61-90/90+金额、逾期金额同比、Top10逾期客户。输出每张发票的核销过程明细，便于核对。",
        "tests": ["部分收款核销", "红冲负数", "作废排除", "坏账核销", "账龄桶", "同比和TopN"],
    }


def write_questions_manifest(cases: list[dict]) -> None:
    lines = [
        "# ChatExcel 压测数据集与测试问题",
        "",
        "这些文件是专门用来测 Excel agent 的复杂业务表。每个问题都可以直接复制到网页里测试。",
        "",
    ]
    manifest = []
    for idx, case in enumerate(cases, start=1):
        rel = Path(case["file"]).relative_to(BASE_DIR)
        lines.extend([
            f"## {idx}. {Path(case['file']).name}",
            "",
            f"文件：`{rel}`",
            "",
            "测试问题：",
            "",
            case["question"],
            "",
            "主要测试点：",
            "",
            *[f"- {item}" for item in case["tests"]],
            "",
        ])
        manifest.append({
            "file": str(rel),
            "question": case["question"],
            "tests": case["tests"],
        })
    (BASE_DIR / "测试问题清单.md").write_text("\n".join(lines), encoding="utf-8")
    (BASE_DIR / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def validate_outputs(cases: list[dict]) -> None:
    for case in cases:
        path = Path(case["file"])
        wb = load_workbook(path, read_only=True, data_only=False)
        if len(wb.sheetnames) < 3:
            raise RuntimeError(f"{path.name} sheet 数过少")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2 or ws.max_column < 2:
                raise RuntimeError(f"{path.name}/{sheet_name} 内容过少")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    cases = [
        create_utility_workbook(),
        create_overload_workbook(),
        create_procurement_workbook(),
        create_sales_workbook(),
        create_inventory_workbook(),
        create_ar_workbook(),
    ]
    validate_outputs(cases)
    write_questions_manifest(cases)
    print(json.dumps({"output_dir": str(OUTPUT_DIR), "files": [Path(c["file"]).name for c in cases]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
