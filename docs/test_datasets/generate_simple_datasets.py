"""生成简单测试数据集，覆盖多种分析场景。

每个数据集刻意保持小规模（10-30行），数据干净无脏值，
便于快速验证不同类型的分析能力。
"""

import datetime
import random
from pathlib import Path

import openpyxl

OUTPUT_DIR = Path(__file__).parent / "simple"
OUTPUT_DIR.mkdir(exist_ok=True)

random.seed(42)


def _save(wb, name):
    path = OUTPUT_DIR / name
    wb.save(path)
    print(f"  ✓ {name}")


# ── 1. 门店月度销售额 ──────────────────────────────────────
def gen_01():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "月度销售"
    ws.append(["门店", "月份", "销售额(万元)", "客单价(元)", "客流量"])
    stores = ["北京旗舰店", "上海南京路店", "广州天河店", "成都春熙路店", "杭州西湖店"]
    for store in stores:
        for m in range(1, 7):
            sales = round(random.uniform(50, 200), 1)
            price = round(random.uniform(80, 300), 0)
            traffic = random.randint(3000, 15000)
            ws.append([store, f"2025-{m:02d}", sales, price, traffic])
    _save(wb, "01_门店月度销售.xlsx")


# ── 2. 员工花名册 ──────────────────────────────────────────
def gen_02():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工信息"
    ws.append(["姓名", "部门", "职级", "入职日期", "月薪(元)", "性别"])
    depts = ["技术部", "市场部", "财务部", "人事部", "运营部"]
    levels = ["P5", "P6", "P7", "P8", "P9"]
    names = ["张伟", "李娜", "王芳", "刘洋", "陈明", "杨柳", "赵磊",
             "黄丽", "周杰", "吴敏", "郑浩", "孙婷", "马超", "朱莉", "徐刚"]
    for i, name in enumerate(names):
        dept = depts[i % len(depts)]
        level = levels[min(i // 3, 4)]
        year = random.randint(2018, 2024)
        month = random.randint(1, 12)
        entry = datetime.date(year, month, 1)
        salary = 8000 + (levels.index(level)) * 5000 + random.randint(-2000, 2000)
        gender = "女" if i % 3 == 1 else "男"
        ws.append([name, dept, level, entry, salary, gender])
    _save(wb, "02_员工花名册.xlsx")


# ── 3. 考试成绩单 ──────────────────────────────────────────
def gen_03():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "期末成绩"
    ws.append(["学号", "姓名", "班级", "语文", "数学", "英语", "物理", "化学"])
    classes = ["高一(1)班", "高一(2)班", "高一(3)班"]
    for cls in classes:
        for j in range(1, 11):
            sid = f"{classes.index(cls)+1}{j:03d}"
            name = f"学生{sid}"
            scores = [random.randint(40, 100) for _ in range(5)]
            ws.append([sid, name, cls, *scores])
    _save(wb, "03_考试成绩单.xlsx")


# ── 4. 产品库存表 ──────────────────────────────────────────
def gen_04():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存明细"
    ws.append(["SKU", "品名", "类别", "库存数量", "安全库存", "单价(元)", "仓库"])
    categories = ["电子产品", "办公用品", "食品饮料", "日用百货"]
    warehouses = ["华东仓", "华南仓", "华北仓"]
    items = [
        ("E001", "蓝牙耳机", "电子产品"), ("E002", "充电宝", "电子产品"),
        ("E003", "数据线", "电子产品"), ("O001", "A4打印纸", "办公用品"),
        ("O002", "签字笔", "办公用品"), ("O003", "文件夹", "办公用品"),
        ("F001", "矿泉水", "食品饮料"), ("F002", "咖啡豆", "食品饮料"),
        ("D001", "纸巾", "日用百货"), ("D002", "洗手液", "日用百货"),
        ("D003", "垃圾袋", "日用百货"), ("E004", "鼠标", "电子产品"),
    ]
    for sku, name, cat in items:
        stock = random.randint(0, 500)
        safety = random.randint(50, 200)
        price = round(random.uniform(5, 500), 2)
        wh = random.choice(warehouses)
        ws.append([sku, name, cat, stock, safety, price, wh])
    _save(wb, "04_产品库存表.xlsx")


# ── 5. 项目进度跟踪 ──────────────────────────────────────────
def gen_05():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目清单"
    ws.append(["项目编号", "项目名称", "负责人", "计划开始", "计划结束",
               "实际开始", "实际结束", "状态", "预算(万元)"])
    projects = [
        ("P001", "官网改版", "张伟"), ("P002", "APP 2.0", "李娜"),
        ("P003", "数据中台", "王芳"), ("P004", "客服系统", "刘洋"),
        ("P005", "ERP升级", "陈明"), ("P006", "小程序", "杨柳"),
        ("P007", "安全审计", "赵磊"), ("P008", "BI看板", "黄丽"),
    ]
    statuses = ["已完成", "进行中", "延期", "已完成", "延期", "进行中", "已完成", "未开始"]
    for i, (pid, pname, owner) in enumerate(projects):
        ps = datetime.date(2025, 1 + i, 1)
        pe = datetime.date(2025, 3 + i, 28)
        ast = ps + datetime.timedelta(days=random.randint(0, 10))
        if statuses[i] == "未开始":
            ast = None
            aen = None
        elif statuses[i] in ("已完成",):
            aen = pe + datetime.timedelta(days=random.randint(-5, 15))
        else:
            aen = None
        budget = round(random.uniform(10, 100), 1)
        ws.append([pid, pname, owner, ps, pe, ast, aen, statuses[i], budget])
    _save(wb, "05_项目进度跟踪.xlsx")


# ── 6. 家庭收支记账 ──────────────────────────────────────────
def gen_06():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收支明细"
    ws.append(["日期", "类型", "分类", "金额(元)", "备注"])
    income_cats = ["工资", "奖金", "兼职", "理财收益"]
    expense_cats = ["餐饮", "交通", "房租", "购物", "娱乐", "水电", "通讯"]
    base = datetime.date(2025, 1, 1)
    for day_offset in range(0, 180, 2):
        d = base + datetime.timedelta(days=day_offset)
        if day_offset % 30 == 0:
            ws.append([d, "收入", "工资", 15000, "月薪"])
        if random.random() > 0.3:
            cat = random.choice(expense_cats)
            amount = round(random.uniform(10, 800), 0)
            ws.append([d, "支出", cat, amount, ""])
    _save(wb, "06_家庭收支记账.xlsx")


# ── 7. 客户订单表（两个 sheet）──────────────────────────────
def gen_07():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "客户信息"
    ws1.append(["客户ID", "客户名称", "行业", "城市", "注册日期"])
    customers = [
        ("C001", "蓝天科技", "互联网", "北京"),
        ("C002", "绿叶食品", "食品", "上海"),
        ("C003", "红星建材", "建材", "广州"),
        ("C004", "白云物流", "物流", "深圳"),
        ("C005", "金桥教育", "教育", "杭州"),
        ("C006", "银河传媒", "传媒", "成都"),
    ]
    for cid, name, ind, city in customers:
        reg = datetime.date(2022, random.randint(1, 12), random.randint(1, 28))
        ws1.append([cid, name, ind, city, reg])

    ws2 = wb.create_sheet("订单明细")
    ws2.append(["订单号", "客户ID", "下单日期", "产品", "数量", "单价(元)", "状态"])
    products = ["服务器", "交换机", "显示器", "笔记本", "打印机"]
    order_statuses = ["已完成", "已发货", "待付款", "已取消"]
    for i in range(1, 26):
        oid = f"ORD-{i:04d}"
        cid = random.choice([c[0] for c in customers])
        odate = datetime.date(2025, random.randint(1, 6), random.randint(1, 28))
        prod = random.choice(products)
        qty = random.randint(1, 20)
        price = round(random.uniform(500, 20000), 0)
        status = random.choice(order_statuses)
        ws2.append([oid, cid, odate, prod, qty, price, status])
    _save(wb, "07_客户与订单.xlsx")


# ── 8. 电商评价数据 ──────────────────────────────────────────
def gen_08():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品评价"
    ws.append(["评价ID", "商品名称", "评分", "评价内容", "评价日期", "是否追评"])
    goods = ["无线鼠标", "机械键盘", "显示器支架", "USB扩展坞", "降噪耳机"]
    comments_good = ["很好用", "性价比高", "质量不错", "物流很快", "推荐购买"]
    comments_bad = ["质量一般", "有点贵", "包装破损", "与描述不符", "客服态度差"]
    for i in range(1, 31):
        good = random.choice(goods)
        score = random.randint(1, 5)
        comment = random.choice(comments_good if score >= 4 else comments_bad)
        d = datetime.date(2025, random.randint(1, 6), random.randint(1, 28))
        follow_up = "是" if random.random() > 0.7 else "否"
        ws.append([f"R{i:04d}", good, score, comment, d, follow_up])
    _save(wb, "08_电商评价数据.xlsx")


# ── 9. 天气与销售关联 ──────────────────────────────────────────
def gen_09():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日数据"
    ws.append(["日期", "最高温(℃)", "最低温(℃)", "天气", "销售额(元)", "客流量"])
    weathers = ["晴", "多云", "阴", "小雨", "大雨", "雪"]
    base = datetime.date(2025, 1, 1)
    for i in range(90):
        d = base + datetime.timedelta(days=i)
        month = d.month
        high = random.randint(0 + month * 5, 10 + month * 5)
        low = high - random.randint(5, 12)
        weather = random.choice(weathers)
        # 天气差时销售额低
        base_sales = random.randint(8000, 25000)
        if weather in ("大雨", "雪"):
            base_sales = int(base_sales * 0.6)
        traffic = base_sales // random.randint(30, 60)
        ws.append([d, high, low, weather, base_sales, traffic])
    _save(wb, "09_天气与销售.xlsx")


# ── 10. 部门费用预算 vs 实际 ────────────────────────────────
def gen_10():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "费用对比"
    ws.append(["部门", "费用科目", "预算(万元)", "实际(万元)", "季度"])
    depts = ["研发部", "市场部", "行政部", "销售部"]
    subjects = ["差旅费", "办公费", "培训费", "招待费", "设备费"]
    for dept in depts:
        for subj in subjects:
            for q in ["Q1", "Q2"]:
                budget = round(random.uniform(5, 50), 1)
                actual = round(budget * random.uniform(0.6, 1.4), 1)
                ws.append([dept, subj, budget, actual, q])
    _save(wb, "10_部门费用预算.xlsx")


# ── 11. 招聘漏斗 ──────────────────────────────────────────
def gen_11():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "招聘数据"
    ws.append(["岗位", "投递数", "筛选通过", "一面通过", "二面通过", "发offer", "入职", "招聘渠道"])
    positions = ["Java开发", "前端开发", "产品经理", "UI设计师",
                 "数据分析师", "运维工程师", "测试工程师", "项目经理"]
    channels = ["Boss直聘", "拉勾", "猎聘", "内推", "校招"]
    for pos in positions:
        total = random.randint(50, 300)
        screen = int(total * random.uniform(0.3, 0.6))
        r1 = int(screen * random.uniform(0.3, 0.6))
        r2 = int(r1 * random.uniform(0.3, 0.7))
        offer = int(r2 * random.uniform(0.5, 1.0))
        onboard = int(offer * random.uniform(0.5, 1.0))
        ch = random.choice(channels)
        ws.append([pos, total, screen, r1, r2, offer, onboard, ch])
    _save(wb, "11_招聘漏斗.xlsx")


# ── 12. 设备巡检记录 ──────────────────────────────────────────
def gen_12():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "巡检记录"
    ws.append(["设备编号", "设备名称", "巡检日期", "巡检人", "运行状态",
               "温度(℃)", "振动值(mm/s)", "是否异常", "处理措施"])
    devices = [
        ("D001", "空压机A"), ("D002", "空压机B"), ("D003", "冷却塔"),
        ("D004", "变压器"), ("D005", "发电机"),
    ]
    inspectors = ["王工", "李工", "张工"]
    base = datetime.date(2025, 1, 1)
    for week in range(12):
        d = base + datetime.timedelta(weeks=week)
        for did, dname in devices:
            inspector = random.choice(inspectors)
            temp = round(random.uniform(35, 85), 1)
            vibration = round(random.uniform(0.5, 8.0), 2)
            abnormal = "是" if temp > 75 or vibration > 6.0 else "否"
            action = "已处理" if abnormal == "是" and random.random() > 0.3 else ""
            status = "异常" if abnormal == "是" else "正常"
            ws.append([did, dname, d, inspector, status, temp, vibration, abnormal, action])
    _save(wb, "12_设备巡检记录.xlsx")


if __name__ == "__main__":
    print("生成简单测试数据集...")
    gen_01()
    gen_02()
    gen_03()
    gen_04()
    gen_05()
    gen_06()
    gen_07()
    gen_08()
    gen_09()
    gen_10()
    gen_11()
    gen_12()
    print(f"\n共 12 个数据集，保存在 {OUTPUT_DIR}/")
