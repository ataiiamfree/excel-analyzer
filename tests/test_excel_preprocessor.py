from pathlib import Path

import openpyxl
import pandas as pd

from app.tools.excel_preprocessor import ExcelPreprocessor
from app.tools.profiler import Profiler
from app.tools.workbook_ingestor import WorkbookIngestor


def test_summary_keyword_inside_business_value_is_not_auto_excluded():
    rows = [
        ["类型", "金额"],
        ["合计包采购", 10],
        ["普通采购", 20],
        ["其他采购", 30],
        ["合计", 60],
    ]

    flags = ExcelPreprocessor().classify_rows(rows, header_row=1, data_end=5)

    assert flags[2]["exclude"] is False
    assert flags[2]["kind"] == "possible_summary"
    assert flags[5]["exclude"] is True
    assert flags[5]["kind"] == "summary"


def test_manifest_path_accepts_new_schema_name():
    manifest = {"manifest_path": "workbook_manifest.json", "files": []}
    assert ExcelPreprocessor().normalize_manifest_path(manifest) == "workbook_manifest.json"


def test_manifest_path_fallback():
    manifest = {"path": "old_path.json"}
    assert ExcelPreprocessor().normalize_manifest_path(manifest) == "old_path.json"


def test_manifest_path_none():
    assert ExcelPreprocessor().normalize_manifest_path({}) is None


def test_dedupe_headers():
    preprocessor = ExcelPreprocessor()
    headers = preprocessor._dedupe_headers(["名称", "金额", "金额", None, "金额"])
    assert headers == ["名称", "金额", "金额_2", "列4", "金额_3"]


def test_merge_multi_level_headers():
    preprocessor = ExcelPreprocessor()
    rows = [
        ["采购金额", "采购金额", "销售金额", "销售金额"],
        ["计划", "实际", "计划", "实际"],
        [100, 110, 200, 190],
    ]
    preprocessor._merge_multi_level_headers(rows, header_row=2)
    assert rows[1] == ["采购金额_计划", "采购金额_实际", "销售金额_计划", "销售金额_实际"]


def test_merge_multi_level_headers_dedup():
    """合并单元格填充后上下层相同时应去重。"""
    preprocessor = ExcelPreprocessor()
    rows = [
        ["总额", "总额"],
        ["总额", "总额"],  # 与上面完全相同（合并填充后的结果）
        [100, 200],
    ]
    preprocessor._merge_multi_level_headers(rows, header_row=2)
    assert rows[1] == ["总额", "总额"]  # 不会变成 "总额_总额"


def test_detect_data_end_with_footnote():
    preprocessor = ExcelPreprocessor()
    rows = [
        ["名称", "金额"],
        ["A", 100],
        ["B", 200],
        ["备注：仅供参考", None],
        [None, None],
    ]
    end = preprocessor._detect_data_end(rows, header_row=1)
    assert end == 3  # 第3行（B, 200）是最后的有效数据行


def test_detect_data_end_no_footnote():
    preprocessor = ExcelPreprocessor()
    rows = [
        ["名称", "金额"],
        ["A", 100],
        ["B", 200],
    ]
    end = preprocessor._detect_data_end(rows, header_row=1)
    assert end == 3


def test_classify_rows_blank_and_title():
    preprocessor = ExcelPreprocessor()
    rows = [
        ["采购报表 2024年", None],  # title (before header)
        [None, None],               # also title (before header, even though blank)
        ["名称", "金额"],           # header (row 3)
        [None, None],               # blank (after header)
        ["A", 100],
    ]
    flags = preprocessor.classify_rows(rows, header_row=3, data_end=5)
    assert flags[1]["kind"] == "title"
    assert flags[1]["exclude"] is True
    assert flags[2]["kind"] == "title"   # before header → title
    assert flags[2]["exclude"] is True
    assert flags[4]["kind"] == "blank"   # after header, empty → blank
    assert flags[4]["exclude"] is True
    assert flags[5]["kind"] == "data"
    assert flags[5]["exclude"] is False


def test_formula_without_cached_value_warns(tmp_path):
    raw_dir = tmp_path / "raw"
    raw_dir.mkdir()
    workbook_path = raw_dir / "formula.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["A", "B", "Total"])
    ws.append([1, 2, "=A2+B2"])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:C2",
                                "header_candidates": [1],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)

    assert result.tables[0].row_count == 1
    assert any("公式没有缓存计算值" in warning for warning in result.tables[0].warnings)


def test_process_writes_to_explicit_output_dir(tmp_path):
    workbook_path = tmp_path / "input.xlsx"
    output_dir = tmp_path / "custom_normalized"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["名称", "金额"])
    ws.append(["A", 10])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:B2",
                                "header_candidates": [1],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest, output_dir=output_dir)

    assert result.tables[0].row_count == 1
    assert output_dir in Path(result.tables[0].parquet_path).parents


def test_process_reuses_previous_headers_for_headerless_continuation_block(tmp_path):
    workbook_path = tmp_path / "continuation.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Division", "Subdivision", "Serial No."])
    ws.append(["Foundation", "Earthwork", 1])
    ws.append([None, None, 2])
    ws.append([None, None, None])
    ws.append([None, None, None])
    ws.append([None, None, None])
    ws.append(["Decoration", "Floor", 1])
    ws.append([None, None, 2])
    ws.append([None, "Doors", 1])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:C3",
                                "header_candidates": [1],
                            },
                            {
                                "table_id": "Sheet1_t2",
                                "range": "A7:C9",
                                # Simulates a later text-heavy row being
                                # mistaken for a header candidate.
                                "header_candidates": [9],
                            },
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    second = result.tables[1]
    path = Path(second.parquet_path)
    dataframe = pd.read_parquet(path) if path.suffix == ".parquet" else pd.read_excel(path)

    assert list(dataframe.columns[:3]) == ["Division", "Subdivision", "Serial No."]
    assert dataframe.loc[0, "Division"] == "Decoration"
    assert dataframe.loc[0, "Serial No."] == 1
    assert any("续表" in warning for warning in second.warnings)


def test_process_converts_group_title_rows_to_context_column(tmp_path):
    workbook_path = tmp_path / "grouped.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Serial No.", "Project Name", "Unit", "Quantity", "Remarks"])
    ws.append(["First Floor", None, None, None, None])
    ws.merge_cells("A2:D2")
    ws.append([1, "Floor Tiles (1000×1000)", "㎡", 72.4, "ok"])
    ws.append([2, "Kitchen Wall Tiles", "㎡", 60, "ok"])
    ws.append(["Second Floor", None, None, None, "XXXXX"])
    ws.append([1, "Floor Tiles (1000×1000)", "㎡", 120, "ok"])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:E6",
                                "header_candidates": [1],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]
    path = Path(table.parquet_path)
    dataframe = pd.read_parquet(path) if path.suffix == ".parquet" else pd.read_excel(path)

    assert table.row_count == 3
    assert "_context_group" in dataframe.columns
    assert dataframe["Project Name"].tolist() == [
        "Floor Tiles (1000×1000)",
        "Kitchen Wall Tiles",
        "Floor Tiles (1000×1000)",
    ]
    assert dataframe["_context_group"].tolist() == [
        "First Floor",
        "First Floor",
        "Second Floor",
    ]
    assert table.enum_columns["_context_group"] == ["First Floor", "Second Floor"]
    assert any("_context_group" in warning for warning in table.warnings)


def test_context_group_ignores_notes_column_header(tmp_path):
    workbook_path = tmp_path / "notes_header.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Notes", "Project Name", "Quantity"])
    ws.append(["First Floor", None, None])
    ws.append(["ok", "Floor Tiles", 72.4])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:C3",
                                "header_candidates": [1],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]
    path = Path(table.parquet_path)
    dataframe = pd.read_parquet(path) if path.suffix == ".parquet" else pd.read_excel(path)

    assert "_context_group" not in dataframe.columns
    assert dataframe["Notes"].tolist()[0] == "First Floor"


def test_context_group_requires_downstream_detail_row(tmp_path):
    workbook_path = tmp_path / "orphan_group.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Serial No.", "Project Name", "Quantity"])
    ws.append(["First Floor", None, None])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:C2",
                                "header_candidates": [1],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]
    path = Path(table.parquet_path)
    dataframe = pd.read_parquet(path) if path.suffix == ".parquet" else pd.read_excel(path)

    assert table.row_count == 1
    assert "_context_group" not in dataframe.columns
    assert dataframe.loc[0, "Serial No."] == "First Floor"


def test_context_group_still_detects_first_floor_pattern():
    preprocessor = ExcelPreprocessor()

    label = preprocessor._context_group_label(
        ["First Floor", None, None],
        ["Serial No.", "Project Name", "Quantity"],
        following_rows=[[1, "Floor Tiles", 72.4]],
    )

    assert label == "First Floor"


def test_process_forward_fills_repeating_key_value_context_rows(tmp_path):
    workbook_path = tmp_path / "repeating_entity_context.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Label", "Value", "Group Label", "Group", "Metric", "Amount"])
    ws.append(["User Name:", "Alice 1001", "Department:", "Support", None, None])
    ws.append(["Queue A", None, None, None, "Chats Serviced", 12])
    ws.append(["Total", None, None, None, "Chats Serviced", 12])
    ws.append(["User Name:", "Bob 1002", "Department:", "Sales", None, None])
    ws.append(["Queue B", None, None, None, "Chats Serviced", 8])
    ws.append(["Total", None, None, None, "Chats Serviced", 8])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:F7",
                                "header_candidates": [1],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]
    path = Path(table.parquet_path)
    dataframe = pd.read_parquet(path) if path.suffix == ".parquet" else pd.read_excel(path)

    assert table.row_count == 4
    assert dataframe["_context_user_name"].tolist() == [
        "Alice 1001",
        "Alice 1001",
        "Bob 1002",
        "Bob 1002",
    ]
    assert dataframe["_context_department"].tolist() == [
        "Support",
        "Support",
        "Sales",
        "Sales",
    ]
    assert any("重复键值分组行" in warning for warning in table.warnings)


def test_repeating_colon_values_with_numeric_data_remain_flat_rows(tmp_path):
    workbook_path = tmp_path / "flat_colon_values.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Event", "Value", "Notes"])
    ws.append(["Status:", 10, "ok"])
    ws.append(["Status:", 20, "ok"])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:C3",
                                "header_candidates": [1],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]
    path = Path(table.parquet_path)
    dataframe = pd.read_parquet(path) if path.suffix == ".parquet" else pd.read_excel(path)

    assert table.row_count == 2
    assert dataframe["Event"].tolist() == ["Status:", "Status:"]
    assert not any(column.startswith("_context_") for column in dataframe.columns)


def test_merge_multi_level_headers_returns_lineage():
    preprocessor = ExcelPreprocessor()
    rows = [
        ["采购金额", "采购金额", "销售金额", "销售金额"],
        ["计划", "实际", "计划", "实际"],
        [100, 110, 200, 190],
    ]

    paths = preprocessor._merge_multi_level_headers(rows, header_row=2)

    assert paths == [
        ["采购金额", "计划"],
        ["采购金额", "实际"],
        ["销售金额", "计划"],
        ["销售金额", "实际"],
    ]


def test_process_preserves_header_lineage_for_grouped_year_columns(tmp_path):
    """Sparse grouped headers (SheetBench 118 shape).

    Parent group `Total Fundraising / Grant / Total Income` spans a
    `£(000)` unit row above per-year leaf columns. Post-normalize we should
    see year columns keyed by their dedup names and a `header_paths` map
    that resolves them back to the parent group.
    """

    workbook_path = tmp_path / "grouped_years.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    # Row 1: parent group (spanned via merged cells)
    ws.append(["Charity", "Total Fundraising", None, "Grant", None])
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")
    # Row 2: unit row
    ws.append([None, "£(000)", "£(000)", "£(000)", "£(000)"])
    # Row 3: leaf year columns
    ws.append([None, "2008/09", "2009/10", "2008/09", "2009/10"])
    # Data rows
    ws.append(["Cancer Research UK", 42, 55, 12, 15])
    ws.append(["Oxfam", 30, 33, 20, 22])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:E5",
                                "header_candidates": [1, 2, 3],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]

    # `_merge_multi_level_headers` concatenates parent → leaf with underscores;
    # header_paths keeps the structured lineage side-by-side for downstream use.
    fund_col = "Total Fundraising_£(000)_2008/09"
    grant_col = "Grant_£(000)_2008/09"
    assert fund_col in table.header_paths
    assert grant_col in table.header_paths
    assert table.header_paths[fund_col] == ["Total Fundraising", "£(000)", "2008/09"]
    assert table.header_paths[grant_col] == ["Grant", "£(000)", "2008/09"]
    # Charity column has no lineage → single-element path equal to the name
    assert table.header_paths["Charity"] == ["Charity"]
    # header_path also shows up on column metadata for downstream consumers
    charity_meta = next(c for c in table.columns if c["name"] == "Charity")
    assert charity_meta["header_path"] == ["Charity"]


def test_process_preserves_header_lineage_for_region_year_columns(tmp_path):
    """Region × year multi-level headers (SheetBench 2292 shape).

    `Landings into` groups regions like `Scotland` and `England`, each
    with per-year sub-columns. Leaf column names collide, so lineage is
    the only reliable way to pick the right column.
    """

    workbook_path = tmp_path / "region_year.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    # Row 1: top-level group
    ws.append(["Species", "Landings into", None, None, None])
    ws.merge_cells("B1:E1")
    # Row 2: region
    ws.append([None, "Scotland", "Scotland", "England", "England"])
    ws.merge_cells("B2:C2")
    ws.merge_cells("D2:E2")
    # Row 3: year
    ws.append([None, "2022", "2023", "2022", "2023"])
    # Data
    ws.append(["Cod", 100, 110, 50, 55])
    ws.append(["Haddock", 80, 90, 40, 45])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:E5",
                                "header_candidates": [1, 2, 3],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]

    # Merged names carry parent+region+year; header_paths keeps the structured
    # lineage so the model can pick e.g. Scotland vs England without parsing
    # the underscored name.
    scotland_2023 = "Landings into_Scotland_2023"
    england_2023 = "Landings into_England_2023"
    assert table.header_paths[scotland_2023] == [
        "Landings into",
        "Scotland",
        "2023",
    ]
    assert table.header_paths[england_2023] == [
        "Landings into",
        "England",
        "2023",
    ]


def test_forward_fill_sparse_header_row_spans_gaps():
    preprocessor = ExcelPreprocessor()
    row = [None, "Landings into", None, None, None, "Total landings", None, None]

    filled = preprocessor._forward_fill_sparse_header_row(row, num_cols=8)

    assert filled == [
        None,
        "Landings into",
        "Landings into",
        "Landings into",
        "Landings into",
        "Total landings",
        "Total landings",
        "Total landings",
    ]


def test_forward_fill_leaves_dense_rows_untouched():
    preprocessor = ExcelPreprocessor()
    row = ["A", "B", "C", "D", "E"]

    filled = preprocessor._forward_fill_sparse_header_row(row, num_cols=5)

    assert filled == ["A", "B", "C", "D", "E"]


def test_merge_multi_level_headers_skips_penultimate_annotation_row():
    """4-row header: parent → region → %-marker annotation → leaf year.

    The %-marker row is per-cell annotation, not spanning. Forward-fill
    would contaminate non-% columns; the merger should leave that row alone.
    """

    preprocessor = ExcelPreprocessor()
    rows = [
        [None, "Landings into", None, None, "Total landings", None, None],
        [None, "Scotland", None, None, "by UK vessels", None, None],
        [None, None, None, "%", None, None, "%"],
        [None, 2022, 2023, "change", 2022, 2023, "change"],
        [None, 100, 110, -5.5, 200, 210, -3.2],
    ]

    paths = preprocessor._merge_multi_level_headers(rows, header_row=4)

    # Scotland's 2022 and 2023 stay clean (no % contamination)
    assert paths[1] == ["Landings into", "Scotland", "2022"]
    assert paths[2] == ["Landings into", "Scotland", "2023"]
    # Scotland's change column correctly carries the % annotation
    assert paths[3] == ["Landings into", "Scotland", "%", "change"]
    # by UK vessels' 2022/2023 also clean
    assert paths[4] == ["Total landings", "by UK vessels", "2022"]
    assert paths[5] == ["Total landings", "by UK vessels", "2023"]


def test_ingest_then_process_preserves_sparse_multi_level_lineage_end_to_end(tmp_path):
    """SheetBench 2292 shape: no merged cells, sparse group cells.

    `Landings into` sits alone at col B expected to span cols B-D; then
    `Scotland`/`England` at cols B and E expected to span 3 cols each.
    Without ingestor sparse-row extension + preprocessor forward-fill, the
    lineage never reaches the year columns and the model can't distinguish
    Scotland/2023 from England/2023.
    """

    from app.tools.workbook_ingestor import WorkbookIngestor

    workbook_path = tmp_path / "sheetbench_2292_shape.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"

    # Structure: parent group row (sparse), region row (sparse), annotation
    # row (very sparse), leaf year row (dense). Regions Scotland/England each
    # span 2 leaf year columns; UK carries a `%` annotation column.
    ws.append([None, "Landings into", None, None, None, "Total landings", None])
    ws.append([None, "Scotland", None, "England", None, "UK", None])
    ws.append([None, None, None, None, None, "%", None])
    ws.append(["Stock", 2022, 2023, 2022, 2023, "quota", "change"])
    ws.append(["NS Herring", 100, 110, 200, 210, 300, -3.2])
    ws.append(["WC Mackerel", 300, 320, 400, 420, 500, 5.0])
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]

    # Region + year lineage is correctly assembled for tonnage columns
    assert table.header_paths["Landings into_Scotland_2022"] == [
        "Landings into",
        "Scotland",
        "2022",
    ]
    assert table.header_paths["Landings into_Scotland_2023"] == [
        "Landings into",
        "Scotland",
        "2023",
    ]
    assert table.header_paths["Landings into_England_2022"] == [
        "Landings into",
        "England",
        "2022",
    ]
    assert table.header_paths["Landings into_England_2023"] == [
        "Landings into",
        "England",
        "2023",
    ]
    # Annotation row's `%` only reaches the column it was actually placed on
    # (UK quota column). Non-annotated year columns stay clean.
    assert table.header_paths["Total landings_UK_%_quota"] == [
        "Total landings",
        "UK",
        "%",
        "quota",
    ]
    assert "%" not in table.header_paths["Landings into_England_2023"]


def test_ingest_then_process_preserves_multi_level_lineage_end_to_end(tmp_path):
    """SheetBench 118 shape driven from ingestor, not a hand-crafted manifest.

    Guards against the failure mode where our preprocessor plumbing works but
    the ingestor's header detection only returns the leaf row, leaving the
    lineage map single-level.
    """

    from app.tools.workbook_ingestor import WorkbookIngestor

    workbook_path = tmp_path / "sheetbench_118_shape.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"

    ws.append([None, "Total Fundraising", None, None, "Grant", None, None])
    ws.merge_cells("B1:D1")
    ws.merge_cells("E1:G1")
    ws.append([None, "£(000)", None, None, "£(000)", None, None])
    ws.merge_cells("B2:D2")
    ws.merge_cells("E2:G2")
    ws.append([
        "Charity",
        "2008/09", "2009/10", "2010/11",
        "2008/09", "2009/10", "2010/11",
    ])
    ws.append(["British Museum", 100, 110, 120, 10, 12, 14])
    ws.append(["Oxfam", 80, 90, 100, 5, 6, 7])
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]

    fund_col = "Total Fundraising_£(000)_2008/09"
    grant_col = "Grant_£(000)_2008/09"
    assert fund_col in table.header_paths
    assert grant_col in table.header_paths
    assert table.header_paths[fund_col] == [
        "Total Fundraising",
        "£(000)",
        "2008/09",
    ]
    assert table.header_paths[grant_col] == ["Grant", "£(000)", "2008/09"]
    # And the metadata surfaces it downstream
    fund_meta = next(c for c in table.columns if c["name"] == fund_col)
    assert fund_meta["header_path"] == ["Total Fundraising", "£(000)", "2008/09"]


def test_ingest_then_process_retains_group_columns_after_numeric_preamble(tmp_path):
    from app.tools.workbook_ingestor import WorkbookIngestor

    workbook_path = tmp_path / "matrix_header.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Matrix"

    ws.append(["Quarterly operating matrix", None, None, None, None, None])
    ws.append([None, 301, 302, 303, 304, 305])
    ws.append([None, "North", None, "South", None, "TOTAL"])
    ws.merge_cells("B3:C3")
    ws.merge_cells("D3:E3")
    ws.merge_cells("F3:F6")
    ws.append([None, "=1+1", "=1+2", "=1+3", "=1+4", None])
    ws.append([None, "=10+1", "=10+2", "=10+3", "=10+4", None])
    ws.append(["Metric", "Retail", "Online", "Retail", "Online", None])
    ws.append(["Revenue", 10, 20, 30, 40, 100])
    ws.append(["Cost", 4, 8, 12, 16, 40])
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    table = ExcelPreprocessor().process(workbook_path, manifest).tables[0]
    normalized_path = Path(table.parquet_path)
    normalized = (
        pd.read_parquet(normalized_path)
        if normalized_path.suffix == ".parquet"
        else pd.read_excel(normalized_path)
    )

    assert "TOTAL" in normalized.columns
    revenue = normalized[
        normalized[normalized.columns[0]].astype(str).str.strip().eq("Revenue")
    ]
    assert len(revenue) == 1
    assert revenue.iloc[0]["TOTAL"] == 100
    assert table.header_paths["TOTAL"] == ["TOTAL"]


def test_process_populates_trivial_header_path_for_single_level_tables(tmp_path):
    """Single-level tables also get header_path, keyed to the column name."""

    workbook_path = tmp_path / "flat.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Name", "Amount"])
    ws.append(["A", 10])
    ws.append(["B", 20])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:B3",
                                "header_candidates": [1],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]
    assert table.header_paths["Name"] == ["Name"]
    assert table.header_paths["Amount"] == ["Amount"]
    for col_meta in table.columns:
        assert col_meta["header_path"] == [col_meta["name"]]


def test_process_distinguishes_excel_display_scaling_from_stored_thousands(tmp_path):
    workbook_path = tmp_path / "display_scale.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["Museum", "Full value", "Stored thousands"])
    ws.append(["A", 8_955_000, 8_955])
    ws.append(["B", 13_555_000, 13_555])
    for row in range(2, 4):
        ws.cell(row, 2).number_format = "#,##0,"
        ws.cell(row, 3).number_format = "#,##0"
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    table = ExcelPreprocessor().process(workbook_path, manifest).tables[0]
    metadata = {item["name"]: item for item in table.columns}

    assert metadata["Full value"]["excel_number_formats"] == ["#,##0,"]
    assert metadata["Full value"]["excel_display_divisor"] == 1000
    assert metadata["Stored thousands"]["excel_number_formats"] == ["#,##0"]
    assert "excel_display_divisor" not in metadata["Stored thousands"]


def test_excel_display_divisor_ignores_grouping_commas_and_literals():
    preprocessor = ExcelPreprocessor()

    assert preprocessor._excel_display_divisor("#,##0") == 1
    assert preprocessor._excel_display_divisor("#,##0,") == 1000
    assert preprocessor._excel_display_divisor('"£"#,##0,, "m"') == 1_000_000


def test_process_records_enum_and_oversized_metadata_without_truncating_data(tmp_path):
    workbook_path = tmp_path / "input.xlsx"
    long_text = "很长的说明" * 50
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.append(["状态", "说明"])
    ws.append(["已完成", long_text])
    ws.append(["未开始", "短说明"])
    ws.append(["已完成", "短说明2"])
    ws.append(["已完成", "短说明3"])
    ws.append(["已完成", "短说明4"])
    workbook.save(workbook_path)

    manifest = {
        "manifest_path": "workbook_manifest.json",
        "files": [
            {
                "path": str(workbook_path),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [
                            {
                                "table_id": "Sheet1_t1",
                                "range": "A1:B6",
                                "header_candidates": [1],
                            }
                        ],
                    }
                ],
            }
        ],
    }

    result = ExcelPreprocessor().process(workbook_path, manifest)
    table = result.tables[0]

    assert table.enum_columns == {"状态": ["已完成", "未开始"]}
    assert table.oversized_cells[0]["column"] == "说明"
    assert table.oversized_cells[0]["source_row"] == 2

    path = Path(table.parquet_path)
    dataframe = pd.read_parquet(path) if path.suffix == ".parquet" else pd.read_excel(path)
    assert dataframe.loc[0, "说明"] == long_text

    profile = Profiler().profile(result.tables)
    sample_value = profile["tables"][0]["sample_rows"][0]["说明"]
    assert "[TRUNCATED:" in sample_value
