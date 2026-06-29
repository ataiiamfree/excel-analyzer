import json
from pathlib import Path

import openpyxl

from scripts.benchmark_data import (
    build_sheetbench_manifest,
    build_spreadsheetbench_manifest,
    build_spreadsheetbench_v2_manifest,
)


def _save_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "value"
    ws["A2"] = 42
    wb.save(path)


def test_build_spreadsheetbench_manifest_pairs_init_and_golden(tmp_path):
    root = tmp_path / "extracted" / "spreadsheetbench_verified_400"
    case_dir = root / "spreadsheet" / "13-1"
    case_dir.mkdir(parents=True)
    (case_dir / "prompt.txt").write_text("Fill the workbook.", encoding="utf-8")
    _save_workbook(case_dir / "1_13-1_init.xlsx")
    _save_workbook(case_dir / "1_13-1_golden.xlsx")
    (root / "dataset.json").write_text(
        json.dumps([
            {
                "id": "13-1",
                "instruction": "Fallback prompt",
                "spreadsheet_path": "spreadsheet/13-1",
                "instruction_type": "Sheet-Level Manipulation",
                "answer_position": "A1:A2",
                "answer_sheet": "Sheet1",
            }
        ]),
        encoding="utf-8",
    )

    manifest_path = tmp_path / "manifest.json"
    build_spreadsheetbench_manifest(tmp_path / "extracted", manifest_path)
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["case_count"] == 1
    case = manifest["cases"][0]
    assert case["id"] == "spreadsheetbench-13-1-1_13-1"
    assert case["file"].endswith("1_13-1_init.xlsx")
    assert case["assertions"]["answer_workbook"]["path"].endswith("1_13-1_golden.xlsx")
    assert case["assertions"]["answer_workbook"]["ranges"] == "A1:A2"


def test_build_spreadsheetbench_manifest_tolerates_single_golden_name_typo(tmp_path):
    root = tmp_path / "extracted" / "spreadsheetbench_verified_400"
    case_dir = root / "spreadsheet" / "42930"
    case_dir.mkdir(parents=True)
    _save_workbook(case_dir / "1_42930_init.xlsx")
    _save_workbook(case_dir / "1_43930_golden.xlsx")
    (root / "dataset.json").write_text(
        json.dumps([
            {
                "id": 42930,
                "instruction": "Complete the workbook.",
                "spreadsheet_path": "spreadsheet/42930",
                "answer_position": "A1:A2",
            }
        ]),
        encoding="utf-8",
    )

    manifest_path = tmp_path / "manifest.json"
    build_spreadsheetbench_manifest(tmp_path / "extracted", manifest_path)
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["case_count"] == 1
    assert manifest["cases"][0]["assertions"]["answer_workbook"]["path"].endswith("1_43930_golden.xlsx")


def test_build_spreadsheetbench_v2_manifest_uses_dataset_paths(tmp_path):
    suite = tmp_path / "extracted" / "data_example_05_11" / "Template"
    _save_workbook(suite / "spreadsheet" / "02_cash_sweep" / "02_01_input.xlsx")
    _save_workbook(suite / "spreadsheet" / "02_cash_sweep" / "02_01_golden.xlsx")
    suite.mkdir(parents=True, exist_ok=True)
    (suite / "dataset.json").write_text(
        json.dumps([
            {
                "id": "02_01",
                "instruction": "Complete the financial model.",
                "spreadsheet_path": "spreadsheet/02_cash_sweep/02_01_input.xlsx",
                "golden_response_path": "spreadsheet/02_cash_sweep/02_01_golden.xlsx",
                "answer_position": "'Revenue_Analysis'!B2:F29",
            }
        ]),
        encoding="utf-8",
    )

    manifest_path = tmp_path / "manifest.json"
    build_spreadsheetbench_v2_manifest(tmp_path / "extracted", manifest_path)
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["case_count"] == 1
    case = manifest["cases"][0]
    assert case["id"] == "spreadsheetbench-v2-template-02_01"
    assert case["file"].endswith("02_01_input.xlsx")
    assert case["assertions"]["answer_workbook"]["path"].endswith("02_01_golden.xlsx")


def test_build_sheetbench_manifest_keeps_qa_and_expected_answer(tmp_path):
    root = tmp_path / "extracted"
    suite_dir = root / "sheetbench" / "complex table cases"
    workbook = root / "sheetbench" / "complex table cases" / "mimo_hitab_xlsx" / "case.xlsx"
    _save_workbook(workbook)
    suite_dir.mkdir(parents=True, exist_ok=True)
    (suite_dir / "complex_tables_mimo_hitab_fixed.json").write_text(
        json.dumps([
            {
                "File": "sheetbench/complex table cases/mimo_hitab_xlsx/case.xlsx",
                "Tags": ["complex table"],
                "QA": ["What is the total?", "42"],
                "Source": "MiMoTable",
                "ID": "33",
                "Type": "QA",
            }
        ]),
        encoding="utf-8",
    )
    (root / "sheetbench" / "complex table cases" / "complex_tables_realhit_fixed.json").write_text(
        "[]",
        encoding="utf-8",
    )

    manifest_path = tmp_path / "manifest.json"
    build_sheetbench_manifest(root, manifest_path, variant="complex-qa")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["case_count"] == 1
    case = manifest["cases"][0]
    assert case["id"] == "sheetbench-complex_mimo_hitab-33-1"
    assert case["file"].endswith("case.xlsx")
    assert "Final Answer:" in case["question"]
    assert case["assertions"]["expected_answer"]["value"] == "42"
    assert case["assertions"]["expected_answer"]["require_marked_answer"] is True
    assert "tag=complex table" in case["tests"]


def test_build_sheetbench_manifest_builds_manipulation_case(tmp_path):
    root = tmp_path / "extracted"
    suite_dir = root / "sheetbench" / "manipulation cases"
    input_wb = suite_dir / "case1" / "input.xlsx"
    answer_wb = suite_dir / "case1" / "answer.xlsx"
    _save_workbook(input_wb)
    _save_workbook(answer_wb)
    suite_dir.mkdir(parents=True, exist_ok=True)
    (suite_dir / "manipulation_cases_question.json").write_text(
        json.dumps([
            {
                "File": "sheetbench/manipulation cases/case1/input.xlsx",
                "Tags": ["large table"],
                "QA": ["Fill the answer.", "sheetbench/manipulation cases/case1/answer.xlsx"],
                "Source": "spreadsheetbench",
                "ID": 1,
                "Type": "manipulation",
            }
        ]),
        encoding="utf-8",
    )

    manifest_path = tmp_path / "manifest.json"
    build_sheetbench_manifest(root, manifest_path, variant="manipulation")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["case_count"] == 1
    case = manifest["cases"][0]
    assert case["assertions"]["answer_workbook"]["path"].endswith("answer.xlsx")
    assert "type=manipulation" in case["tests"]
