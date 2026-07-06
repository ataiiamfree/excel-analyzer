import openpyxl

from app.tools.workbook_ingestor import WorkbookIngestor


def test_scan_uses_autofilter_as_strong_header_signal(tmp_path):
    workbook_path = tmp_path / "autofilter.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "采购"
    ws.append(["2026 年采购台账", None, None])
    ws.append(["类别", "金额", "状态"])
    ws.append(["工程", 100, "已完成"])
    ws.append(["服务", 200, "未开始"])
    ws.auto_filter.ref = "A2:C4"
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    sheet = manifest["files"][0]["sheets"][0]
    table = sheet["tables"][0]

    assert sheet["auto_filter_ref"] == "A2:C4"
    assert table["header_candidates"] == [2]
    assert table["confidence"] == 0.95
    assert any("AutoFilter" in note for note in table["notes"])


def test_autofilter_header_does_not_absorb_text_heavy_data_rows(tmp_path):
    workbook_path = tmp_path / "autofilter_text_rows.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "过超去重"
    ws.append(["用户编号", "用户名称", "供电单位", "计量点编号", "类型"])
    ws.append(["0950000075750333", "陈斌", "龙华供电局", "300127910614703843", "过"])
    ws.append(["0947000065600055", "深圳市同乐股份合作公司新布分公司", "龙城供电分局", "300176473213656014", "过"])
    ws.auto_filter.ref = "A1:E3"
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    table = manifest["files"][0]["sheets"][0]["tables"][0]

    assert table["header_candidates"] == [1]


def test_scan_detects_multi_level_header_with_merged_parent_row(tmp_path):
    """SheetBench 118 shape: parent group + unit row + leaf year row.

    The parent group and unit rows use merged cells; without merge-aware
    density detection, they look sparse and get skipped, collapsing the
    three-level header into a single-row leaf.
    """

    workbook_path = tmp_path / "grouped_years.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"

    ws.append([None, "Total Fundraising", None, None, None, "Grant", None, None, None])
    ws.merge_cells("B1:E1")
    ws.merge_cells("F1:I1")
    ws.append([None, "£(000)", None, None, None, "£(000)", None, None, None])
    ws.merge_cells("B2:E2")
    ws.merge_cells("F2:I2")
    ws.append([
        "Charity",
        "2008/09", "2009/10", "2010/11", "2011/12",
        "2008/09", "2009/10", "2010/11", "2011/12",
    ])
    ws.append(["British Museum", 100, 110, 120, 130, 10, 12, 14, 16])
    ws.append(["Oxfam", 80, 90, 100, 110, 5, 6, 7, 8])
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    table = manifest["files"][0]["sheets"][0]["tables"][0]

    assert table["header_candidates"] == [1, 2, 3]


def test_scan_extends_backwards_through_sparse_parent_rows(tmp_path):
    """SheetBench 2292 shape: no merges, sparse parent-group rows.

    Rows 1-3 are individually below the 40% density threshold, but they sit
    consecutively above a valid leaf header (row 4). The ingestor should
    walk upward from the first primary candidate and include them.
    """

    workbook_path = tmp_path / "sparse_headers.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "PELAGIC"

    ws.append([None, "Landings into", None, None, None, "Total landings", None])
    ws.append([None, "Scotland", None, "England", None, "by UK vessels", None])
    ws.append([None, None, None, "%", None, None, "%"])
    ws.append(["Stock", 2022, 2023, "change", 2022, 2023, "change"])
    ws.append(["NS Herring", 100, 110, -5.5, 200, 210, -3.2])
    ws.append(["WC Mackerel", 300, 320, 6.7, 400, 420, 5.0])
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    table = manifest["files"][0]["sheets"][0]["tables"][0]

    assert table["header_candidates"] == [1, 2, 3, 4]


def test_scan_treats_year_like_ints_as_header_text(tmp_path):
    """Leaf year rows like `2022, 2023, 2024` should not be rejected as data."""

    workbook_path = tmp_path / "year_header.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"

    ws.append(["Region", 2022, 2023, 2024])
    ws.append(["Scotland", 100, 110, 120])
    ws.append(["England", 200, 210, 220])
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    table = manifest["files"][0]["sheets"][0]["tables"][0]

    assert table["header_candidates"] == [1]


def test_scan_ignores_single_value_title_row(tmp_path):
    """A report title merged across all columns should not become a header."""

    workbook_path = tmp_path / "title.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"

    ws.append(["Annual Report 2025", None, None, None])
    ws.merge_cells("A1:D1")
    ws.append([None, None, None, None])
    ws.append(["Item", "Category", "Amount", "Note"])
    ws.append(["A", "X", 100, "ok"])
    ws.append(["B", "Y", 200, "ok"])
    workbook.save(workbook_path)

    manifest = WorkbookIngestor().scan(workbook_path)
    table = manifest["files"][0]["sheets"][0]["tables"][0]

    # Row 1 title is rejected; the real single-level header at row 3 is picked.
    assert table["header_candidates"] == [3]
