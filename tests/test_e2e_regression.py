"""End-to-end regression tests with dirty Excel samples.

Each test builds a realistic .xlsx with openpyxl, then runs the full
Ingestor → Preprocessor → Profiler pipeline and asserts structural correctness.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from app.tools.workbook_ingestor import WorkbookIngestor
from app.tools.excel_preprocessor import ExcelPreprocessor
from app.tools.profiler import Profiler


def _run_pipeline(wb_path: Path) -> dict:
    """Run Ingestor → Preprocessor → Profiler and return results dict."""
    ingestor = WorkbookIngestor()
    preprocessor = ExcelPreprocessor()
    profiler = Profiler()

    out_dir = wb_path.parent / "normalized"
    out_dir.mkdir(exist_ok=True)

    manifest = ingestor.scan(wb_path)
    result = preprocessor.process(
        file_path=wb_path,
        manifest=manifest,
        output_dir=out_dir,
    )
    profile = profiler.profile(result.tables)

    return {
        "manifest": manifest,
        "tables": result.tables,
        "warnings": result.report.get("warnings", []),
        "profile": profile,
    }


# ── Sample 1: 多层表头 + 合并单元格 ──

def test_multi_level_header_with_merged_cells(tmp_path):
    """两层表头，上层合并跨列，下层具体字段名。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采购明细"
    # Row 1: merged headers
    ws.merge_cells("A1:B1")
    ws["A1"] = "基本信息"
    ws.merge_cells("C1:D1")
    ws["C1"] = "金额信息"
    # Row 2: sub-headers
    ws["A2"] = "供应商"
    ws["B2"] = "日期"
    ws["C2"] = "含税金额"
    ws["D2"] = "不含税金额"
    # Data rows
    for i in range(3, 13):
        ws[f"A{i}"] = f"供应商{i - 2}"
        ws[f"B{i}"] = f"2025-0{(i % 9) + 1}-15"
        ws[f"C{i}"] = (i - 2) * 10000
        ws[f"D{i}"] = (i - 2) * 8500

    path = tmp_path / "multi_header.xlsx"
    wb.save(path)

    result = _run_pipeline(path)
    assert len(result["tables"]) == 1
    table = result["tables"][0]
    # Headers should be merged: 基本信息_供应商, etc.
    col_names = [c["name"] for c in table.columns if not c["name"].startswith("_source")]
    assert any("供应商" in c for c in col_names)
    assert any("金额" in c for c in col_names)
    assert table.row_count == 10


# ── Sample 2: AutoFilter + 隐藏列 ──

def test_autofilter_with_hidden_column(tmp_path):
    """有 AutoFilter 的标准表，其中一列被隐藏。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单"
    headers = ["订单号", "客户", "金额", "内部编码", "状态"]
    ws.append(headers)
    for i in range(1, 21):
        ws.append([f"ORD-{i:04d}", f"客户{i % 5}", i * 150, f"INT-{i}", "已完成" if i % 3 else "进行中"])
    ws.auto_filter.ref = "A1:E21"
    # Hide column D (内部编码)
    ws.column_dimensions["D"].hidden = True

    path = tmp_path / "autofilter_hidden.xlsx"
    wb.save(path)

    result = _run_pipeline(path)
    manifest = result["manifest"]
    sheet = manifest["files"][0]["sheets"][0]
    assert sheet["auto_filter_ref"] == "A1:E21"
    assert "D" in sheet["hidden_cols"]
    assert len(result["tables"]) == 1
    assert result["tables"][0].row_count == 20


# ── Sample 3: 同一 sheet 多个表（空行分隔） ──

def test_multi_table_same_sheet(tmp_path):
    """同一个 sheet 上有两个表，中间隔 3 行空行。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"
    # Table 1: rows 1-6
    ws.append(["部门", "人数", "预算"])
    for dept in ["研发", "销售", "财务", "人事", "市场"]:
        ws.append([dept, 10, 50000])
    # Gap: rows 7-9 are empty (3 blank rows)
    ws.append([])
    ws.append([])
    ws.append([])
    # Table 2: rows 10-14
    ws.append(["月份", "收入", "支出"])
    for m in range(1, 5):
        ws.append([f"2025-{m:02d}", m * 100000, m * 80000])

    path = tmp_path / "multi_table.xlsx"
    wb.save(path)

    result = _run_pipeline(path)
    # Should detect 2 table regions
    manifest = result["manifest"]
    tables_in_manifest = manifest["files"][0]["sheets"][0]["tables"]
    assert len(tables_in_manifest) >= 2, f"Expected ≥2 tables, got {len(tables_in_manifest)}"
    # Both should produce normalized tables
    assert len(result["tables"]) >= 2


# ── Sample 4: 脚注 + 汇总行 ──

def test_footnote_and_summary_row(tmp_path):
    """表末有汇总行和脚注，应被正确识别/排除。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "费用"
    ws.append(["项目", "Q1", "Q2", "Q3", "Q4"])
    items = ["差旅", "办公", "培训", "招待", "其他"]
    for item in items:
        ws.append([item, 1000, 2000, 1500, 1800])
    # Summary row
    ws.append(["合计", 5000, 10000, 7500, 9000])
    # Footnote
    ws.append([])
    ws.append(["备注：以上数据均为预算数，实际请以财务系统为准。"])

    path = tmp_path / "footnote_summary.xlsx"
    wb.save(path)

    result = _run_pipeline(path)
    table = result["tables"][0]
    # 5 data rows, summary and footnote excluded
    assert table.row_count == 5, f"Expected 5 data rows, got {table.row_count}"


# ── Sample 5: 多 sheet 工作簿 ──

def test_multi_sheet_workbook(tmp_path):
    """3 个 sheet，每个有不同结构的表。"""
    wb = openpyxl.Workbook()
    # Sheet 1: 员工表
    ws1 = wb.active
    ws1.title = "员工"
    ws1.append(["姓名", "部门", "入职日期"])
    for i in range(5):
        ws1.append([f"员工{i}", "技术" if i % 2 else "产品", f"2024-0{i+1}-01"])
    # Sheet 2: 项目表
    ws2 = wb.create_sheet("项目")
    ws2.append(["项目名", "负责人", "预算", "进度"])
    for i in range(3):
        ws2.append([f"项目{i}", f"员工{i}", (i + 1) * 100000, f"{(i + 1) * 30}%"])
    # Sheet 3: 空 sheet
    wb.create_sheet("备用")

    path = tmp_path / "multi_sheet.xlsx"
    wb.save(path)

    result = _run_pipeline(path)
    manifest = result["manifest"]
    assert len(manifest["files"][0]["sheets"]) == 3
    # Should produce 2 tables (空 sheet 不产出)
    assert len(result["tables"]) == 2
    sheet_names = {t.source_sheet for t in result["tables"]}
    assert "员工" in sheet_names
    assert "项目" in sheet_names


# ── Sample 6: 合并单元格作为分组标签 ──

def test_merged_cells_as_group_labels(tmp_path):
    """左侧列有合并单元格用作分组标签（如"华东区"跨3行）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "区域销售"
    ws.append(["区域", "城市", "销售额"])
    # 华东区: rows 2-4
    ws["A2"] = "华东区"
    ws.merge_cells("A2:A4")
    ws["B2"], ws["C2"] = "上海", 500000
    ws["B3"], ws["C3"] = "杭州", 300000
    ws["B4"], ws["C4"] = "南京", 200000
    # 华北区: rows 5-6
    ws["A5"] = "华北区"
    ws.merge_cells("A5:A6")
    ws["B5"], ws["C5"] = "北京", 800000
    ws["B6"], ws["C6"] = "天津", 250000

    path = tmp_path / "merged_groups.xlsx"
    wb.save(path)

    result = _run_pipeline(path)
    table = result["tables"][0]
    # After unmerge-and-fill, all rows should have region populated
    assert table.row_count == 5
    # Check profile has the table
    assert len(result["profile"]["tables"]) == 1


# ── Sample 7: 标题行在第3行（前面有 logo/标题文字） ──

def test_title_rows_before_header(tmp_path):
    """前两行是报告标题和日期，第3行才是表头。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "月度报告"
    ws.append(["2025年第一季度采购报告"])
    ws.append(["制表日期：2025-04-01"])
    ws.append(["物料编号", "名称", "数量", "单价", "总价"])
    for i in range(1, 8):
        ws.append([f"MAT-{i:03d}", f"物料{i}", i * 10, 50.0, i * 500.0])

    path = tmp_path / "title_rows.xlsx"
    wb.save(path)

    result = _run_pipeline(path)
    table = result["tables"][0]
    col_names = [c["name"] for c in table.columns if not c["name"].startswith("_source")]
    assert "物料编号" in col_names
    assert "名称" in col_names
    assert table.row_count == 7


# ── Sample 8: 数据中包含汇总关键词但不是汇总行 ──

def test_summary_keyword_in_data_not_excluded(tmp_path):
    """数据列中包含"合计"等关键词，但行本身是正常数据，不应被排除。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目"
    ws.append(["项目名", "说明", "金额"])
    ws.append(["合计系统升级", "对现有系统进行合计功能升级", 50000])
    ws.append(["数据汇总平台", "建设汇总展示平台", 80000])
    ws.append(["普通项目A", "常规项目", 30000])
    ws.append(["普通项目B", "常规项目", 25000])

    path = tmp_path / "summary_keyword_data.xlsx"
    wb.save(path)

    result = _run_pipeline(path)
    table = result["tables"][0]
    # All 4 rows should be kept — "合计" in project name is business data
    assert table.row_count == 4, f"Expected 4 data rows, got {table.row_count}"
